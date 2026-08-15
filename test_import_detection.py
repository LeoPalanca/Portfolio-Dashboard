from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.portfolio_dashboard.imports import detect_statement_source, import_destination


class ImportDetectionTest(unittest.TestCase):
    def test_detects_csv_sources_from_headers(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            trade_republic = root / "export.csv"
            trade_republic.write_text("type,category,shares,amount,date\n", encoding="utf-8")
            revolut = root / "statement.csv"
            revolut.write_text(
                "Tipo,Data di completamento,Descrizione,Importo,State,Valuta\n",
                encoding="utf-8",
            )

            self.assertEqual(detect_statement_source(trade_republic), "trade_republic")
            self.assertEqual(detect_statement_source(revolut), "revolut")

    def test_detects_workbook_source_from_sheet_names(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "export.xlsx"
            workbook = Workbook()
            workbook.active.title = "Attività account"
            workbook.create_sheet("Dividendi")
            workbook.save(path)

            self.assertEqual(detect_statement_source(path), "etoro")

    def test_detects_fineco_current_account_headers_below_the_preamble(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "movements.xlsx"
            workbook = Workbook()
            workbook.active.title = "Movimenti"
            workbook.active.append(("Conto Corrente: 1234567",))
            workbook.active.append(())
            workbook.active.append(
                (
                    "Data_Operazione",
                    "Data_Valuta",
                    "Entrate",
                    "Uscite",
                    "Descrizione",
                    "Descrizione_Completa",
                    "Stato",
                )
            )
            workbook.save(path)

            self.assertEqual(detect_statement_source(path), "fineco")

    def test_destination_matches_existing_discovery_layout(self) -> None:
        destination = import_destination(Path("/private/source"), "fineco", "a" * 64, "export.xlsx")

        self.assertEqual(destination.parent, Path("/private/source/broker_exports/fineco"))
        self.assertIn("fineco-", destination.name)

    def test_manual_source_selection_rejects_the_wrong_format(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "trade-republic.xlsx"
            path.write_bytes(b"not needed because the extension is rejected first")

            with self.assertRaisesRegex(ValueError, "Trade Republic imports require CSV; this file is XLSX"):
                detect_statement_source(path, "trade_republic")


if __name__ == "__main__":
    unittest.main()
