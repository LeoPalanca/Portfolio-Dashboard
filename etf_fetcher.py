from __future__ import annotations

import argparse
import csv
import io
import json
import re
import shutil
import sys
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from html import unescape
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urljoin, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook

import app as dashboard


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DOCUMENTS_JSON = DATA_DIR / "etf_documents.json"
HOLDINGS_DIR = DATA_DIR / "etf_holdings"
CLASSIFICATION_CSV = DATA_DIR / "holding_classifications.csv"
EXPOSURES_CSV = APP_DIR / "asset_exposures.csv"

EXPOSURE_FIELDS = [
    "asset_name",
    "isin",
    "holding_name",
    "holding_ticker",
    "weight_pct",
    "sector",
    "geo",
    "asset_class",
]
RAW_HOLDING_FIELDS = EXPOSURE_FIELDS + ["holding_isin", "source_url", "fetched_at"]

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Safari/537.36"
)

MONEY_MARKET_ISINS = {"LU0290358497"}

PRODUCT_URL_BY_ISIN = {
    "IE00B4L5Y983": "https://www.ishares.com/uk/individual/en/products/251882/ishares-core-msci-world-ucits-etf?switchLocale=y&siteEntryPassthrough=true",
    "IE00023EZQ82": "https://www.ishares.com/uk/individual/en/products/326051/ishares-digital-entertainment-and-education-ucits-etf?siteEntryPassthrough=true&switchLocale=y",
    "IE00B5BMR087": "https://www.ishares.com/uk/individual/en/products/253743/ishares-core-sp-500-ucits-etf-de-fund?switchLocale=y&siteEntryPassthrough=true",
    "IE00BKM4GZ66": "https://www.ishares.com/uk/individual/en/products/264659/ishares-core-msci-em-imi-ucits-etf?switchLocale=y&siteEntryPassthrough=true",
    "IE00BK5BQT80": "https://www.vanguard.co.uk/professional/product/etf/equity/9679/ftse-all-world-ucits-etf-usd-accumulating",
    "IE00B5MTXJ97": "https://www.invesco.com/ie/en/financial-products/etfs/invesco-stoxx-europe-600-optimised-insurance-ucits-etf-acc.html",
    "IE00BGV5VN51": "https://etf.dws.com/en-gb/IE00BGV5VN51-artificial-intelligence-big-data-ucits-etf-1c/",
    "IE00BM67HL84": "https://etf.dws.com/en-gb/IE00BM67HL84-msci-world-financials-ucits-etf-1c/",
    "IE000YDOORK7": "https://etf.dws.com/en-gb/IE000YDOORK7-msci-fintech-innovation-ucits-etf-1c/",
}

DIRECT_HOLDINGS_URL_BY_ISIN = {
    "IE00BK5BQT80": "https://www.vanguard.co.uk/professional/product/etf/equity/9679/ftse-all-world-ucits-etf-usd-accumulating",
    "IE00B5MTXJ97": "https://dng-api.invesco.com/cache/v1/accounts/en_IE/shareclasses/IE00B5MTXJ97/holdings/index?idType=isin",
    "IE00BGV5VN51": "https://etf.dws.com/api/pdp/en-gb/etf/IE00BGV5VN51/holdings",
    "IE00BM67HL84": "https://etf.dws.com/api/pdp/en-gb/etf/IE00BM67HL84/holdings",
    "IE000YDOORK7": "https://etf.dws.com/api/pdp/en-gb/etf/IE000YDOORK7/holdings",
}

STATIC_DOCUMENTS_BY_ISIN = {
    "LU1681045370": {
        "factsheet_url": "https://www.amundietf.com/pdfDocuments/monthly-factsheet/LU1681045370/ENG/FIN/INSTITUTIONNEL/ETF",
        "kid_url": "https://www.amundietf.com/pdfDocuments/kid-priips/LU1681045370/ENG/LUX/20251205",
    },
    "NL0011683594": {
        "product_url": "https://www.vaneck.com/uk/en/investments/dividend-etf",
        "factsheet_url": "https://www.vaneck.com/ucits/en/library/fact-sheets/tdiv-fact-sheet.pdf",
    },
}

ISSUER_BY_ISIN = {
    "IE00BK5BQT80": "Vanguard",
    "IE00B4L5Y983": "iShares",
    "IE00023EZQ82": "iShares",
    "IE00B5BMR087": "iShares",
    "IE00BKM4GZ66": "iShares",
    "IE00B5MTXJ97": "Invesco",
    "IE00BGV5VN51": "Xtrackers",
    "IE00BM67HL84": "Xtrackers",
    "NL0011683594": "VanEck",
    "LU1681045370": "Amundi",
    "LU1900066033": "Amundi",
    "IE000YDOORK7": "Xtrackers",
}

COUNTRY_BY_ISIN_PREFIX = {
    "AU": "Australia",
    "BE": "Belgium",
    "CA": "Canada",
    "CH": "Switzerland",
    "DE": "Germany",
    "DK": "Denmark",
    "ES": "Spain",
    "FI": "Finland",
    "FR": "France",
    "GB": "United Kingdom",
    "HK": "Hong Kong",
    "IE": "Ireland",
    "IT": "Italy",
    "JP": "Japan",
    "NL": "Netherlands",
    "NO": "Norway",
    "SE": "Sweden",
    "SG": "Singapore",
    "US": "United States",
}

COMPOSITION_HINTS = (
    " etf",
    " ucits",
    " acc",
    " msci ",
    " ftse ",
    " stoxx",
    " nasdaq",
    " overnight",
    " swap",
    " lev ",
    " leveraged",
    " cocoa",
    " coffee",
    " energy",
    " emerging markets",
    " semiconductors",
    " financials",
)


@dataclass(frozen=True)
class Position:
    asset_name: str
    isin: str


@dataclass
class Holding:
    asset_name: str
    isin: str
    holding_name: str
    holding_ticker: str
    weight_pct: Decimal
    sector: str
    geo: str
    asset_class: str
    holding_isin: str = ""
    source_url: str = ""
    fetched_at: str = ""


@dataclass
class DocumentRecord:
    isin: str
    asset_name: str
    issuer: str
    status: str
    fetched_at: str
    official_only: bool = True
    product_url: str = ""
    holdings_url: str = ""
    kid_url: str = ""
    factsheet_url: str = ""
    prospectus_url: str = ""
    parser: str = ""
    message: str = ""
    rows: int = 0
    weight_sum: str = ""
    normalized: bool = False
    data_path: str = ""


@dataclass(frozen=True)
class ProviderAdapter:
    issuer: str
    domains: tuple[str, ...]
    search_urls: tuple[str, ...]

    def matches(self, position: Position) -> bool:
        issuer = ISSUER_BY_ISIN.get(position.isin)
        if issuer:
            return issuer == self.issuer
        return self.issuer.casefold() in position.asset_name.casefold()

    def discover(self, position: Position, fetcher: "OfficialFetcher") -> dict[str, str]:
        pages: list[str] = []
        seeded_product_url = PRODUCT_URL_BY_ISIN.get(position.isin, "")
        if seeded_product_url and host_allowed(seeded_product_url, self.domains):
            pages.append(seeded_product_url)
        for template in self.search_urls:
            url = template.format(isin=quote(position.isin), asset=quote(position.asset_name))
            try:
                body, final_url = fetcher.fetch_text(url)
            except FetchError:
                continue
            if position.isin in body or self.issuer.casefold() in body.casefold():
                pages.append(final_url)
                pages.extend(find_candidate_pages(body, final_url, self.domains, position.isin))
            if len(pages) >= 8:
                break

        seen: set[str] = set()
        documents: dict[str, str] = {}
        for page_url in pages:
            if page_url in seen:
                continue
            seen.add(page_url)
            try:
                body, final_url = fetcher.fetch_text(page_url)
            except FetchError:
                continue
            found = find_document_links(body, final_url, self.domains)
            if not documents.get("product_url") and position.isin in body:
                documents["product_url"] = final_url
            for key, value in found.items():
                documents.setdefault(key, value)
            if documents.get("holdings_url") and documents.get("kid_url"):
                break
        return documents


PROVIDERS = [
    ProviderAdapter(
        "iShares",
        ("ishares.com", "blackrock.com"),
        (
            "https://www.ishares.com/uk/individual/en/search?switchLocale=y&siteEntryPassthrough=true&search={isin}",
            "https://www.ishares.com/uk/individual/en/search?search={isin}",
            "https://www.ishares.com/it/investitore-privato/it/ricerca?search={isin}",
        ),
    ),
    ProviderAdapter(
        "Vanguard",
        ("vanguard.co.uk", "vanguardinvestor.co.uk", "vanguard.com"),
        (
            "https://www.vanguardinvestor.co.uk/search?query={isin}",
            "https://www.vanguard.co.uk/professional/search?query={isin}",
        ),
    ),
    ProviderAdapter(
        "VanEck",
        ("vaneck.com",),
        (
            "https://www.vaneck.com/it/it/search/?query={isin}",
            "https://www.vaneck.com/uk/en/search/?query={isin}",
        ),
    ),
    ProviderAdapter(
        "Xtrackers",
        ("xtrackers.com", "dws.com"),
        (
            "https://etf.dws.com/en-gb/search-results/?query={isin}",
            "https://etf.dws.com/it-it/risultati-ricerca/?query={isin}",
        ),
    ),
    ProviderAdapter(
        "Invesco",
        ("invesco.com",),
        (
            "https://www.invesco.com/it/it/search.html?query={isin}",
            "https://www.invesco.com/uk/en/search.html?query={isin}",
        ),
    ),
    ProviderAdapter(
        "Amundi",
        ("amundietf.it", "amundietf.com", "amundi.com"),
        (
            "https://www.amundietf.it/it/professionali/ricerca?text={isin}",
            "https://www.amundietf.com/en/professional/search?text={isin}",
        ),
    ),
    ProviderAdapter(
        "WisdomTree",
        ("wisdomtree.eu",),
        (
            "https://www.wisdomtree.eu/it-it/search?query={isin}",
            "https://www.wisdomtree.eu/en-gb/search?query={isin}",
        ),
    ),
]


class FetchError(Exception):
    pass


class OfficialFetcher:
    def __init__(self, offline: bool = False, timeout: int = 25) -> None:
        self.offline = offline
        self.timeout = timeout

    def fetch_bytes(self, url: str) -> tuple[bytes, str, str]:
        if self.offline:
            raise FetchError("offline mode")
        request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
        request.add_header("Cookie", "iShares_userType=individual; siteEntryPassthrough=true")
        try:
            with urlopen(request, timeout=self.timeout) as response:
                content_type = response.headers.get("Content-Type", "")
                return response.read(), response.geturl(), content_type
        except (HTTPError, URLError, TimeoutError) as exc:
            raise FetchError(str(exc)) from exc

    def fetch_text(self, url: str) -> tuple[str, str]:
        payload, final_url, content_type = self.fetch_bytes(url)
        encoding = "utf-8"
        match = re.search(r"charset=([\w-]+)", content_type, flags=re.I)
        if match:
            encoding = match.group(1)
        try:
            return payload.decode(encoding), final_url
        except UnicodeDecodeError:
            return payload.decode("latin-1", errors="replace"), final_url


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_decimal(value: Any) -> Decimal:
    raw = clean(value)
    if not raw or raw in {"-", "—"}:
        return Decimal("0")
    raw = raw.replace("%", "").replace("\u00a0", "").replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    raw = re.sub(r"[^0-9.+-]", "", raw)
    try:
        return Decimal(raw)
    except InvalidOperation:
        return Decimal("0")


def decimal_str(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.0001")).normalize(), "f")


def host_allowed(url: str, domains: tuple[str, ...]) -> bool:
    host = urlparse(url).netloc.casefold()
    return any(host == domain or host.endswith("." + domain) for domain in domains)


def hrefs(html: str, base_url: str) -> list[tuple[str, str]]:
    matches = re.finditer(
        r"<a\b(?P<attrs>[^>]*)>(?P<label>.*?)</a>",
        html,
        flags=re.I | re.S,
    )
    links = []
    for match in matches:
        href_match = re.search(r"""href=["'](?P<href>[^"']+)["']""", match.group("attrs"), flags=re.I)
        if not href_match:
            continue
        label = clean(re.sub(r"<[^>]+>", " ", unescape(match.group("label"))))
        links.append((urljoin(base_url, unescape(href_match.group("href"))), label))
    return links


def find_candidate_pages(html: str, base_url: str, domains: tuple[str, ...], isin: str) -> list[str]:
    candidates: list[tuple[int, str]] = []
    for url, label in hrefs(html, base_url):
        if not host_allowed(url, domains):
            continue
        text = f"{url} {label}".casefold()
        score = 0
        if isin.casefold() in text:
            score += 20
        if any(word in text for word in ("fund", "etf", "product", "scheda", "prodotto")):
            score += 5
        if score:
            candidates.append((score, url))
    return [url for _, url in sorted(candidates, reverse=True)[:8]]


def find_document_links(html: str, base_url: str, domains: tuple[str, ...]) -> dict[str, str]:
    found: dict[str, tuple[int, str]] = {}
    raw_links = hrefs(html, base_url)

    for regex in (
        r"""["'](?P<url>[^"']+(?:csv|xlsx?|json|pdf)(?:\?[^"']*)?)["']""",
        r"""(?P<url>https?://[^"'\\\s]+(?:csv|xlsx?|json|pdf)(?:\?[^"'\\\s]*)?)""",
    ):
        for match in re.finditer(regex, html, flags=re.I):
            raw_links.append((urljoin(base_url, unescape(match.group("url"))), ""))

    for url, label in raw_links:
        if not host_allowed(url, domains):
            continue
        text = f"{url} {label}".casefold()
        key = ""
        score = 1
        is_download = any(token in text for token in (".csv", ".xls", ".xlsx", ".json", ".ajax", "filetype=", "download"))
        if any(word in text for word in ("holdings", "holding", "portfolio", "constituents", "positions", "composizione")):
            if is_download:
                key = "holdings_url"
                score = 50
        elif any(word in text for word in ("kid", "kiid", "kıd", "documents-clés", "kinder")):
            key = "kid_url"
            score = 30
        elif "factsheet" in text or "fact-sheet" in text or "scheda" in text:
            key = "factsheet_url"
            score = 20
        elif "prospectus" in text or "prospetto" in text:
            key = "prospectus_url"
            score = 10
        if key:
            old = found.get(key)
            if not old or score > old[0]:
                found[key] = (score, url)
    return {key: value for key, (_, value) in found.items()}


def detect_provider(position: Position) -> ProviderAdapter | None:
    for provider in PROVIDERS:
        if provider.matches(position):
            return provider
    return None


def load_existing_documents() -> dict[str, dict[str, Any]]:
    if not DOCUMENTS_JSON.exists():
        return {}
    try:
        payload = json.loads(DOCUMENTS_JSON.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if isinstance(payload, dict) and "funds" in payload:
        return {clean(item.get("isin")).upper(): item for item in payload.get("funds", []) if item.get("isin")}
    if isinstance(payload, dict):
        return {clean(key).upper(): value for key, value in payload.items() if isinstance(value, dict)}
    return {}


def save_documents(records: list[DocumentRecord]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at": now_iso(),
        "source_policy": "official_only",
        "funds": [asdict(record) for record in sorted(records, key=lambda item: item.isin)],
    }
    DOCUMENTS_JSON.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def current_open_positions(person: str = dashboard.PRIMARY_PORTFOLIO_ID) -> list[Position]:
    positions = []
    mappings = dashboard.read_mappings()
    
    if person.lower() in (dashboard.PRIMARY_PORTFOLIO_ID, "all"):
        trades, _ = dashboard.read_trades()
        summary = dashboard.summarize_trades(trades)
        positions.extend(summary["positions"])
        
    if person.lower() != dashboard.PRIMARY_PORTFOLIO_ID:
        if person.lower() == "all":
            persons = list(dashboard.FAMILY_PORTFOLIOS)
        else:
            persons = [person.lower()]
            
        for p in persons:
            try:
                snapshot = dashboard.read_family_snapshot(p)
                for pos in snapshot["positions"]:
                    if "is_open" not in pos:
                        pos["is_open"] = True
                positions.extend(snapshot["positions"])
            except Exception:
                pass
                
            try:
                tr_file = dashboard.latest_family_trade_republic_export(p)
                if tr_file:
                    tr_trades = dashboard.read_trade_republic_trades(tr_file)
                    tr_summary = dashboard.summarize_trades(tr_trades)
                    for pos in tr_summary["positions"]:
                        exists = False
                        for existing_p in positions:
                            if existing_p["asset"] == pos["asset"] or (existing_p.get("isin") and existing_p["isin"] == pos.get("isin")):
                                existing_p["is_open"] = existing_p.get("is_open", True) or pos.get("is_open", True)
                                if not existing_p.get("isin") and pos.get("isin"):
                                    existing_p["isin"] = pos["isin"]
                                exists = True
                                break
                        if not exists:
                            positions.append(pos)
            except Exception:
                pass

    output: list[Position] = []
    seen: set[str] = set()
    for item in positions:
        if not item.get("is_open", True):
            continue
        asset = clean(item.get("asset"))
        isin = clean(item.get("isin")).upper()
        if not isin:
            mapping = dashboard.mapping_for(asset, "", mappings)
            isin = clean(mapping.get("isin")).upper()
        key = isin or asset.casefold()
        if not key or key in seen:
            continue
        seen.add(key)
        output.append(Position(asset, isin))
    return sorted(output, key=lambda item: (item.asset_name.casefold(), item.isin))


def load_existing_exposure_rows() -> list[dict[str, str]]:
    if not EXPOSURES_CSV.exists():
        return []
    with EXPOSURES_CSV.open(newline="", encoding="utf-8-sig") as handle:
        return [{field: clean(row.get(field)) for field in EXPOSURE_FIELDS} for row in csv.DictReader(handle)]


def load_classifications() -> dict[str, dict[str, str]]:
    classifications: dict[str, dict[str, str]] = {}

    for row in load_existing_exposure_rows():
        keys = [row.get("holding_ticker", ""), row.get("holding_name", "")]
        for key in keys:
            if key:
                classifications.setdefault(
                    key.casefold(),
                    {
                        "sector": row.get("sector") or "Unknown from issuer data",
                        "geo": row.get("geo") or "Unknown from issuer data",
                        "asset_class": row.get("asset_class") or "ETF underlying",
                    },
                )

    if CLASSIFICATION_CSV.exists():
        with CLASSIFICATION_CSV.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                keys = [clean(row.get("holding_ticker")), clean(row.get("holding_name")), clean(row.get("isin")).upper()]
                value = {
                    "sector": clean(row.get("sector")) or "Unknown from issuer data",
                    "geo": clean(row.get("geo")) or "Unknown from issuer data",
                    "asset_class": clean(row.get("asset_class")) or "ETF underlying",
                }
                for key in keys:
                    if key:
                        classifications[key.casefold()] = value
    return classifications


def classify_holding(holding: Holding, classifications: dict[str, dict[str, str]]) -> Holding:
    if holding.sector and holding.geo:
        return holding
    for key in (holding.holding_ticker, holding.holding_isin, holding.holding_name):
        cached = classifications.get(key.casefold()) if key else None
        if cached:
            holding.sector = holding.sector or cached["sector"]
            holding.geo = holding.geo or cached["geo"]
            holding.asset_class = holding.asset_class or cached["asset_class"]
            return holding
    holding.sector = holding.sector or "Unknown from issuer data"
    holding.geo = holding.geo or "Unknown from issuer data"
    holding.asset_class = holding.asset_class or "ETF underlying"
    return holding


def apply_issuer_specific_classification(position: Position, holdings: list[Holding]) -> None:
    if position.isin == "IE00B5MTXJ97":
        for holding in holdings:
            if not holding.sector or holding.sector == "Unknown from issuer data":
                holding.sector = "Financials"
            if (not holding.geo or holding.geo == "Unknown from issuer data") and holding.holding_isin:
                holding.geo = COUNTRY_BY_ISIN_PREFIX.get(holding.holding_isin[:2], "")


def target_needs_composition(position: Position, existing_rows: list[dict[str, str]]) -> bool:
    if position.isin in MONEY_MARKET_ISINS:
        return True
    issuer = ISSUER_BY_ISIN.get(position.isin)
    if issuer:
        return True
    for row in existing_rows:
        if row.get("isin", "").upper() == position.isin and row.get("asset_class") != "Single share":
            return True
    text = f" {position.asset_name.casefold()} "
    return any(hint in text for hint in COMPOSITION_HINTS)


def is_money_market(position: Position) -> bool:
    text = position.asset_name.casefold()
    return position.isin in MONEY_MARKET_ISINS or ("overnight" in text and "swap" in text)


def direct_share_row(position: Position, existing_rows: list[dict[str, str]]) -> dict[str, str] | None:
    for row in existing_rows:
        if row.get("isin", "").upper() == position.isin and row.get("asset_class") == "Single share":
            return {field: row.get(field, "") for field in EXPOSURE_FIELDS}
    if position.isin and not target_needs_composition(position, existing_rows):
        return {
            "asset_name": position.asset_name,
            "isin": position.isin,
            "holding_name": position.asset_name,
            "holding_ticker": "",
            "weight_pct": "100",
            "sector": "Unknown from issuer data",
            "geo": "Unknown from issuer data",
            "asset_class": "Single share",
        }
    return None


def money_market_holding(position: Position, fetched_at: str) -> Holding:
    return Holding(
        asset_name=position.asset_name,
        isin=position.isin,
        holding_name="EUR overnight cash equivalent",
        holding_ticker="XEON" if position.isin == "LU0290358497" else "",
        weight_pct=Decimal("100"),
        sector="Cash / Money Market",
        geo="Eurozone",
        asset_class="Cash equivalent",
        source_url="cash-equivalent-classification",
        fetched_at=fetched_at,
    )


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", clean(value).casefold()).strip("_")


def find_column(headers: list[str], options: tuple[str, ...]) -> str:
    normalized = {normalize_header(header): header for header in headers}
    for option in options:
        option_norm = normalize_header(option)
        for norm, original in normalized.items():
            if norm == option_norm or option_norm in norm:
                return original
    return ""


def row_to_holding(row: dict[str, Any], position: Position, source_url: str, fetched_at: str) -> Holding | None:
    headers = list(row)
    name_col = find_column(headers, ("holding name", "security name", "name", "issuer name", "titolo", "descrizione"))
    weight_col = find_column(headers, ("% of fund", "% of market value", "weight", "weight (%)", "% net assets", "fund weight", "ponderazione", "peso"))
    if not name_col or not weight_col:
        return None
    name = clean(row.get(name_col))
    weight = parse_decimal(row.get(weight_col))
    if not name or weight <= 0:
        return None
    ticker_col = find_column(headers, ("ticker", "ticker symbol", "bbg ticker", "exchange ticker", "sedol"))
    isin_col = find_column(headers, ("isin", "isin code"))
    sector_col = find_column(headers, ("sector", "gics sector", "industry sector", "industry", "settore"))
    country_col = find_column(headers, ("country", "location", "market location", "country of risk", "paese", "geo"))
    asset_class_col = find_column(headers, ("asset class", "asset type", "type"))
    return Holding(
        asset_name=position.asset_name,
        isin=position.isin,
        holding_name=name,
        holding_ticker=clean(row.get(ticker_col)),
        weight_pct=weight,
        sector=clean(row.get(sector_col)),
        geo=clean(row.get(country_col)),
        asset_class=clean(row.get(asset_class_col)) or "ETF underlying",
        holding_isin=clean(row.get(isin_col)).upper(),
        source_url=source_url,
        fetched_at=fetched_at,
    )


def parse_csv_holdings(payload: bytes, position: Position, source_url: str, fetched_at: str) -> list[Holding]:
    text = payload.decode("utf-8-sig", errors="replace")
    if "\ufffd" in text:
        text = payload.decode("latin-1", errors="replace")
    lines = text.splitlines()
    best_rows: list[Holding] = []
    for start in range(min(40, len(lines))):
        sample = "\n".join(lines[start : start + 5])
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(lines[start:], dialect=dialect)
        if not reader.fieldnames:
            continue
        rows = [row_to_holding(row, position, source_url, fetched_at) for row in reader]
        holdings = [row for row in rows if row]
        if len(holdings) > len(best_rows):
            best_rows = holdings
        if len(best_rows) >= 5:
            break
    return best_rows


def parse_xlsx_holdings(payload: bytes, position: Position, source_url: str, fetched_at: str) -> list[Holding]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    best: list[Holding] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        for index, values in enumerate(rows[:80]):
            headers = [clean(value) for value in values]
            if not any(headers):
                continue
            data = []
            for row_values in rows[index + 1 :]:
                row = {headers[i]: row_values[i] if i < len(row_values) else "" for i in range(len(headers)) if headers[i]}
                holding = row_to_holding(row, position, source_url, fetched_at)
                if holding:
                    data.append(holding)
            if len(data) > len(best):
                best = data
            if len(best) >= 5:
                break
    return best


def iter_json_lists(value: Any) -> list[list[Any]]:
    lists: list[list[Any]] = []
    if isinstance(value, list):
        if value and all(isinstance(item, dict) for item in value):
            lists.append(value)
        for item in value:
            lists.extend(iter_json_lists(item))
    elif isinstance(value, dict):
        for item in value.values():
            lists.extend(iter_json_lists(item))
    return lists


def parse_json_holdings(payload: bytes, position: Position, source_url: str, fetched_at: str) -> list[Holding]:
    data = json.loads(payload.decode("utf-8-sig"))
    dws_holdings = parse_dws_json_holdings(data, position, source_url, fetched_at)
    if dws_holdings:
        return dws_holdings
    best: list[Holding] = []
    for items in iter_json_lists(data):
        rows = [row_to_holding(item, position, source_url, fetched_at) for item in items if isinstance(item, dict)]
        holdings = [row for row in rows if row]
        if len(holdings) > len(best):
            best = holdings
    return best


def parse_dws_json_holdings(data: Any, position: Position, source_url: str, fetched_at: str) -> list[Holding]:
    if not isinstance(data, dict) or not isinstance(data.get("tables"), list):
        return []
    best: list[Holding] = []
    for table in data["tables"]:
        if not isinstance(table, dict):
            continue
        columns = {
            clean(column.get("key")): clean(column.get("value"))
            for column in table.get("columns", [])
            if isinstance(column, dict)
        }
        rows = []
        for raw_row in table.get("values", []):
            if not isinstance(raw_row, dict):
                continue
            row = {}
            for key, label in columns.items():
                cell = raw_row.get(key)
                if isinstance(cell, dict):
                    row[label] = cell.get("value")
                else:
                    row[label] = cell
            holding = row_to_holding(row, position, source_url, fetched_at)
            if holding:
                rows.append(holding)
        if len(rows) > len(best):
            best = rows
    return best


def parse_html_holdings(payload: bytes, position: Position, source_url: str, fetched_at: str) -> list[Holding]:
    text = payload.decode("utf-8", errors="replace")
    best: list[Holding] = []
    for table_match in re.finditer(r"<table\b[^>]*>(.*?)</table>", text, flags=re.I | re.S):
        table_html = table_match.group(1)
        header_match = re.search(r"<thead\b[^>]*>(.*?)</thead>", table_html, flags=re.I | re.S)
        header_html = header_match.group(1) if header_match else table_html
        headers = [
            clean(re.sub(r"<[^>]+>", " ", unescape(cell)))
            for cell in re.findall(r"<th\b[^>]*>(.*?)</th>", header_html, flags=re.I | re.S)
        ]
        if not headers:
            continue
        rows = []
        body_match = re.search(r"<tbody\b[^>]*>(.*?)</tbody>", table_html, flags=re.I | re.S)
        body_html = body_match.group(1) if body_match else table_html
        for row_match in re.finditer(r"<tr\b[^>]*>(.*?)</tr>", body_html, flags=re.I | re.S):
            cells = [
                clean(re.sub(r"<[^>]+>", " ", unescape(cell)))
                for cell in re.findall(r"<td\b[^>]*>(.*?)</td>", row_match.group(1), flags=re.I | re.S)
            ]
            if len(cells) < 2:
                continue
            row = {headers[index]: cells[index] for index in range(min(len(headers), len(cells)))}
            holding = row_to_holding(row, position, source_url, fetched_at)
            if holding:
                rows.append(holding)
        if len(rows) > len(best):
            best = rows
    return best


def parse_holdings(payload: bytes, url: str, content_type: str, position: Position, fetched_at: str) -> tuple[str, list[Holding]]:
    lowered = f"{url} {content_type}".casefold()
    stripped_payload = payload.lstrip()
    if ".xlsx" in lowered or "spreadsheetml" in lowered:
        return "xlsx", parse_xlsx_holdings(payload, position, url, fetched_at)
    if ".xls" in lowered and ".xlsx" not in lowered:
        raise ValueError("legacy XLS files are not supported; save as XLSX/CSV first")
    if ".json" in lowered or "json" in lowered or stripped_payload.startswith((b"{", b"[")):
        return "json", parse_json_holdings(payload, position, url, fetched_at)
    if "text/html" in lowered or ".html" in lowered:
        return "html", parse_html_holdings(payload, position, url, fetched_at)
    if ".pdf" in lowered or "pdf" in lowered:
        raise ValueError("PDF documents are stored as metadata, not parsed as full holdings")
    return "csv", parse_csv_holdings(payload, position, url, fetched_at)


def validate_holdings(holdings: list[Holding]) -> tuple[list[Holding], Decimal, bool]:
    if not holdings:
        return holdings, Decimal("0"), False
    weight_sum = sum((holding.weight_pct for holding in holdings), Decimal("0"))
    if weight_sum <= Decimal("1.5"):
        for holding in holdings:
            holding.weight_pct *= Decimal("100")
        weight_sum = sum((holding.weight_pct for holding in holdings), Decimal("0"))

    normalized = False
    if weight_sum > Decimal("100"):
        for holding in holdings:
            holding.weight_pct = holding.weight_pct / weight_sum * Decimal("100")
        weight_sum = Decimal("100")
        normalized = True
    elif Decimal("0") < weight_sum < Decimal("99.5"):
        template = holdings[0]
        holdings.append(
            Holding(
                asset_name=template.asset_name,
                isin=template.isin,
                holding_name="Other issuer holdings",
                holding_ticker="",
                weight_pct=Decimal("100") - weight_sum,
                sector="Unknown from issuer data",
                geo="Unknown from issuer data",
                asset_class="ETF underlying",
                source_url=template.source_url,
                fetched_at=template.fetched_at,
            )
        )
        weight_sum = Decimal("100")
    return holdings, weight_sum, normalized


def holding_to_exposure_row(holding: Holding) -> dict[str, str]:
    return {
        "asset_name": holding.asset_name,
        "isin": holding.isin,
        "holding_name": holding.holding_name,
        "holding_ticker": holding.holding_ticker,
        "weight_pct": decimal_str(holding.weight_pct),
        "sector": holding.sector,
        "geo": holding.geo,
        "asset_class": holding.asset_class,
    }


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def save_raw_holdings(isin: str, holdings: list[Holding]) -> str:
    rows = []
    for holding in holdings:
        row = holding_to_exposure_row(holding)
        row.update(
            {
                "holding_isin": holding.holding_isin,
                "source_url": holding.source_url,
                "fetched_at": holding.fetched_at,
            }
        )
        rows.append(row)
    path = HOLDINGS_DIR / f"{isin}.csv"
    write_csv(path, rows, RAW_HOLDING_FIELDS)
    return str(path.relative_to(APP_DIR))


def load_cached_holdings(position: Position) -> list[Holding]:
    path = HOLDINGS_DIR / f"{position.isin}.csv"
    if not path.exists():
        return []
    holdings = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            holdings.append(
                Holding(
                    asset_name=clean(row.get("asset_name")) or position.asset_name,
                    isin=clean(row.get("isin")).upper() or position.isin,
                    holding_name=clean(row.get("holding_name")),
                    holding_ticker=clean(row.get("holding_ticker")),
                    weight_pct=parse_decimal(row.get("weight_pct")),
                    sector=clean(row.get("sector")),
                    geo=clean(row.get("geo")),
                    asset_class=clean(row.get("asset_class")) or "ETF underlying",
                    holding_isin=clean(row.get("holding_isin")).upper(),
                    source_url=clean(row.get("source_url")),
                    fetched_at=clean(row.get("fetched_at")),
                )
            )
    return [holding for holding in holdings if holding.weight_pct > 0]


def discover_from_seed(url: str, provider: ProviderAdapter, fetcher: OfficialFetcher) -> dict[str, str]:
    lowered = url.casefold()
    if any(token in lowered for token in ("csv", "xls", "xlsx", "json")):
        return {"holdings_url": url}
    body, final_url = fetcher.fetch_text(url)
    documents = {"product_url": final_url}
    documents.update(find_document_links(body, final_url, provider.domains))
    return documents


def fetch_position(
    position: Position,
    fetcher: OfficialFetcher,
    classifications: dict[str, dict[str, str]],
    seed_url: str = "",
    cache_only: bool = False,
    dry_run: bool = False,
) -> tuple[DocumentRecord, list[Holding]]:
    fetched_at = now_iso()
    issuer = ISSUER_BY_ISIN.get(position.isin, "")
    provider = detect_provider(position)
    if provider:
        issuer = provider.issuer

    if is_money_market(position):
        record = DocumentRecord(
            isin=position.isin,
            asset_name=position.asset_name,
            issuer=issuer or "Xtrackers",
            status="cash_equivalent",
            fetched_at=fetched_at,
            message="Classified as cash/money-market exposure; stock holdings are not expected.",
            parser="cash_equivalent",
            rows=1,
            weight_sum="100",
        )
        return record, [money_market_holding(position, fetched_at)]

    cached = load_cached_holdings(position)
    if cached and (cache_only or fetcher.offline):
        for holding in cached:
            classify_holding(holding, classifications)
        apply_issuer_specific_classification(position, cached)
        holdings, weight_sum, normalized = validate_holdings(cached)
        record = DocumentRecord(
            isin=position.isin,
            asset_name=position.asset_name,
            issuer=issuer,
            status="cached",
            fetched_at=fetched_at,
            parser="csv_cache",
            rows=len(holdings),
            weight_sum=decimal_str(weight_sum),
            normalized=normalized,
            data_path=str((HOLDINGS_DIR / f"{position.isin}.csv").relative_to(APP_DIR)),
        )
        return record, holdings

    if cache_only:
        return (
            DocumentRecord(
                isin=position.isin,
                asset_name=position.asset_name,
                issuer=issuer,
                status="official_source_not_found",
                fetched_at=fetched_at,
                message="No cached official holdings file is available.",
            ),
            [],
        )

    if not provider:
        return (
            DocumentRecord(
                isin=position.isin,
                asset_name=position.asset_name,
                issuer=issuer or "Unknown",
                status="official_source_not_found",
                fetched_at=fetched_at,
                message="No official issuer adapter is configured for this ISIN.",
            ),
            [],
        )

    if fetcher.offline:
        return (
            DocumentRecord(
                isin=position.isin,
                asset_name=position.asset_name,
                issuer=provider.issuer,
                status="official_source_not_found",
                fetched_at=fetched_at,
                message="Offline mode is enabled and no cached official holdings file is available.",
            ),
            [],
        )

    try:
        direct_holdings_url = DIRECT_HOLDINGS_URL_BY_ISIN.get(position.isin, "")
        if direct_holdings_url:
            documents = {
                "product_url": PRODUCT_URL_BY_ISIN.get(position.isin, ""),
                "holdings_url": direct_holdings_url,
            }
        else:
            documents = discover_from_seed(seed_url, provider, fetcher) if seed_url else provider.discover(position, fetcher)
        documents = {**STATIC_DOCUMENTS_BY_ISIN.get(position.isin, {}), **documents}
    except FetchError as exc:
        if cached:
            for holding in cached:
                classify_holding(holding, classifications)
            apply_issuer_specific_classification(position, cached)
            holdings, weight_sum, normalized = validate_holdings(cached)
            return (
                DocumentRecord(
                    isin=position.isin,
                    asset_name=position.asset_name,
                    issuer=issuer,
                    status="cached",
                    fetched_at=fetched_at,
                    parser="csv_cache",
                    message=f"Live fetch failed; reused cached official holdings. {exc}",
                    rows=len(holdings),
                    weight_sum=decimal_str(weight_sum),
                    normalized=normalized,
                    data_path=str((HOLDINGS_DIR / f"{position.isin}.csv").relative_to(APP_DIR)),
                ),
                holdings,
            )
        return (
            DocumentRecord(
                isin=position.isin,
                asset_name=position.asset_name,
                issuer=provider.issuer,
                status="official_source_not_found",
                fetched_at=fetched_at,
                message=str(exc),
            ),
            [],
        )

    holdings_url = documents.get("holdings_url", "")
    if not holdings_url:
        return (
            DocumentRecord(
                isin=position.isin,
                asset_name=position.asset_name,
                issuer=provider.issuer,
                status="official_source_not_found",
                fetched_at=fetched_at,
                product_url=documents.get("product_url", ""),
                kid_url=documents.get("kid_url", ""),
                factsheet_url=documents.get("factsheet_url", ""),
                prospectus_url=documents.get("prospectus_url", ""),
                message="Official product page was checked, but no full holdings file was discovered.",
            ),
            [],
        )

    try:
        payload, final_url, content_type = fetcher.fetch_bytes(holdings_url)
        parser, holdings = parse_holdings(payload, final_url, content_type, position, fetched_at)
    except (FetchError, ValueError, json.JSONDecodeError) as exc:
        return (
            DocumentRecord(
                isin=position.isin,
                asset_name=position.asset_name,
                issuer=provider.issuer,
                status="parse_failed",
                fetched_at=fetched_at,
                product_url=documents.get("product_url", ""),
                holdings_url=holdings_url,
                kid_url=documents.get("kid_url", ""),
                factsheet_url=documents.get("factsheet_url", ""),
                prospectus_url=documents.get("prospectus_url", ""),
                message=str(exc),
            ),
            [],
        )

    for holding in holdings:
        classify_holding(holding, classifications)
    apply_issuer_specific_classification(position, holdings)
    holdings, weight_sum, normalized = validate_holdings(holdings)
    if not holdings:
        status = "parse_failed"
        message = "Holdings file was downloaded but no weighted holding rows were detected."
    else:
        status = "ok"
        message = ""
        if position.isin == "IE00BK5BQT80" and parser == "html" and len(holdings) <= 20:
            status = "partial_official_holdings"
            message = (
                "Official Vanguard page exposed the visible holdings table; "
                "the full spreadsheet download endpoint was not discoverable."
            )

    data_path = "" if dry_run or not holdings else save_raw_holdings(position.isin, holdings)
    return (
        DocumentRecord(
            isin=position.isin,
            asset_name=position.asset_name,
            issuer=provider.issuer,
            status=status,
            fetched_at=fetched_at,
            product_url=documents.get("product_url", ""),
            holdings_url=holdings_url,
            kid_url=documents.get("kid_url", ""),
            factsheet_url=documents.get("factsheet_url", ""),
            prospectus_url=documents.get("prospectus_url", ""),
            parser=parser if holdings else "",
            message=message,
            rows=len(holdings),
            weight_sum=decimal_str(weight_sum),
            normalized=normalized,
            data_path=data_path,
        ),
        holdings,
    )


def backup_exposures() -> Path | None:
    if not EXPOSURES_CSV.exists():
        return None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = EXPOSURES_CSV.with_name(f"asset_exposures.backup_{stamp}.csv")
    shutil.copy2(EXPOSURES_CSV, backup)
    return backup


def build_seed_map(values: list[str]) -> dict[str, str]:
    output: dict[str, str] = {}
    for value in values:
        if "=" not in value:
            raise SystemExit(f"Invalid --seed-url value {value!r}; expected ISIN=URL.")
        isin, url = value.split("=", 1)
        output[isin.strip().upper()] = url.strip()
    return output


def update(args: argparse.Namespace) -> int:
    if not args.official_only:
        raise SystemExit("Only --official-only is supported.")

    positions = current_open_positions(args.person)
    existing_rows = load_existing_exposure_rows()
    classifications = load_classifications()
    fetcher = OfficialFetcher(offline=args.offline, timeout=args.timeout)
    seed_urls = build_seed_map(args.seed_url or [])

    records: list[DocumentRecord] = []
    exposure_rows: list[dict[str, str]] = []
    target_count = 0
    resolved_count = 0

    for position in positions:
        direct = direct_share_row(position, existing_rows)
        if direct:
            exposure_rows.append(direct)
            continue
        if not target_needs_composition(position, existing_rows):
            continue
        target_count += 1
        record, holdings = fetch_position(
            position,
            fetcher,
            classifications,
            seed_url=seed_urls.get(position.isin, ""),
            cache_only=args.cache_only,
            dry_run=args.dry_run,
        )
        records.append(record)
        if holdings:
            resolved_count += 1
            exposure_rows.extend(holding_to_exposure_row(holding) for holding in holdings)

    print(f"Open positions: {len(positions)}")
    print(f"Composition targets: {target_count}")
    print(f"Official/cached compositions resolved: {resolved_count}")
    for record in records:
        suffix = f" ({record.rows} rows, {record.weight_sum}%)" if record.rows else f" - {record.message}"
        print(f"{record.isin or 'NOISIN'} {record.asset_name}: {record.status}{suffix}")

    if args.dry_run:
        print("Dry-run enabled: no files were written.")
        return 0

    backup = None if args.no_backup else backup_exposures()
    if backup:
        print(f"Backed up existing exposures to {backup.relative_to(APP_DIR)}")
    save_documents(records)
    write_csv(EXPOSURES_CSV, exposure_rows, EXPOSURE_FIELDS)
    print(f"Wrote {len(exposure_rows)} rows to {EXPOSURES_CSV.relative_to(APP_DIR)}")
    print(f"Wrote metadata to {DOCUMENTS_JSON.relative_to(APP_DIR)}")
    return 0


def list_targets(args: argparse.Namespace) -> int:
    rows = load_existing_exposure_rows()
    for position in current_open_positions(args.person):
        if target_needs_composition(position, rows):
            provider = detect_provider(position)
            issuer = provider.issuer if provider else ISSUER_BY_ISIN.get(position.isin, "Unknown")
            print(f"{position.isin or 'NOISIN'}\t{issuer}\t{position.asset_name}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Fetch official ETF holdings and regenerate asset_exposures.csv.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    update_parser = subparsers.add_parser("update", help="Discover, fetch, parse, and write official exposures.")
    update_parser.add_argument("--official-only", action="store_true", default=True, help="Use official issuer sources only.")
    update_parser.add_argument("--dry-run", action="store_true", help="Do not write metadata, holdings, or exposures.")
    update_parser.add_argument("--offline", action="store_true", help="Do not use the network; use cached files only.")
    update_parser.add_argument("--cache-only", action="store_true", help="Use saved data/etf_holdings files only.")
    update_parser.add_argument("--person", default=dashboard.PRIMARY_PORTFOLIO_ID, help="Portfolio key to inspect.")
    update_parser.add_argument("--seed-url", action="append", default=[], help="Official URL override as ISIN=URL.")
    update_parser.add_argument("--timeout", type=int, default=25, help="HTTP timeout in seconds.")
    update_parser.add_argument("--no-backup", action="store_true", help="Do not back up asset_exposures.csv before writing.")
    update_parser.set_defaults(func=update)

    list_parser = subparsers.add_parser("list-targets", help="List open ETF/fund composition targets.")
    list_parser.add_argument("--person", default=dashboard.PRIMARY_PORTFOLIO_ID, help="Portfolio key to inspect.")
    list_parser.set_defaults(func=list_targets)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
