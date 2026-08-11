from __future__ import annotations

import csv
import io
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import app
from src.portfolio_dashboard.imports import detect_statement_source

PERSONAL_ROWS = [
    ["2026-01-02", "BUY", "Example ETF", "IE0000000001", "Personal", "EUR", 2, 50, 1, 0, 101],
    ["2026-02-03", "SELL", "Example ETF", "IE0000000001", "Personal", "EUR", 1, 60, 1, 2, 57],
]


def personal_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trades"
    sheet.append(app.PERSONAL_TRADE_COLUMNS)
    for row in PERSONAL_ROWS:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


class PersonalTradeImportTest(unittest.TestCase):
    def test_legacy_seventeen_column_csv_remains_supported(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "legacy.csv"
            with path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerow(
                    ["Nome", "Ticker", "Acq/Ven", "Val", "Saveback?", "Giorno", "Mese", "Anno", "Val", "Prezzo", "QT", "QT diff", "TOT + saveback", "TOT spesa", "FEES", "TAX", "GRAND TOT"]
                )
                writer.writerow(["Example ETF", "", "Acquisto", "EUR", "", "02/01", "", "2026", "", "50", "2", "2", "", "101", "1", "0", "101"])

            trades = app.read_manual_trades(path)

        self.assertEqual(len(trades), 1)
        self.assertEqual(trades[0].asset, "Example ETF")
        self.assertEqual(trades[0].quantity_diff, app.Decimal("2"))

    def test_standard_csv_and_xlsx_produce_the_same_trades(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            csv_path = root / "personal.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerow(app.PERSONAL_TRADE_COLUMNS)
                writer.writerows(PERSONAL_ROWS)
            xlsx_path = root / "personal.xlsx"
            xlsx_path.write_bytes(personal_workbook_bytes())

            csv_trades = app.read_manual_trades(csv_path)
            xlsx_trades = app.read_manual_trades(xlsx_path)

        self.assertEqual(csv_trades, xlsx_trades)
        self.assertEqual(csv_trades[0].quantity_diff, app.Decimal("2"))
        self.assertEqual(csv_trades[1].quantity_diff, app.Decimal("-1"))
        self.assertEqual(csv_trades[1].grand_total, app.Decimal("57"))

    def test_auto_detection_recognizes_personal_xlsx_headers(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "personal.xlsx"
            path.write_bytes(personal_workbook_bytes())

            source = detect_statement_source(path)

        self.assertEqual(source, "manual")

    def test_template_endpoints_return_fillable_csv_and_xlsx(self) -> None:
        with app.app.test_client() as client:
            csv_response = client.get("/api/imports/template?format=csv")
            xlsx_response = client.get("/api/imports/template?format=xlsx")

        self.assertEqual(csv_response.status_code, 200)
        self.assertIn(",".join(app.PERSONAL_TRADE_COLUMNS), csv_response.get_data(as_text=True))
        workbook = load_workbook(io.BytesIO(xlsx_response.data), read_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["Trades", "Instructions"])
            self.assertEqual(tuple(cell.value for cell in workbook["Trades"][1]), app.PERSONAL_TRADE_COLUMNS)
        finally:
            workbook.close()

    def test_browser_import_accepts_personal_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            with (
                patch.object(app, "ROOT_DIR", root / "sources"),
                patch.object(app, "MOVEMENT_DATABASE", root / "private" / "movements.sqlite3"),
                patch.object(app, "_movement_store", None),
                app.app.test_client() as client,
            ):
                response = client.post(
                    "/api/imports",
                    data={"source": "auto", "file": (io.BytesIO(personal_workbook_bytes()), "personal.xlsx")},
                    content_type="multipart/form-data",
                )
            imported = list((root / "sources" / "broker_exports" / "manual").glob("*.xlsx"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["source"], "manual")
        self.assertEqual(response.get_json()["movements"], 2)
        self.assertEqual(len(imported), 1)


if __name__ == "__main__":
    unittest.main()
