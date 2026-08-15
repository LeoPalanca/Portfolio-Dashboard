from __future__ import annotations

import csv
import email.utils
import hashlib
import html
import io
import json
import math
import os
import re
import shutil
import subprocess
import tempfile
import threading
import time
import urllib.parse
import urllib.request
import warnings
import xml.etree.ElementTree as ET
from bisect import bisect_right
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from src.portfolio_dashboard import APP_VERSION, display_version
from src.portfolio_dashboard.cache import HistoryStore
from src.portfolio_dashboard.config import get_settings
from src.portfolio_dashboard.domain import (
    ZERO,
    Dividend,
    ExpenseRule,
    FrictionEvent,
    Trade,
    decimal_to_float,
    money,
    parse_decimal,
)
from src.portfolio_dashboard.imports import (
    SOURCE_EXTENSIONS,
    SOURCE_LABELS,
    detect_statement_source,
    fineco_statement_kind,
    import_destination,
    read_fineco_bank_movements,
    source_format_label,
)
from src.portfolio_dashboard.ingest import BrokerAdapter, FunctionBrokerAdapter
from src.portfolio_dashboard.movements import MovementStore

try:
    import yfinance as yf
except Exception:  # pragma: no cover - handled at runtime in the dashboard
    yf = None


SETTINGS = get_settings()
DISPLAY_VERSION = display_version(APP_VERSION, SETTINGS.edition_suffix)
DEFAULT_PROXY_MODE = SETTINGS.default_proxy_mode
APP_DIR = SETTINGS.project_dir
ROOT_DIR = SETTINGS.source_dir
PRIMARY_PORTFOLIO_ID = SETTINGS.primary_portfolio_id.lower()
PRIMARY_PORTFOLIO_NAME = SETTINGS.primary_portfolio_name
SINCE_2024_PORTFOLIO_IDS = {portfolio_id.strip().lower() for portfolio_id in SETTINGS.since_2024_portfolio_ids}
TRADES_CSV = ROOT_DIR / SETTINGS.manual_trades_file
PERSONAL_TRADE_COLUMNS = (
    "date",
    "action",
    "asset",
    "isin",
    "broker",
    "currency",
    "quantity",
    "price",
    "fees",
    "tax",
    "total",
)
TRADE_REPUBLIC_PATTERN = SETTINGS.trade_republic_pattern
FINECO_PATTERN = SETTINGS.fineco_pattern
IB_PATTERN = SETTINGS.interactive_brokers_pattern
ETORO_PATTERN = SETTINGS.etoro_pattern
REVOLUT_PATTERN = SETTINGS.revolut_pattern
REVOLUT_DOWNLOADS_PATTERN = SETTINGS.revolut_downloads_pattern
INTESA_OPERATIONS_PATTERN = SETTINGS.intesa_pattern
INTESA_DOWNLOADS_PATTERN = SETTINGS.intesa_downloads_pattern
FAMILY_PORTFOLIOS = SETTINGS.family_portfolios
MAPPINGS_CSV = SETTINGS.data_path("asset_mappings.csv")
EXPOSURES_CSV = SETTINGS.data_path("asset_exposures.csv")
ETF_DOCUMENTS_JSON = APP_DIR / "data" / "etf_documents.json"
BERKSHIRE_HOLDINGS_CSV = APP_DIR / "data" / "berkshire_holdings.csv"
BERKSHIRE_ISIN = "US0846707026"
PROXY_EXPOSURES_CSV = APP_DIR / "data" / "proxy_exposures.csv"
CRYPTO_WALLETS_CSV = SETTINGS.data_path("crypto_wallets.csv")
CRYPTO_WALLET_POSITIONS_CSV = SETTINGS.data_path("crypto_wallet_positions.csv")
CRYPTO_WALLET_TRANSACTIONS_CSV = SETTINGS.data_path("crypto_wallet_transactions.csv")
EXPENSE_RULES_CSV = SETTINGS.data_path("expense_category_rules.csv")
MOVEMENT_DATABASE = SETTINGS.data_path(SETTINGS.movement_database_file)
PROXY_ISSUERS = {
    "IE00BK5BQT80": "Vanguard",
    "LU1681045370": "Amundi",
    "IE00BKM4GZ66": "iShares",
    "NL0011683594": "VanEck",
    "IE00B2NLMV86": "Mediolanum",
    "IE00BYZ2Y955": "Mediolanum",
    "IT0001019329": "Mediolanum",
    "IT0001280541": "Eurizon",
    "IT0005285157": "Eurizon",
    "IT0005104424": "Eurizon",
    "LU0497415702": "Eurizon",
    "IT0004896715": "Anima",
    "LU2146152231": "JPMorgan",
    "LU0129441100": "JPMorgan",
    "LU0109394709": "Franklin Templeton",
    "LU0195953079": "Franklin Templeton",
    "LU0755218046": "Fidelity",
    "LU0106239527": "Schroders",
}
PROXY_SOURCE_URLS = {
    "IE00BK5BQT80": "https://companiesmarketcap.com/vanguard-ftse-all-world-ucits-etf-usd-accumulation/holdings/",
    "LU1681045370": "https://www.ishares.com/uk/individual/en/products/264659/fund/1506575576011.ajax?fileType=csv&fileName=EIMI_holdings&dataType=fund",
    "IE00BKM4GZ66": "https://www.ishares.com/uk/individual/en/products/264659/fund/1506575576011.ajax?fileType=csv&fileName=EIMI_holdings&dataType=fund",
    "NL0011683594": "https://companiesmarketcap.com/vaneck-morningstar-developed-markets-dividend-leaders-ucits-etf/holdings/",
    "IE00B2NLMV86": "https://www.ishares.com/uk/individual/en/products/251882/ishares-msci-world-ucits-etf-acc-fund",
    "IE00BYZ2Y955": "https://companiesmarketcap.com/vaneck-morningstar-developed-markets-dividend-leaders-ucits-etf/holdings/",
    "LU0755218046": "https://www.ishares.com/uk/individual/en/products/251900/ishares-core-sp-500-ucits-etf-acc-fund",
}
PROXY_MESSAGES = {
    "IE00BK5BQT80": "Non-official 1,500-row constituent proxy normalized to 100%; switch to Official only to hide it.",
    "LU1681045370": "Official iShares MSCI EM IMI holdings used as an economic proxy and normalized to 100%; switch to Official only to hide it.",
    "IE00BKM4GZ66": "Official iShares MSCI EM IMI holdings normalized to 100%; switch to Official only to hide it.",
    "NL0011683594": "Non-official 101-row constituent proxy; switch to Official only to hide it.",
    "IE00B2NLMV86": "Official iShares MSCI World holdings used as a global equity fund proxy and normalized to 100%; switch to Official only to hide it.",
    "IE00BYZ2Y955": "Non-official VanEck Developed Markets Dividend Leaders holdings used as a value equity proxy and normalized to 100%; switch to Official only to hide it.",
    "IT0001019329": "Mock PIR-compliant proxy (70% top Italian equities, 30% cash/gov bonds) normalized to 100%; switch to Official only to hide it.",
    "IT0001280541": "Mock Euro High Yield Corporate Bond proxy (80% corporate bonds, 20% short-term gov bonds) normalized to 100%; switch to Official only to hide it.",
    "IT0005285157": "Mock Defensive Profilo Flessibile proxy (10% MSCI World, 60% Eurozone gov bonds, 30% cash) normalized to 100%; switch to Official only to hide it.",
    "IT0005104424": "Mock Short-term Gov Bond proxy (70% Eurozone short-term gov bonds, 30% cash) normalized to 100%; switch to Official only to hide it.",
    "LU0497415702": "Mock Flexible Equity Strategy proxy (55% MSCI World, 30% Eurozone bonds, 15% cash) normalized to 100%; switch to Official only to hide it.",
    "IT0004896715": "Mock Balanced proxy (60% MSCI World, 40% Euro cash/bonds) normalized to 100%; switch to Official only to hide it.",
    "LU2146152231": "Mock MSCI Europe Index proxy using top 9 European equity giants normalized to 100%; switch to Official only to hide it.",
    "LU0129441100": "Mock MSCI Europe Index proxy using top 9 European equity giants normalized to 100%; switch to Official only to hide it.",
    "LU0109394709": "Mock Biotechnology Index proxy using top 5 biotech leaders normalized to 100%; switch to Official only to hide it.",
    "LU0195953079": "Mock Global Bond Index proxy (40% US, 40% Eurozone, 20% Japan) normalized to 100%; switch to Official only to hide it.",
    "LU0755218046": "Official iShares S&P 500 holdings used as an economic proxy and normalized to 100%; switch to Official only to hide it.",
    "LU0106239527": "Mock Italian Equity Index proxy using top 6 Italian leaders normalized to 100%; switch to Official only to hide it.",
}
FULL_OFFICIAL_COMPOSITION_STATUSES = {"ok", "cached", "cash_equivalent", "official_sec_13f"}
SYMBOL_CACHE = SETTINGS.cache_path("price-symbols.json")
PRICE_CACHE = SETTINGS.cache_path("prices.json")
HISTORY_CACHE = SETTINGS.cache_path("history.json")
HISTORY_CACHE_DIR = SETTINGS.cache_dir / "history"
NEWS_CACHE = SETTINGS.cache_path("news.json")
PRICE_TTL_SECONDS = 15 * 60
HISTORY_TTL_SECONDS = 12 * 60 * 60
NEWS_TTL_SECONDS = 60 * 60

FINECO_DIVIDEND_NET_RATE = Decimal(1) - SETTINGS.fineco_withholding_tax_rate
BBVA_INTEREST_NET_RATE = Decimal(1) - SETTINGS.bbva_interest_tax_rate
_CACHE_WRITE_LOCK = threading.RLock()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = SETTINGS.import_max_bytes
app.jinja_env.keep_trailing_newline = True


@app.errorhandler(413)
def upload_too_large(_error):
    return jsonify({"error": f"Statement files must be smaller than {SETTINGS.import_max_bytes // (1024 * 1024)} MB"}), 413


def parse_trade_date(day_value: str, year_value: str) -> date:
    raw = day_value.strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass

    parts = raw.split("/")
    if len(parts) == 2 and year_value.strip():
        day, month = parts
        return date(int(year_value.strip()), int(month), int(day))

    raise ValueError(f"Unsupported trade date: {day_value!r}")


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_json(path: Path, payload: Any) -> None:
    encoded = json.dumps(payload, indent=2, sort_keys=True)
    tmp_path = path.with_name(f".{path.name}.{os.getpid()}.{threading.get_ident()}.tmp")
    with _CACHE_WRITE_LOCK:
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path.write_text(encoded, encoding="utf-8")
            tmp_path.replace(path)
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except OSError:
                    pass


def configured_path_label(path: Path) -> str:
    """Return a stable path label without exposing an absolute local path."""

    for prefix, base in (
        ("private", SETTINGS.data_dir),
        ("sources", ROOT_DIR),
        ("public", APP_DIR),
        ("cache", SETTINGS.cache_dir),
    ):
        try:
            return str(Path(prefix) / path.relative_to(base))
        except ValueError:
            continue
    return path.name


def latest_trade_republic_export() -> Path | None:
    files = sorted(set(ROOT_DIR.glob(TRADE_REPUBLIC_PATTERN)) | set((ROOT_DIR / "broker_exports" / "trade_republic").glob("*.csv")))
    family_markers = {
        (profile.trade_republic_name or portfolio_id).lower()
        for portfolio_id, profile in SETTINGS.portfolios.items()
        if portfolio_id.lower() != PRIMARY_PORTFOLIO_ID
    }
    files = [path for path in files if not any(marker in path.name.lower() for marker in family_markers)]
    return files[-1] if files else None


def latest_family_trade_republic_export(person: str) -> Path | None:
    files = sorted(set(ROOT_DIR.glob(TRADE_REPUBLIC_PATTERN)) | set((ROOT_DIR / "broker_exports" / "trade_republic").glob("*.csv")))
    profile = SETTINGS.portfolios.get(person.lower())
    marker = (profile.trade_republic_name if profile else None) or person
    files = [f for f in files if marker.lower() in f.name.lower()]
    return files[-1] if files else None


def latest_fineco_export() -> Path | None:
    files = sorted(set(ROOT_DIR.glob(FINECO_PATTERN)) | set((ROOT_DIR / "broker_exports" / "fineco").glob("*.xlsx")))
    return files[-1] if files else None


def latest_ib_export() -> Path | None:
    files = sorted(set(ROOT_DIR.glob(IB_PATTERN)) | set((ROOT_DIR / "broker_exports" / "interactive_brokers").glob("*.pdf")))
    return files[-1] if files else None


def latest_etoro_export() -> Path | None:
    files = sorted(set(ROOT_DIR.glob(ETORO_PATTERN)) | set((ROOT_DIR / "broker_exports" / "etoro").glob("*.xlsx")))
    return files[-1] if files else None


def intesa_operations_sort_key(path: Path) -> tuple[date, float, str]:
    match = re.search(r"(\d{8})", path.stem)
    export_date = date.min
    if match:
        try:
            export_date = datetime.strptime(match.group(1), "%d%m%Y").date()
        except ValueError:
            export_date = date.min
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0.0
    return export_date, mtime, path.name


def intesa_operations_files() -> list[Path]:
    candidates: dict[Path, Path] = {}
    for path in ROOT_DIR.glob(INTESA_OPERATIONS_PATTERN):
        candidates[path.resolve()] = path
    for path in (ROOT_DIR / "cash_exports" / "intesa").glob("*.xlsx"):
        candidates[path.resolve()] = path
    downloads = Path.home() / "Downloads"
    if SETTINGS.scan_downloads and downloads.exists():
        for path in downloads.glob(INTESA_DOWNLOADS_PATTERN):
            candidates[path.resolve()] = path
    return sorted(candidates.values(), key=intesa_operations_sort_key)


def latest_intesa_operations_export() -> Path | None:
    files = intesa_operations_files()
    return files[-1] if files else None


def bbva_statement_files() -> list[Path]:
    return sorted(set(ROOT_DIR.glob(SETTINGS.bbva_pattern)) | set((ROOT_DIR / "cash_exports" / "bbva").glob("*.xls")))


def parse_revolut_datetime(value: str | None) -> datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def revolut_statement_profile(path: Path) -> tuple[tuple[str, ...], date, int] | None:
    currencies: set[str] = set()
    latest_date: date | None = None
    completed_rows = 0
    try:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                if row.get("State") != "COMPLETATO":
                    continue
                event_time = parse_revolut_datetime(row.get("Data di completamento")) or parse_revolut_datetime(row.get("Data di inizio"))
                if event_time is None:
                    continue
                currencies.add(normalize_currency_code(row.get("Valuta") or "EUR"))
                latest_date = max(latest_date or event_time.date(), event_time.date())
                completed_rows += 1
    except OSError:
        return None
    if not completed_rows or latest_date is None:
        return None
    return tuple(sorted(currencies)), latest_date, completed_rows


def revolut_statement_files() -> list[Path]:
    candidates: list[Path] = []
    candidates.extend(ROOT_DIR.glob(REVOLUT_PATTERN))
    candidates.extend((ROOT_DIR / "cash_exports" / "revolut").glob("*.csv"))
    downloads = Path.home() / "Downloads"
    if SETTINGS.scan_downloads and downloads.exists():
        candidates.extend(downloads.glob(REVOLUT_DOWNLOADS_PATTERN))

    best_by_currency: dict[tuple[str, ...], tuple[date, int, float, Path]] = {}
    for path in candidates:
        profile = revolut_statement_profile(path)
        if profile is None:
            continue
        currencies, latest_date, completed_rows = profile
        try:
            mtime = path.stat().st_mtime
        except OSError:
            mtime = 0.0
        candidate = (latest_date, completed_rows, mtime, path)
        current = best_by_currency.get(currencies)
        if current is None or candidate[:3] > current[:3]:
            best_by_currency[currencies] = candidate

    return sorted((item[3] for item in best_by_currency.values()), key=lambda item: item.name)


def broker_adapters() -> tuple[BrokerAdapter, ...]:
    return (
        FunctionBrokerAdapter("Trade Republic", latest_trade_republic_export, read_trade_republic_trades),
        FunctionBrokerAdapter("Fineco", latest_fineco_export, read_fineco_trades),
        FunctionBrokerAdapter("Interactive Brokers", latest_ib_export, read_interactive_brokers_trades),
        FunctionBrokerAdapter("eToro", latest_etoro_export, read_etoro_trades),
    )


def read_trades() -> tuple[list[Trade], dict[str, Any]]:
    broker_sources: list[dict[str, Any]] = []
    trades: list[Trade] = []

    for adapter in broker_adapters():
        export_path = adapter.discover()
        if export_path is None:
            continue
        trades.extend(adapter.parse(export_path))
        broker_sources.append(
            {
                "path": export_path,
                "kind": adapter.name,
                "relative_path": configured_path_label(export_path),
            }
        )

    if broker_sources:
        return sorted(trades, key=lambda item: (item.date, item.broker, item.asset, item.action)), {
            "kind": " + ".join(source["kind"] for source in broker_sources),
            "relative_path": ", ".join(source["relative_path"] for source in broker_sources),
            "sources": broker_sources,
        }

    personal_files = manual_trade_files()
    if not personal_files:
        return [], {
            "kind": "No imported statements",
            "relative_path": "",
            "sources": [],
        }
    personal_file = personal_files[-1]
    trades = read_manual_trades(personal_file)
    return trades, {
        "path": personal_file,
        "kind": "Personal trades",
        "relative_path": configured_path_label(personal_file),
    }


def manual_trade_files() -> list[Path]:
    candidates = [TRADES_CSV] if TRADES_CSV.exists() else []
    personal_directory = ROOT_DIR / "broker_exports" / "manual"
    candidates.extend(personal_directory.glob("*.csv"))
    candidates.extend(personal_directory.glob("*.xlsx"))

    def sort_key(path: Path) -> tuple[float, str]:
        try:
            return path.stat().st_mtime, path.name
        except OSError:
            return 0.0, path.name

    return sorted(set(candidates), key=sort_key)


def normalize_personal_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")


def parse_personal_trade_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported personal trade date: {raw!r}; use YYYY-MM-DD")


def personal_trade_from_row(row: dict[str, Any], line_number: int) -> Trade:
    values = {normalize_personal_header(key): value for key, value in row.items()}
    asset = str(values.get("asset") or "").strip()
    if not asset:
        raise ValueError(f"Personal trades row {line_number} is missing asset")
    action_raw = str(values.get("action") or "").strip().upper()
    if action_raw in {"BUY", "ACQUISTO", "PURCHASE"}:
        action = "Acquisto"
        direction = Decimal("1")
    elif action_raw in {"SELL", "VENDITA", "SALE"}:
        action = "Vendita"
        direction = Decimal("-1")
    else:
        raise ValueError(f"Personal trades row {line_number} action must be BUY or SELL")

    quantity = abs(parse_decimal(str(values.get("quantity") or "")))
    price = abs(parse_decimal(str(values.get("price") or "")))
    if quantity <= ZERO or price <= ZERO:
        raise ValueError(f"Personal trades row {line_number} requires positive quantity and price")
    fees = abs(parse_decimal(str(values.get("fees") or "")))
    tax = abs(parse_decimal(str(values.get("tax") or "")))
    total_raw = str(values.get("total") or "").strip()
    gross = price * quantity
    derived_total = gross + fees + tax if direction > ZERO else max(ZERO, gross - fees - tax)
    total = abs(parse_decimal(total_raw)) if total_raw else derived_total
    currency = str(values.get("currency") or "EUR").strip().upper() or "EUR"
    return Trade(
        asset=asset,
        isin=str(values.get("isin") or "").strip().upper(),
        broker=str(values.get("broker") or "Personal").strip() or "Personal",
        action=action,
        currency_hint=currency,
        cash_currency=currency,
        date=parse_personal_trade_date(values.get("date")),
        price=price,
        quantity=quantity,
        quantity_diff=quantity * direction,
        total_spend=total,
        fees=fees,
        tax=tax,
        grand_total=total,
        grand_total_present=bool(total_raw),
        source="manual",
    )


def read_standard_personal_csv(path: Path) -> list[Trade]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        headers = {normalize_personal_header(value) for value in (reader.fieldnames or [])}
        required = {"date", "action", "asset", "quantity", "price"}
        if not required.issubset(headers):
            missing = ", ".join(sorted(required - headers))
            raise ValueError(f"Personal trades CSV is missing required columns: {missing}")
        return [
            personal_trade_from_row(row, line_number)
            for line_number, row in enumerate(reader, start=2)
            if any(str(value or "").strip() for value in row.values())
        ]


def read_personal_xlsx(path: Path) -> list[Trade]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook["Trades"] if "Trades" in workbook.sheetnames else workbook.active
        rows = iter(sheet.iter_rows(values_only=True))
        headers: list[str] | None = None
        header_row_number = 0
        for header_row_number, row in enumerate(rows, start=1):
            candidate = [normalize_personal_header(value) for value in row]
            if {"date", "action", "asset", "quantity", "price"}.issubset(set(candidate)):
                headers = candidate
                break
            if header_row_number >= 30:
                break
        if headers is None:
            raise ValueError("Personal trades XLSX is missing date, action, asset, quantity, and price columns")
        trades = []
        for row_number, row in enumerate(rows, start=header_row_number + 1):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            record = {header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
            trades.append(personal_trade_from_row(record, row_number))
        return trades
    finally:
        workbook.close()


def read_legacy_manual_csv(path: Path) -> list[Trade]:
    trades: list[Trade] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header = next(reader, None)
        if not header or len(header) < 17:
            raise ValueError("Trades CSV does not match the expected 17-column export.")

        for line_number, row in enumerate(reader, start=2):
            if len(row) != len(header):
                raise ValueError(f"Invalid column count on row {line_number}.")
            if not row[0].strip():
                continue

            trade_date = parse_trade_date(row[5], row[7])
            grand_total_raw = row[16].strip()
            trades.append(
                Trade(
                    asset=row[0].strip(),
                    isin="",
                    broker="Manual",
                    action=row[2].strip(),
                    currency_hint=(row[3].strip() or "E").upper(),
                    cash_currency="EUR",
                    date=trade_date,
                    price=parse_decimal(row[9]),
                    quantity=parse_decimal(row[10]),
                    quantity_diff=parse_decimal(row[11]),
                    total_spend=parse_decimal(row[13]),
                    fees=parse_decimal(row[14]),
                    tax=parse_decimal(row[15]),
                    grand_total=parse_decimal(grand_total_raw),
                    grand_total_present=bool(grand_total_raw),
                    source="manual",
                )
            )

    return sorted(trades, key=lambda item: (item.date, item.asset, item.action))


def read_manual_trades(path: Path) -> list[Trade]:
    if path.suffix.casefold() == ".xlsx":
        trades = read_personal_xlsx(path)
    else:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            headers = {normalize_personal_header(value) for value in next(csv.reader(handle), [])}
        if {"date", "action", "asset", "quantity", "price"}.issubset(headers):
            trades = read_standard_personal_csv(path)
        else:
            trades = read_legacy_manual_csv(path)
    return sorted(trades, key=lambda item: (item.date, item.asset, item.action))


def read_trade_republic_trades(path: Path) -> list[Trade]:
    trades: list[Trade] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if row.get("type") not in {"BUY", "SELL"}:
                continue
            if row.get("category") != "TRADING" or not row.get("shares"):
                continue

            raw_fee = parse_decimal(row.get("fee"))
            raw_tax = parse_decimal(row.get("tax"))
            cash_effect = parse_decimal(row.get("amount")) + raw_fee + raw_tax
            trade_type = row["type"]
            trades.append(
                Trade(
                    asset=(row.get("name") or row.get("symbol") or "Unknown").strip(),
                    isin=(row.get("symbol") or "").strip().upper(),
                    broker="Trade Republic",
                    action="Acquisto" if trade_type == "BUY" else "Vendita",
                    currency_hint=(row.get("currency") or "EUR").strip().upper(),
                    cash_currency="EUR",
                    date=datetime.strptime(row["date"], "%Y-%m-%d").date(),
                    price=parse_decimal(row.get("price")),
                    quantity=abs(parse_decimal(row.get("shares"))),
                    quantity_diff=parse_decimal(row.get("shares")),
                    total_spend=cash_effect,
                    fees=-raw_fee,
                    tax=-raw_tax,
                    grand_total=cash_effect,
                    grand_total_present=True,
                    source="trade_republic",
                )
            )

    return sorted(trades, key=lambda item: (item.date, item.asset, item.action))


def parse_fineco_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()


def fineco_decimal(value: Any) -> Decimal:
    if value is None:
        return ZERO
    return parse_decimal(str(value))


def fineco_dividend_tax_from_net(net_amount: Decimal) -> Decimal:
    if net_amount <= ZERO:
        return ZERO
    gross = net_amount / FINECO_DIVIDEND_NET_RATE
    return gross - net_amount


def fineco_capital_gain_tax_from_net_gain(net_gain: Decimal) -> Decimal:
    if net_gain <= ZERO:
        return ZERO
    gross_gain = net_gain / FINECO_DIVIDEND_NET_RATE
    return gross_gain - net_gain


def read_fineco_trades(path: Path) -> list[Trade]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Movimenti Dossier Titoli"]
    trades: list[Trade] = []

    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not row or row[2] != "Compravendita titoli":
            continue
        sign = str(row[5] or "").strip().upper()
        if sign not in {"A", "V"}:
            continue

        quantity = abs(fineco_decimal(row[6]))
        value_eur = fineco_decimal(row[10])
        commission = fineco_decimal(row[14])
        is_buy = sign == "A"
        cash_effect = -(value_eur + commission) if is_buy else value_eur - commission
        trades.append(
            Trade(
                asset=str(row[3]).strip(),
                isin=str(row[4] or "").strip().upper(),
                broker="Fineco",
                action="Acquisto" if is_buy else "Vendita",
                currency_hint=str(row[7] or "EUR").strip().upper(),
                cash_currency="EUR",
                date=parse_fineco_date(row[0]),
                price=fineco_decimal(row[8]),
                quantity=quantity,
                quantity_diff=quantity if is_buy else -quantity,
                total_spend=cash_effect,
                fees=commission,
                tax=ZERO,
                grand_total=cash_effect,
                grand_total_present=True,
                source="fineco",
            )
        )

    return sorted(trades, key=lambda item: (item.date, item.asset, item.action))


def pdf_text(path: Path) -> str:
    result = subprocess.run(["pdftotext", "-layout", str(path), "-"], check=True, capture_output=True, text=True)
    return result.stdout


def parse_ib_instruments(text: str) -> dict[str, dict[str, str]]:
    instruments: dict[str, dict[str, str]] = {}
    for line in text.splitlines():
        match = re.match(r"^\s*([A-Z0-9.]+)\s+(.+?)\s+(\d+)\s+([A-Z]{2}[A-Z0-9]{10})\s+([A-Z0-9.]+)\s+([A-Z]+)", line)
        if not match:
            continue
        symbol, description, _, isin, _, exchange = match.groups()
        if symbol == "EUR.USD":
            continue
        instruments[symbol] = {
            "asset": re.sub(r"\s+", " ", description).strip(),
            "isin": isin,
            "exchange": exchange,
        }
    return instruments


def read_interactive_brokers_trades(path: Path) -> list[Trade]:
    text = pdf_text(path)
    instruments = parse_ib_instruments(text)
    trades: list[Trade] = []
    section_currency = "EUR"
    stock_section = False
    pattern = re.compile(
        r"^U16961051\s+(\S+)\s+(\d{4}-\d{2}-\d{2}),\s+\d{2}:\d{2}:\d{2}\s+"
        r"(\d{4}-\d{2}-\d{2})\s+-\s+(BUY|SELL)\s+(-?\d+(?:\.\d+)?)\s+"
        r"(\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)"
    )

    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        stripped = line.strip()
        if stripped == "Azioni":
            stock_section = True
            continue
        if stripped == "Forex":
            stock_section = False
            continue
        if stock_section and stripped in {"EUR", "USD", "GBP"}:
            section_currency = stripped
            continue
        if not stock_section:
            continue

        match = pattern.match(line)
        if not match:
            continue
        symbol, trade_date, _, side, quantity, price, proceeds, commission, tax = match.groups()
        if "." in symbol:
            continue
        instrument = instruments.get(symbol, {})
        qty = parse_decimal(quantity)
        cash_effect = parse_decimal(proceeds) + parse_decimal(commission) + parse_decimal(tax)
        is_buy = side == "BUY"
        trades.append(
            Trade(
                asset=instrument.get("asset") or symbol,
                isin=instrument.get("isin", ""),
                broker="Interactive Brokers",
                action="Acquisto" if is_buy else "Vendita",
                currency_hint=section_currency,
                cash_currency=section_currency,
                date=datetime.strptime(trade_date, "%Y-%m-%d").date(),
                price=parse_decimal(price),
                quantity=abs(qty),
                quantity_diff=abs(qty) if is_buy else -abs(qty),
                total_spend=cash_effect,
                fees=abs(parse_decimal(commission)),
                tax=abs(parse_decimal(tax)),
                grand_total=cash_effect,
                grand_total_present=True,
                source="interactive_brokers",
            )
        )

    return sorted(trades, key=lambda item: (item.date, item.asset, item.action))


def read_etoro_trades(path: Path) -> list[Trade]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    dividends = workbook["Dividendi"]
    position_isins: dict[str, str] = {}
    for row in dividends.iter_rows(min_row=2, values_only=True):
        position_id = str(row[11] or "").strip()
        isin = str(row[13] or "").strip().upper()
        if position_id and isin:
            position_isins[position_id] = isin

    activity = workbook["Attività account"]
    trades: list[Trade] = []
    for row in activity.iter_rows(min_row=2, values_only=True):
        if not row or row[1] != "Apri posizione":
            continue
        opened_at = datetime.strptime(str(row[0]), "%d/%m/%Y %H:%M:%S").date()
        detail = str(row[2] or "").strip()
        position_id = str(row[8] or "").strip()
        amount = fineco_decimal(row[3])
        units = fineco_decimal(row[4])
        isin = position_isins.get(position_id, "")
        asset = detail.split("/")[0]
        trades.append(
            Trade(
                asset=asset,
                isin=isin,
                broker="eToro",
                action="Acquisto",
                currency_hint="USD",
                cash_currency="USD",
                date=opened_at,
                price=amount / units if units else ZERO,
                quantity=units,
                quantity_diff=units,
                total_spend=-amount,
                fees=ZERO,
                tax=ZERO,
                grand_total=-amount,
                grand_total_present=True,
                source="etoro",
            )
        )

    return sorted(trades, key=lambda item: (item.date, item.asset, item.action))


def read_dividends() -> list[Dividend]:
    dividends: list[Dividend] = []
    tr_file = latest_trade_republic_export()
    fineco_file = latest_fineco_export()
    etoro_file = latest_etoro_export()
    if tr_file:
        dividends.extend(read_trade_republic_dividends(tr_file))
    if fineco_file:
        dividends.extend(read_fineco_dividends(fineco_file))
    if etoro_file:
        dividends.extend(read_etoro_dividends(etoro_file))
    return sorted(dividends, key=lambda item: (item.date, item.broker, item.asset))


def read_portfolio_dividends(person: str = PRIMARY_PORTFOLIO_ID) -> list[Dividend]:
    return read_ledger_dividends(person)


def read_trade_republic_dividends(path: Path) -> list[Dividend]:
    dividends: list[Dividend] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if row.get("type") != "DIVIDEND":
                continue
            dividends.append(
                Dividend(
                    broker="Trade Republic",
                    asset=(row.get("name") or row.get("symbol") or "Unknown").strip(),
                    isin=(row.get("symbol") or "").strip().upper(),
                    date=datetime.strptime(row["date"], "%Y-%m-%d").date(),
                    amount_eur=parse_decimal(row.get("amount")),
                    tax_eur=abs(parse_decimal(row.get("tax"))),
                )
            )
    return dividends


def read_fineco_dividends(path: Path) -> list[Dividend]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Movimenti Dossier Titoli"]
    dividends: list[Dividend] = []
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not row or row[2] != "Dividendo":
            continue
        net_amount = fineco_decimal(row[10])
        dividends.append(
            Dividend(
                broker="Fineco",
                asset=str(row[3]).strip(),
                isin=str(row[4] or "").strip().upper(),
                date=parse_fineco_date(row[0]),
                amount_eur=net_amount,
                tax_eur=fineco_dividend_tax_from_net(net_amount),
            )
        )
    return dividends


def read_etoro_dividends(path: Path) -> list[Dividend]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Dividendi"]
    dividends: list[Dividend] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        dividends.append(
            Dividend(
                broker="eToro",
                asset=str(row[1] or "").strip(),
                isin=str(row[13] or "").strip().upper(),
                date=datetime.strptime(str(row[0]), "%d/%m/%Y").date(),
                amount_eur=fineco_decimal(row[7]),
                tax_eur=fineco_decimal(row[10]),
            )
        )
    return dividends


def summarize_dividends(dividends: list[Dividend]) -> dict[str, Any]:
    by_asset: dict[str, Decimal] = {}
    by_broker: dict[str, Decimal] = {}
    total = ZERO
    tax = ZERO
    gross_total = ZERO
    rows = []
    for dividend in dividends:
        gross = dividend.amount_eur + dividend.tax_eur
        total += dividend.amount_eur
        tax += dividend.tax_eur
        gross_total += gross
        by_asset[dividend.asset] = by_asset.get(dividend.asset, ZERO) + dividend.amount_eur
        by_broker[dividend.broker] = by_broker.get(dividend.broker, ZERO) + dividend.amount_eur
        rows.append(
            {
                "date": dividend.date.isoformat(),
                "broker": dividend.broker,
                "asset": dividend.asset,
                "isin": dividend.isin,
                "amount_eur": money(dividend.amount_eur),
                "tax_eur": money(dividend.tax_eur),
                "gross_eur": money(gross),
            }
        )
    return {
        "rows": sorted(rows, key=lambda item: item["date"], reverse=True),
        "total_eur": money(total),
        "tax_eur": money(tax),
        "gross_eur": money(gross_total),
        "count": len(dividends),
        "by_asset": [{"asset": key, "amount_eur": money(value)} for key, value in sorted(by_asset.items())],
        "by_broker": [{"broker": key, "amount_eur": money(value)} for key, value in sorted(by_broker.items())],
    }


def read_cash_interests(person: str = PRIMARY_PORTFOLIO_ID) -> list[dict[str, Any]]:
    # Statement imports are normalized once; runtime analytics read the ledger only.
    return read_ledger_interests(person)


def read_cash_interests_from_raw_files(person: str = PRIMARY_PORTFOLIO_ID) -> list[dict[str, Any]]:
    """Legacy parser kept only for import/migration compatibility."""

    interests: list[dict[str, Any]] = []
    person = person.lower()
    
    # 1. Load Trade Republic Interest
    tr_file = None
    if person == PRIMARY_PORTFOLIO_ID:
        tr_file = latest_trade_republic_export()
    else:
        tr_file = latest_family_trade_republic_export(person)
        
    if tr_file and tr_file.exists():
        try:
            with tr_file.open(newline="", encoding="utf-8-sig") as handle:
                for row in csv.DictReader(handle):
                    if row.get("type") == "INTEREST_PAYMENT":
                        amount = parse_decimal(row.get("amount"))
                        tax = abs(parse_decimal(row.get("tax") or "0"))
                        interests.append({
                            "broker": "Trade Republic",
                            "date": datetime.strptime(row["date"], "%Y-%m-%d").date(),
                            "net_eur": amount - tax,
                            "tax_eur": tax,
                            "gross_eur": amount,
                            "description": row.get("description", "Cash Interest").strip() or "Cash Interest"
                        })
        except Exception:
            pass
                    
    # 2. Load primary-portfolio BBVA interest.
    if person == PRIMARY_PORTFOLIO_ID:
        bbva_files = bbva_statement_files()
        if bbva_files:
            bbva_file = bbva_files[-1]
            import shutil
            import tempfile
            temp_dir = tempfile.gettempdir()
            temp_xlsx = Path(temp_dir) / "bbva_temp.xlsx"
            try:
                shutil.copy(bbva_file, temp_xlsx)
                import openpyxl
                wb = openpyxl.load_workbook(temp_xlsx, read_only=True)
                sheet = wb.active
                for row in sheet.iter_rows(values_only=True):
                    if len(row) < 7:
                        continue
                    val_date_str, op_date_str, causale, mov, ben, imp = row[1:7]
                    if causale and "INTERESSI" in str(causale).upper() and imp:
                        try:
                            op_date = datetime.strptime(str(op_date_str).strip(), "%d/%m/%Y").date()
                        except Exception:
                            op_date = date.today()
                        try:
                            clean_imp = str(imp).replace(" EUR", "").replace(",", ".").strip()
                            amount = Decimal(clean_imp)
                        except Exception:
                            amount = ZERO
                        
                        net_val = amount
                        gross_val = net_val / BBVA_INTEREST_NET_RATE if net_val > ZERO else ZERO
                        tax_val = gross_val - net_val
                        
                        interests.append({
                            "broker": "BBVA",
                            "date": op_date,
                            "net_eur": net_val,
                            "tax_eur": tax_val,
                            "gross_eur": gross_val,
                            "description": str(causale).strip()
                        })
            except Exception:
                pass
            finally:
                if temp_xlsx.exists():
                    try:
                        temp_xlsx.unlink()
                    except Exception:
                        pass
                    
    return sorted(interests, key=lambda x: x["date"], reverse=True)


def is_revolut_internal_movement(row: dict[str, str]) -> bool:
    description = (row.get("Descrizione") or "").strip().lower()
    movement_type = (row.get("Tipo") or "").strip().lower()
    if not description and not movement_type:
        return False
    if movement_type == "cambia valuta" or description.startswith("conversione in "):
        return True
    if "balance migration to another region or legal entity" in description:
        return True
    if description == "closing transaction":
        return True
    if description.startswith("prelievo da pocket"):
        return True
    if re.fullmatch(r"to [a-z]{3}", description):
        return True
    if description.startswith("accredita ") and " da " in description:
        return True
    return False


def read_revolut_cash_events(paths: list[Path] | None = None) -> list[dict[str, Any]]:
    source_paths = paths if paths is not None else revolut_statement_files()
    events: list[dict[str, Any]] = []

    for path in source_paths:
        try:
            handle = path.open(newline="", encoding="utf-8-sig")
        except OSError:
            continue
        with handle:
            for row in csv.DictReader(handle):
                if row.get("State") != "COMPLETATO":
                    continue
                event_time = parse_revolut_datetime(row.get("Data di completamento")) or parse_revolut_datetime(row.get("Data di inizio"))
                if event_time is None:
                    continue
                amount = parse_decimal(row.get("Importo"))
                fee = abs(parse_decimal(row.get("Costo")))
                cash_change_native = amount - fee
                if cash_change_native == ZERO:
                    continue
                currency = normalize_currency_code(row.get("Valuta") or "EUR")
                contrib_change_native = ZERO if is_revolut_internal_movement(row) else cash_change_native
                events.append(
                    {
                        "datetime": event_time,
                        "date": event_time.date(),
                        "cash_change": cash_change_native,
                        "contrib_change": contrib_change_native,
                        "currency": currency,
                    }
                )

    return sorted(events, key=lambda item: item["datetime"])


def add_revolut_fx_histories(
    revolut_events: list[dict[str, Any]],
    start: date,
    end: date,
    fx_histories: dict[str, dict[str, Any]],
    refresh: bool = False,
) -> None:
    event_dates = [event["date"] for event in revolut_events if event.get("date")]
    if not event_dates:
        return
    history_start = min(min(event_dates), start) - timedelta(days=10)
    history_end = max(max(event_dates), end)
    currencies = {fx_base_currency(event.get("currency") or "EUR") for event in revolut_events}
    for currency in sorted(currencies):
        if not currency or currency == "EUR":
            continue
        symbol, _ = fx_symbol_for(currency)
        if symbol not in fx_histories:
            fx_histories[symbol] = fetch_history(symbol, history_start, history_end, refresh=refresh)


def revolut_cash_balance_eur(
    revolut_events: list[dict[str, Any]],
    target: date,
    fx_histories: dict[str, dict[str, Any]],
    live_fx_prices: dict[str, dict[str, Any]] | None = None,
) -> Decimal:
    balances: dict[str, Decimal] = {}
    for event in revolut_events:
        event_date = event.get("date")
        if not event_date or event_date > target:
            continue
        currency = normalize_currency_code(event.get("currency") or "EUR")
        balances[currency] = balances.get(currency, ZERO) + Decimal(str(event.get("cash_change") or 0))

    total = ZERO
    for currency, balance in balances.items():
        if balance == ZERO:
            continue
        rate = historical_fx_rate(currency, fx_histories, target, live_fx_prices=live_fx_prices)
        if rate is None:
            if fx_base_currency(currency) != "EUR":
                continue
            rate = 1.0
        total += balance * Decimal(str(rate))
    return total


def revolut_contributions_eur(
    revolut_events: list[dict[str, Any]],
    target: date,
    fx_histories: dict[str, dict[str, Any]],
) -> Decimal:
    total = ZERO
    for event in revolut_events:
        event_date = event.get("date")
        if not event_date or event_date > target:
            continue
        contrib_change = Decimal(str(event.get("contrib_change") or 0))
        if contrib_change == ZERO:
            continue
        currency = normalize_currency_code(event.get("currency") or "EUR")
        rate = historical_fx_rate(currency, fx_histories, event_date)
        if rate is None:
            if fx_base_currency(currency) != "EUR":
                continue
            rate = 1.0
        total += contrib_change * Decimal(str(rate))
    return total


def read_revolut_cash_history(paths: list[Path] | None = None) -> list[tuple[date, Decimal, Decimal]]:
    events = read_revolut_cash_events(paths)
    if not events:
        return []
    fx_histories: dict[str, dict[str, Any]] = {}
    event_dates = [event["date"] for event in events]
    add_revolut_fx_histories(events, min(event_dates), max(event_dates), fx_histories)
    cash_history: list[tuple[date, Decimal, Decimal]] = []
    for event in events:
        rate = historical_fx_rate(event["currency"], fx_histories, event["date"])
        if rate is None:
            if fx_base_currency(event["currency"]) != "EUR":
                continue
            rate = 1.0
        rate_dec = Decimal(str(rate))
        cash_history.append((event["date"], event["cash_change"] * rate_dec, event["contrib_change"] * rate_dec))
    return cash_history


EXPENSE_CATEGORIES = {
    "Groceries",
    "Dining",
    "Transport & Fuel",
    "Travel & Lodging",
    "Shopping",
    "Subscriptions & Digital",
    "Housing & Utilities",
    "Health & Personal Care",
    "Entertainment",
    "Education",
    "Cash Withdrawals",
    "Fees",
    "Personal Transfers",
    "Investments",
    "Credits",
    "Income",
    "Uncategorized",
}
TR_EXPENSE_TYPES = {"CARD_TRANSACTION", "CARD_TRANSACTION_INTERNATIONAL", "CARD_ORDERING_FEE"}
SELF_TRANSFER_NAMES = SETTINGS.self_transfer_names


def normalize_expense_source(value: str | None) -> str:
    raw = (value or "all").strip().lower().replace("-", "_").replace(" ", "_")
    if raw in {"trade_republic", "traderepublic", "tr"}:
        return "trade_republic"
    if raw in {"revolut", "rv"}:
        return "revolut"
    if raw in {"intesa", "intesasanpaolo", "intesa_sanpaolo"}:
        return "intesa"
    return raw or "all"


def expense_source_label(source: str) -> str:
    if source == "trade_republic":
        return "Trade Republic"
    if source == "revolut":
        return "Revolut"
    if source == "intesa":
        return "Intesa"
    return source.replace("_", " ").title()


def normalize_expense_category(value: str | None) -> str:
    category = (value or "").strip()
    return category if category in EXPENSE_CATEGORIES else "Uncategorized"


def expense_default_category(flow_kind: str) -> str:
    if flow_kind == "income":
        return "Income"
    if flow_kind == "investment":
        return "Investments"
    if flow_kind == "credit":
        return "Credits"
    if flow_kind == "personal_transfer":
        return "Personal Transfers"
    if flow_kind == "cash_withdrawal":
        return "Cash Withdrawals"
    if flow_kind == "fee":
        return "Fees"
    return "Uncategorized"


def is_self_giroconto_expense(source_category: str, merchant: str, description: str, flow_kind: str) -> bool:
    combined = f"{source_category} {merchant} {description}".lower()
    if re.search(r"\brevolut\*\*\d+\*", combined):
        return True
    if not any(name in combined for name in SELF_TRANSFER_NAMES):
        return False
    category = source_category.lower()
    if "bonific" in category or "giroconto" in category:
        return True
    return flow_kind in {"personal_transfer", "income"}


def read_expense_category_rules(path: Path = EXPENSE_RULES_CSV) -> list[ExpenseRule]:
    if not path.exists():
        return []
    rules: list[ExpenseRule] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for index, row in enumerate(csv.DictReader(handle), start=1):
            enabled = (row.get("enabled") or "1").strip().lower()
            if enabled in {"0", "false", "no", "off"}:
                continue
            try:
                priority = int((row.get("priority") or index).strip())
            except ValueError:
                priority = index
            rules.append(
                ExpenseRule(
                    priority=priority,
                    source=normalize_expense_source(row.get("source") or "all"),
                    match_field=(row.get("match_field") or "description").strip().lower(),
                    match_type=(row.get("match_type") or "contains").strip().lower(),
                    pattern=(row.get("pattern") or "").strip(),
                    category=normalize_expense_category(row.get("category")),
                    subcategory=(row.get("subcategory") or "").strip(),
                    merchant=(row.get("merchant") or "").strip(),
                )
            )
    return sorted(rules, key=lambda item: item.priority)


def expense_match_text(event: dict[str, Any], field: str) -> str:
    if field == "all":
        parts = [
            event.get("source", ""),
            event.get("merchant", ""),
            event.get("description", ""),
            event.get("flow_kind", ""),
            event.get("category", ""),
            event.get("subcategory", ""),
            event.get("source_category", ""),
            str(event.get("native_amount") or ""),
            str(event.get("amount_eur") or ""),
        ]
        return " ".join(str(part or "") for part in parts)
    return str(event.get(field) or "")


def expense_rule_matches(event: dict[str, Any], rule: ExpenseRule) -> bool:
    if rule.source not in {"", "all", normalize_expense_source(str(event.get("source") or ""))}:
        return False
    pattern = rule.pattern
    if not pattern:
        return False
    value = expense_match_text(event, rule.match_field)
    if rule.match_type == "exact":
        return value.strip().lower() == pattern.strip().lower()
    if rule.match_type == "regex":
        try:
            return re.search(pattern, value, flags=re.IGNORECASE) is not None
        except re.error:
            return False
    return pattern.lower() in value.lower()


def classify_expense_event(event: dict[str, Any], rules: list[ExpenseRule]) -> dict[str, Any]:
    classified = dict(event)
    classified["category"] = normalize_expense_category(classified.get("category"))
    classified.setdefault("subcategory", "")
    classified.setdefault("confidence", 0.35)
    for rule in rules:
        if not expense_rule_matches(classified, rule):
            continue
        classified["category"] = rule.category
        classified["subcategory"] = rule.subcategory
        if rule.merchant:
            classified["merchant"] = rule.merchant
        classified["confidence"] = 0.95
        break
    if classified["category"] == "Credits":
        classified["flow_kind"] = "credit"
    return classified


def make_expense_event(
    *,
    event_date: date,
    source: str,
    merchant: str,
    description: str,
    flow_kind: str,
    amount_eur: Decimal,
    currency: str,
    native_amount: Decimal,
    rules: list[ExpenseRule],
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    event = {
        "date": event_date,
        "source": normalize_expense_source(source),
        "merchant": (merchant or description or "Unknown").strip() or "Unknown",
        "description": (description or merchant or "").strip(),
        "flow_kind": flow_kind,
        "category": expense_default_category(flow_kind),
        "subcategory": "",
        "amount_eur": amount_eur,
        "currency": normalize_currency_code(currency),
        "native_amount": native_amount,
        "confidence": 0.65 if flow_kind != "spend" else 0.35,
    }
    if extra:
        event.update(extra)
    return classify_expense_event(event, rules)


def fineco_bank_flow_kind(source_category: str, description: str, amount: Decimal) -> str:
    """Classify a signed Fineco current-account movement for expense analytics."""

    combined = f"{source_category} {description}".casefold()
    if any(term in combined for term in ("bonific", "girocont", "trasferiment")):
        return "personal_transfer"
    if any(
        term in combined
        for term in (
            "compravendita titoli",
            "acquisto titoli",
            "vendita titoli",
            "sottoscrizione titoli",
            "rimborso titoli",
        )
    ):
        return "investment"
    if "preliev" in combined:
        return "cash_withdrawal"
    if amount > ZERO:
        return "income"
    if any(
        term in combined
        for term in ("commission", "canone", "imposta", "riten", "interessi passivi", "spese", "bollo")
    ):
        return "fee"
    return "spend"


def read_fineco_bank_expense_events(
    path: Path,
    rules: list[ExpenseRule] | None = None,
) -> list[dict[str, Any]]:
    """Convert a native Fineco current-account XLSX into classified ledger events."""

    active_rules = rules if rules is not None else read_expense_category_rules()
    events: list[dict[str, Any]] = []
    for row in read_fineco_bank_movements(path):
        amount = Decimal(str(row["amount"]))
        source_category = str(row.get("source_category") or "").strip()
        description = str(row.get("description") or source_category).strip()
        flow_kind = fineco_bank_flow_kind(source_category, description, amount)
        merchant = source_category or description or "Fineco"
        if is_self_giroconto_expense(source_category, merchant, description, flow_kind):
            continue
        events.append(
            make_expense_event(
                event_date=row["date"],
                source="fineco",
                merchant=merchant,
                description=description,
                flow_kind=flow_kind,
                amount_eur=abs(amount),
                currency="EUR",
                native_amount=amount,
                rules=active_rules,
                extra={
                    "source_category": source_category,
                    "value_date": row.get("value_date"),
                    "source_state": row.get("state"),
                },
            )
        )
    return sorted(events, key=lambda item: (item["date"], item["source"], item["merchant"], item["native_amount"]))


def revolut_expense_flow_kind(row: dict[str, str], cash_change_native: Decimal) -> str:
    movement_type = (row.get("Tipo") or "").strip().lower()
    description = (row.get("Descrizione") or "").strip().lower()
    if is_revolut_internal_movement(row):
        return ""
    if "al conto di investimento" in description:
        return "investment"
    if movement_type == "pagamento con carta":
        return "income" if cash_change_native > ZERO else "spend"
    if movement_type in {"rimborso su carta", "cashback"}:
        return "income"
    if movement_type == "prelievo" and cash_change_native < ZERO:
        return "cash_withdrawal"
    if movement_type == "commissione" and cash_change_native < ZERO:
        return "fee"
    if movement_type == "pagamento" and cash_change_native < ZERO:
        return "personal_transfer"
    return ""


def expense_amount_to_eur(
    signed_native_amount: Decimal,
    currency: str,
    event_date: date,
    fx_histories: dict[str, dict[str, Any]],
) -> Decimal | None:
    currency = normalize_currency_code(currency)
    rate = historical_fx_rate(currency, fx_histories, event_date)
    if rate is None:
        if fx_base_currency(currency) != "EUR":
            return None
        rate = 1.0
    return abs(signed_native_amount) * Decimal(str(rate))


def read_revolut_expense_events(
    paths: list[Path] | None = None,
    rules: list[ExpenseRule] | None = None,
    fx_histories: dict[str, dict[str, Any]] | None = None,
    refresh: bool = False,
) -> list[dict[str, Any]]:
    source_paths = paths if paths is not None else revolut_statement_files()
    active_rules = rules if rules is not None else read_expense_category_rules()
    raw_events: list[dict[str, Any]] = []

    for path in source_paths:
        try:
            handle = path.open(newline="", encoding="utf-8-sig")
        except OSError:
            continue
        with handle:
            for row in csv.DictReader(handle):
                if row.get("State") != "COMPLETATO":
                    continue
                event_time = parse_revolut_datetime(row.get("Data di completamento")) or parse_revolut_datetime(row.get("Data di inizio"))
                if event_time is None:
                    continue
                amount = parse_decimal(row.get("Importo"))
                fee = abs(parse_decimal(row.get("Costo")))
                cash_change_native = amount - fee
                if cash_change_native == ZERO:
                    continue
                flow_kind = revolut_expense_flow_kind(row, cash_change_native)
                if not flow_kind:
                    continue
                merchant = (row.get("Descrizione") or "").strip() or (row.get("Tipo") or "Revolut").strip()
                description = (row.get("Descrizione") or row.get("Tipo") or "").strip()
                if is_self_giroconto_expense("", merchant, description, flow_kind):
                    continue
                raw_events.append(
                    {
                        "date": event_time.date(),
                        "datetime": event_time,
                        "source": "revolut",
                        "merchant": merchant,
                        "description": description,
                        "flow_kind": flow_kind,
                        "currency": normalize_currency_code(row.get("Valuta") or "EUR"),
                        "native_amount": cash_change_native,
                    }
                )

    if not raw_events:
        return []

    histories = fx_histories if fx_histories is not None else {}
    event_dates = [event["date"] for event in raw_events]
    add_revolut_fx_histories(raw_events, min(event_dates), max(event_dates), histories, refresh=refresh)

    events: list[dict[str, Any]] = []
    for raw_event in raw_events:
        amount_eur = expense_amount_to_eur(
            Decimal(str(raw_event["native_amount"])),
            raw_event["currency"],
            raw_event["date"],
            histories,
        )
        if amount_eur is None:
            continue
        events.append(
            make_expense_event(
                event_date=raw_event["date"],
                source=raw_event["source"],
                merchant=raw_event["merchant"],
                description=raw_event["description"],
                flow_kind=raw_event["flow_kind"],
                amount_eur=amount_eur,
                currency=raw_event["currency"],
                native_amount=raw_event["native_amount"],
                rules=active_rules,
            )
        )
    return sorted(events, key=lambda item: (item["date"], item["source"], item["merchant"]))


def read_trade_republic_expense_events(path: Path, rules: list[ExpenseRule] | None = None) -> list[dict[str, Any]]:
    active_rules = rules if rules is not None else read_expense_category_rules()
    events: list[dict[str, Any]] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            t_type = (row.get("type") or "").strip()
            if t_type not in TR_EXPENSE_TYPES:
                continue
            event_date_raw = row.get("date")
            if not event_date_raw:
                continue
            event_date = datetime.strptime(event_date_raw, "%Y-%m-%d").date()
            amount = parse_decimal(row.get("amount"))
            fee = parse_decimal(row.get("fee"))
            tax = parse_decimal(row.get("tax"))
            cash_change = amount + fee + tax
            if cash_change == ZERO:
                continue
            original_currency = normalize_currency_code(row.get("original_currency") or "")
            native_currency = original_currency if original_currency not in {"", "EUR"} else normalize_currency_code(row.get("currency") or "EUR")
            original_amount = parse_decimal(row.get("original_amount"))
            native_amount = original_amount if original_amount != ZERO and native_currency != "EUR" else cash_change
            merchant = (row.get("name") or row.get("description") or row.get("type") or "Trade Republic").strip()
            description = (row.get("description") or merchant or t_type).strip()
            if t_type == "CARD_ORDERING_FEE":
                flow_kind = "fee"
                merchant = merchant if merchant != t_type else "Trade Republic Card"
            else:
                flow_kind = "income" if cash_change > ZERO else "spend"
            if is_self_giroconto_expense("", merchant, description, flow_kind):
                continue
            events.append(
                make_expense_event(
                    event_date=event_date,
                    source="trade_republic",
                    merchant=merchant,
                    description=description,
                    flow_kind=flow_kind,
                    amount_eur=abs(cash_change),
                    currency=native_currency,
                    native_amount=native_amount,
                    rules=active_rules,
                )
            )
    return sorted(events, key=lambda item: (item["date"], item["source"], item["merchant"]))


def intesa_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def parse_intesa_expense_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = intesa_cell_text(value)
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def intesa_expense_flow_kind(source_category: str, merchant: str, description: str, amount: Decimal) -> str:
    if amount > ZERO:
        return "income"
    category = source_category.strip().lower()
    combined = f"{merchant} {description}".lower()
    if "preliev" in category or "preliev" in combined:
        return "cash_withdrawal"
    if "imposte" in category or "commission" in category or "commission" in combined or "interessi debitori" in combined:
        return "fee"
    if "bonifici in uscita" in category or "giroconto in uscita" in category:
        return "personal_transfer"
    return "spend"


def read_intesa_expense_events(
    path: Path,
    rules: list[ExpenseRule] | None = None,
    fx_histories: dict[str, dict[str, Any]] | None = None,
    refresh: bool = False,
) -> list[dict[str, Any]]:
    active_rules = rules if rules is not None else read_expense_category_rules()
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Print area cannot be set*", category=UserWarning)
        workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Lista Operazione"] if "Lista Operazione" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    header_row_number: int | None = None
    headers: list[str] = []
    required_headers = {"Data", "Operazione", "Importo"}
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        candidate = [intesa_cell_text(value) for value in row]
        if required_headers.issubset(set(candidate)):
            header_row_number = row_number
            headers = candidate
            break
    if header_row_number is None:
        return []

    raw_events: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        if not row or not any(intesa_cell_text(value) for value in row):
            continue
        record = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        accounted = intesa_cell_text(record.get("Contabilizzazione")).upper()
        if accounted and not (accounted.startswith("S") or accounted in {"YES", "Y", "TRUE", "1"}):
            continue
        event_date = parse_intesa_expense_date(record.get("Data"))
        if event_date is None:
            continue
        amount = fineco_decimal(record.get("Importo"))
        if amount == ZERO:
            continue
        source_category = intesa_cell_text(record.get("Categoria"))
        merchant = intesa_cell_text(record.get("Operazione"))
        description = intesa_cell_text(record.get("Dettagli")) or merchant
        currency = normalize_currency_code(record.get("Valuta") or "EUR")
        flow_kind = intesa_expense_flow_kind(source_category, merchant, description, amount)
        if is_self_giroconto_expense(source_category, merchant, description, flow_kind):
            continue
        raw_events.append(
            {
                "date": event_date,
                "source": "intesa",
                "merchant": merchant or description or "Intesa",
                "description": description,
                "flow_kind": flow_kind,
                "currency": currency,
                "native_amount": amount,
                "source_category": source_category,
            }
        )

    if not raw_events:
        return []

    histories = fx_histories if fx_histories is not None else {}
    event_dates = [event["date"] for event in raw_events]
    add_revolut_fx_histories(raw_events, min(event_dates), max(event_dates), histories, refresh=refresh)

    events: list[dict[str, Any]] = []
    for raw_event in raw_events:
        amount_eur = expense_amount_to_eur(
            Decimal(str(raw_event["native_amount"])),
            raw_event["currency"],
            raw_event["date"],
            histories,
        )
        if amount_eur is None:
            continue
        events.append(
            make_expense_event(
                event_date=raw_event["date"],
                source=raw_event["source"],
                merchant=raw_event["merchant"],
                description=raw_event["description"],
                flow_kind=raw_event["flow_kind"],
                amount_eur=amount_eur,
                currency=raw_event["currency"],
                native_amount=raw_event["native_amount"],
                rules=active_rules,
                extra={"source_category": raw_event["source_category"]},
            )
        )
    return sorted(events, key=lambda item: (item["date"], item["source"], item["merchant"]))


def read_bbva_expense_events(
    paths: list[Path] | None = None,
    rules: list[ExpenseRule] | None = None,
    fx_histories: dict[str, dict[str, Any]] | None = None,
    refresh: bool = False,
) -> list[dict[str, Any]]:
    active_rules = rules if rules is not None else read_expense_category_rules()
    bbva_files = paths if paths is not None else bbva_statement_files()
    if not bbva_files:
        return []

    raw_events: list[dict[str, Any]] = []
    seen_keys = set()

    import os
    import shutil
    import tempfile
    import threading

    for bbva_file in bbva_files:
        temp_xlsx = Path(tempfile.gettempdir()) / f"bbva_temp_{os.getpid()}_{threading.get_ident()}.xlsx"
        try:
            shutil.copy(bbva_file, temp_xlsx)
            wb = load_workbook(temp_xlsx, read_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if len(row) < 7:
                    continue
                # columns B to G (index 1 to 6)
                _data_valuta = row[1]
                op_date_str = row[2]
                causale = row[3]
                movimento = row[4]
                beneficiario = row[5]
                imp = row[6]

                if causale and causale != "Causale" and imp:
                    try:
                        op_date = datetime.strptime(str(op_date_str).strip(), "%d/%m/%Y").date()
                    except Exception:
                        continue
                    try:
                        clean_imp = str(imp).replace(" EUR", "").replace(",", ".").strip()
                        amount = Decimal(clean_imp)
                    except Exception:
                        continue

                    if amount == ZERO:
                        continue

                    # Clean counterparty / beneficiary
                    benef = ""
                    if beneficiario:
                        benef = str(beneficiario).split("\n")[0].strip()
                        if benef == "-":
                            benef = ""

                    # Dedup check
                    key = (op_date, str(causale), str(movimento), benef, amount)
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)

                    flow_kind = "spend"
                    if amount > 0:
                        flow_kind = "income"
                    else:
                        causale_upper = str(causale).upper()
                        if "COMMISS" in causale_upper or "SPESE" in causale_upper:
                            flow_kind = "fee"

                    merchant = benef if benef else (str(movimento).strip() if movimento else (str(causale).strip() if causale else "BBVA"))
                    description = str(movimento).strip() if movimento else (str(causale).strip() if causale else "")
                    source_category = str(causale).strip() if causale else ""

                    if is_self_giroconto_expense(source_category, merchant, description, flow_kind):
                        continue

                    raw_events.append(
                        {
                            "date": op_date,
                            "source": "bbva",
                            "merchant": merchant,
                            "description": description,
                            "flow_kind": flow_kind,
                            "currency": "EUR",
                            "native_amount": amount,
                            "source_category": source_category,
                        }
                    )
        except Exception:
            pass
        finally:
            if temp_xlsx.exists():
                try:
                    temp_xlsx.unlink()
                except OSError:
                    pass

    if not raw_events:
        return []

    histories = fx_histories if fx_histories is not None else {}
    events: list[dict[str, Any]] = []
    for raw_event in raw_events:
        amount_eur = expense_amount_to_eur(
            Decimal(str(raw_event["native_amount"])),
            raw_event["currency"],
            raw_event["date"],
            histories,
        )
        if amount_eur is None:
            continue
        events.append(
            make_expense_event(
                event_date=raw_event["date"],
                source=raw_event["source"],
                merchant=raw_event["merchant"],
                description=raw_event["description"],
                flow_kind=raw_event["flow_kind"],
                amount_eur=amount_eur,
                currency=raw_event["currency"],
                native_amount=raw_event["native_amount"],
                rules=active_rules,
                extra={"source_category": raw_event["source_category"]},
            )
        )
    return sorted(events, key=lambda item: (item["date"], item["source"], item["merchant"]))


def read_expense_events(person: str = PRIMARY_PORTFOLIO_ID, refresh: bool = False) -> list[dict[str, Any]]:
    if (person or PRIMARY_PORTFOLIO_ID).lower() != PRIMARY_PORTFOLIO_ID:
        return []
    rules = read_expense_category_rules()
    events: list[dict[str, Any]] = []
    tr_file = latest_trade_republic_export()
    if tr_file:
        try:
            events.extend(read_trade_republic_expense_events(tr_file, rules=rules))
        except Exception:
            pass
    intesa_file = latest_intesa_operations_export()
    if intesa_file:
        try:
            events.extend(read_intesa_expense_events(intesa_file, rules=rules, refresh=refresh))
        except Exception:
            pass
    try:
        events.extend(read_revolut_expense_events(rules=rules, refresh=refresh))
    except Exception:
        pass
    try:
        events.extend(read_bbva_expense_events(rules=rules, refresh=refresh))
    except Exception:
        pass
    return sorted(events, key=lambda item: (item["date"], item["source"], item["merchant"]))


def expense_row_kind(row: dict[str, Any]) -> str:
    flow_kind = str(row.get("flow_kind") or "")
    category = str(row.get("category") or "")
    if flow_kind == "income" or category == "Income":
        return "income"
    if flow_kind == "credit" or category == "Credits":
        return "credits"
    if flow_kind == "investment" or category == "Investments":
        return "investments"
    if flow_kind == "personal_transfer" or category == "Personal Transfers":
        return "transfers"
    return "spend"


def expense_event_payload(event: dict[str, Any]) -> dict[str, Any]:
    event_date = event.get("date")
    if isinstance(event_date, date):
        date_value = event_date.isoformat()
    else:
        date_value = str(event_date or "")
    return {
        "date": date_value,
        "source": event.get("source") or "",
        "source_label": expense_source_label(str(event.get("source") or "")),
        "merchant": event.get("merchant") or "Unknown",
        "description": event.get("description") or "",
        "flow_kind": event.get("flow_kind") or "spend",
        "category": normalize_expense_category(str(event.get("category") or "")),
        "subcategory": event.get("subcategory") or "",
        "amount_eur": money(Decimal(str(event.get("amount_eur") or 0))),
        "currency": normalize_currency_code(event.get("currency") or "EUR"),
        "native_amount": decimal_to_float(Decimal(str(event.get("native_amount") or 0))),
        "confidence": round(float(event.get("confidence") or 0), 2),
    }


def summarize_expense_events(events: list[dict[str, Any]]) -> dict[str, Any]:
    rows = [expense_event_payload(event) for event in sorted(events, key=lambda item: item["date"], reverse=True)]
    spend = ZERO
    income = ZERO
    transfers = ZERO
    investments = ZERO
    credits = ZERO
    by_category: dict[str, dict[str, Any]] = {}
    by_source: dict[str, dict[str, Any]] = {}
    by_month: dict[str, dict[str, Decimal]] = {}
    by_merchant: dict[str, dict[str, Any]] = {}
    credit_rows: list[dict[str, Any]] = []

    for row in rows:
        amount = Decimal(str(row.get("amount_eur") or 0))
        kind = expense_row_kind(row)
        if kind == "income":
            income += amount
        elif kind == "credits":
            credits += amount
            credit_rows.append(row)
        elif kind == "investments":
            investments += amount
        elif kind == "transfers":
            transfers += amount
        else:
            spend += amount

        category = normalize_expense_category(str(row.get("category") or ""))
        cat_bucket = by_category.setdefault(category, {"category": category, "amount": ZERO, "count": 0})
        cat_bucket["amount"] += amount
        cat_bucket["count"] += 1

        source = row.get("source_label") or expense_source_label(str(row.get("source") or ""))
        source_bucket = by_source.setdefault(str(source), {"source": str(source), "spend": ZERO, "income": ZERO, "transfers": ZERO, "investments": ZERO, "credits": ZERO, "count": 0})
        source_bucket[kind if kind != "spend" else "spend"] += amount
        source_bucket["count"] += 1

        month = str(row.get("date") or "")[:7]
        if month:
            month_bucket = by_month.setdefault(month, {"month": month, "spend": ZERO, "income": ZERO, "transfers": ZERO, "investments": ZERO, "credits": ZERO, "count": 0})
            month_bucket[kind if kind != "spend" else "spend"] += amount
            month_bucket["count"] += 1

        if kind not in {"income", "credits"}:
            merchant = str(row.get("merchant") or "Unknown")
            merch_bucket = by_merchant.setdefault(merchant, {"merchant": merchant, "category": category, "amount": ZERO, "count": 0})
            merch_bucket["amount"] += amount
            merch_bucket["count"] += 1

    net_outflow = spend + investments - income
    total_outflow = spend + investments
    category_rows = [
        {
            "category": values["category"],
            "amount_eur": money(values["amount"]),
            "count": values["count"],
            "share_pct": float((values["amount"] / total_outflow * Decimal("100")).quantize(Decimal("0.01"))) if total_outflow and values["category"] not in {"Income", "Credits", "Personal Transfers"} else None,
        }
        for values in by_category.values()
        if values["category"] not in {"Income", "Credits", "Personal Transfers"}
    ]
    source_rows = [
        {
            "source": values["source"],
            "spend_eur": money(values["spend"]),
            "income_eur": money(values["income"]),
            "transfers_eur": money(values["transfers"]),
            "investments_eur": money(values["investments"]),
            "credits_eur": money(values["credits"]),
            "net_outflow_eur": money(values["spend"] + values["investments"] - values["income"]),
            "count": values["count"],
        }
        for values in by_source.values()
    ]
    month_rows = [
        {
            "month": values["month"],
            "spend_eur": money(values["spend"]),
            "income_eur": money(values["income"]),
            "transfers_eur": money(values["transfers"]),
            "investments_eur": money(values["investments"]),
            "credits_eur": money(values["credits"]),
            "net_outflow_eur": money(values["spend"] + values["investments"] - values["income"]),
            "count": values["count"],
        }
        for values in by_month.values()
    ]
    merchant_rows = [
        {
            "merchant": values["merchant"],
            "category": values["category"],
            "amount_eur": money(values["amount"]),
            "count": values["count"],
        }
        for values in by_merchant.values()
    ]

    return {
        "status": "available" if rows else "empty",
        "summary": {
            "spend_eur": money(spend),
            "income_eur": money(income),
            "transfers_eur": money(transfers),
            "investments_eur": money(investments),
            "credits_eur": money(credits),
            "net_outflow_eur": money(net_outflow),
            "rows_count": len(rows),
        },
        "by_category": sorted(category_rows, key=lambda item: float(item.get("amount_eur") or 0), reverse=True),
        "by_source": sorted(source_rows, key=lambda item: abs(float(item.get("net_outflow_eur") or 0)), reverse=True),
        "by_month": sorted(month_rows, key=lambda item: item["month"]),
        "top_merchants": sorted(merchant_rows, key=lambda item: float(item.get("amount_eur") or 0), reverse=True)[:20],
        "credits": sorted(credit_rows, key=lambda item: item["date"], reverse=True),
        "recent_rows": rows[:30],
        "rows": rows,
        "rules_file": configured_path_label(EXPENSE_RULES_CSV),
    }


def empty_expenses(message: str = "No expense data available for this selection.") -> dict[str, Any]:
    return {
        "status": "empty",
        "message": message,
        "summary": {
            "spend_eur": 0.0,
            "income_eur": 0.0,
            "transfers_eur": 0.0,
            "investments_eur": 0.0,
            "credits_eur": 0.0,
            "net_outflow_eur": 0.0,
            "rows_count": 0,
        },
        "by_category": [],
        "by_source": [],
        "by_month": [],
        "top_merchants": [],
        "credits": [],
        "recent_rows": [],
        "rows": [],
        "rules_file": configured_path_label(EXPENSE_RULES_CSV),
    }


def summarize_cash_interests(interests: list[dict[str, Any]]) -> dict[str, Any]:
    total_net = ZERO
    total_tax = ZERO
    total_gross = ZERO
    
    by_broker_dict = {}
    for item in interests:
        total_net += item["net_eur"]
        total_tax += item["tax_eur"]
        total_gross += item["gross_eur"]
        
        b = item["broker"]
        by_broker_dict.setdefault(b, {"net_eur": ZERO, "tax_eur": ZERO, "gross_eur": ZERO, "count": 0})
        by_broker_dict[b]["net_eur"] += item["net_eur"]
        by_broker_dict[b]["tax_eur"] += item["tax_eur"]
        by_broker_dict[b]["gross_eur"] += item["gross_eur"]
        by_broker_dict[b]["count"] += 1
        
    by_broker = []
    for b, vals in sorted(by_broker_dict.items()):
        by_broker.append({
            "broker": b,
            "net_eur": money(vals["net_eur"]),
            "tax_eur": money(vals["tax_eur"]),
            "gross_eur": money(vals["gross_eur"]),
            "payments_count": vals["count"]
        })
        
    payments = []
    for item in interests:
        payments.append({
            "broker": item["broker"],
            "date": item["date"].isoformat(),
            "net_eur": money(item["net_eur"]),
            "tax_eur": money(item["tax_eur"]),
            "gross_eur": money(item["gross_eur"]),
            "description": item["description"]
        })
        
    return {
        "summary": {
            "total_net_eur": money(total_net),
            "total_tax_eur": money(total_tax),
            "total_gross_eur": money(total_gross),
            "payments_count": len(interests),
        },
        "by_broker": by_broker,
        "payments": payments
    }


def summarize_net_contributions(trades: list[Trade]) -> dict[str, Any]:
    by_broker: dict[str, dict[str, Decimal]] = {}
    by_date: dict[str, dict[str, Decimal]] = {}
    total_buys = ZERO
    total_sells = ZERO

    for trade in trades:
        amount, _ = trade_cash_amount(trade)
        broker = trade.broker
        day = trade.date.isoformat()
        by_broker.setdefault(broker, {"buys": ZERO, "sells": ZERO})
        by_date.setdefault(day, {"buys": ZERO, "sells": ZERO})
        if trade.quantity_diff >= ZERO:
            total_buys += amount
            by_broker[broker]["buys"] += amount
            by_date[day]["buys"] += amount
        else:
            total_sells += amount
            by_broker[broker]["sells"] += amount
            by_date[day]["sells"] += amount

    total_net = total_buys - total_sells
    broker_rows = []
    for broker, values in sorted(by_broker.items()):
        net = values["buys"] - values["sells"]
        broker_rows.append(
            {
                "broker": broker,
                "buys_eur": money(values["buys"]),
                "sells_eur": money(values["sells"]),
                "net_eur": money(net),
                "share_pct": float((net / total_net * Decimal("100")).quantize(Decimal("0.01"))) if total_net else None,
            }
        )

    date_rows = []
    for day, values in sorted(by_date.items(), reverse=True):
        net = values["buys"] - values["sells"]
        if net == ZERO:
            continue
        date_rows.append(
            {
                "date": day,
                "buys_eur": money(values["buys"]),
                "sells_eur": money(values["sells"]),
                "net_eur": money(net),
            }
        )

    return {
        "total_buys_eur": money(total_buys),
        "total_sells_eur": money(total_sells),
        "net_eur": money(total_net),
        "by_broker": broker_rows,
        "by_date": date_rows,
    }


def trade_friction_events(trades: list[Trade]) -> list[FrictionEvent]:
    events: list[FrictionEvent] = []
    for trade in trades:
        if trade.fees:
            amount, _ = convert_cash_to_eur(trade.fees, trade.cash_currency, trade.date)
            if amount:
                events.append(
                    FrictionEvent(
                        broker=trade.broker,
                        event_type="cost",
                        date=trade.date,
                        amount_eur=amount,
                        description=f"{trade.action} commission - {trade.asset}",
                    )
                )
        if trade.tax:
            amount, _ = convert_cash_to_eur(trade.tax, trade.cash_currency, trade.date)
            if amount:
                events.append(
                    FrictionEvent(
                        broker=trade.broker,
                        event_type="tax",
                        date=trade.date,
                        amount_eur=amount,
                        description=f"{trade.action} tax - {trade.asset}",
                    )
                )
    return events


def inferred_fineco_sell_tax_events(trades: list[Trade]) -> list[FrictionEvent]:
    quantities: dict[str, Decimal] = {}
    cost_basis: dict[str, Decimal] = {}
    events: list[FrictionEvent] = []

    for trade in trades:
        asset_key = trade.isin or trade.asset
        quantities.setdefault(asset_key, ZERO)
        cost_basis.setdefault(asset_key, ZERO)
        amount, _ = trade_cash_amount(trade)

        if trade.quantity_diff >= ZERO:
            quantities[asset_key] += trade.quantity_diff
            cost_basis[asset_key] += amount
            continue

        sell_quantity = abs(trade.quantity_diff)
        previous_quantity = quantities[asset_key]
        previous_cost = cost_basis[asset_key]
        if previous_quantity > ZERO and sell_quantity > ZERO:
            removed_cost = min(previous_cost, previous_cost * sell_quantity / previous_quantity)
        else:
            removed_cost = ZERO

        net_gain = amount - removed_cost
        if trade.broker == "Fineco" and trade.tax == ZERO and net_gain > ZERO:
            tax = fineco_capital_gain_tax_from_net_gain(net_gain)
            if tax:
                events.append(
                    FrictionEvent(
                        broker="Fineco",
                        event_type="tax",
                        date=trade.date,
                        amount_eur=tax,
                        description=f"Implied capital gain tax - {trade.asset}",
                    )
                )

        cost_basis[asset_key] = max(ZERO, previous_cost - removed_cost)
        quantities[asset_key] = previous_quantity - sell_quantity

    return events


def read_trade_republic_tax_events(path: Path) -> list[FrictionEvent]:
    events: list[FrictionEvent] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if row.get("type") not in {"TAX_OPTIMIZATION", "DIVIDEND"}:
                continue
            raw_tax = parse_decimal(row.get("tax"))
            if not raw_tax:
                continue
            paid_tax = -raw_tax
            event_date = datetime.strptime(row["date"], "%Y-%m-%d").date()
            events.append(
                FrictionEvent(
                    broker="Trade Republic",
                    event_type="dividend_tax" if row.get("type") == "DIVIDEND" else "tax",
                    date=event_date,
                    amount_eur=paid_tax,
                    description=(row.get("description") or row.get("name") or row.get("type") or "Tax").strip(),
                )
            )
    return events


def read_fineco_friction_events(path: Path) -> list[FrictionEvent]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["Movimenti Dossier Titoli"]
    events: list[FrictionEvent] = []
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not row or row[2] != "Dividendo":
            continue
        net_amount = fineco_decimal(row[10])
        tax = fineco_dividend_tax_from_net(net_amount)
        if not tax:
            continue
        events.append(
            FrictionEvent(
                broker="Fineco",
                event_type="dividend_tax",
                date=parse_fineco_date(row[0]),
                amount_eur=tax,
                description=f"Implied dividend withholding - {str(row[3] or '').strip()}",
            )
        )
    return events


def read_etoro_friction_events(path: Path) -> list[FrictionEvent]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    events: list[FrictionEvent] = []

    if "Dividendi" in workbook.sheetnames:
        dividends = workbook["Dividendi"]
        for row in dividends.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            tax = fineco_decimal(row[10])
            if not tax:
                continue
            events.append(
                FrictionEvent(
                    broker="eToro",
                    event_type="dividend_tax",
                    date=datetime.strptime(str(row[0]), "%d/%m/%Y").date(),
                    amount_eur=tax,
                    description=f"Dividend withholding - {str(row[1] or '').strip()}",
                )
            )

    if "Posizioni chiuse" in workbook.sheetnames:
        closed = workbook["Posizioni chiuse"]
        for row in closed.iter_rows(min_row=2, values_only=True):
            if not row or not row[6]:
                continue
            closed_at = datetime.strptime(str(row[6]), "%d/%m/%Y %H:%M:%S").date()
            asset = str(row[1] or "Closed position").strip()
            spread_usd = abs(fineco_decimal(row[8]))
            overnight_usd = abs(fineco_decimal(row[18]))
            for label, amount_usd in (("Spread fee", spread_usd), ("Overnight fee/dividend", overnight_usd)):
                if not amount_usd:
                    continue
                amount_eur, _ = convert_cash_to_eur(amount_usd, "USD", closed_at)
                events.append(
                    FrictionEvent(
                        broker="eToro",
                        event_type="cost",
                        date=closed_at,
                        amount_eur=amount_eur,
                        description=f"{label} - {asset}",
                    )
                )
    return events


def read_extra_friction_events() -> list[FrictionEvent]:
    events: list[FrictionEvent] = []
    tr_file = latest_trade_republic_export()
    fineco_file = latest_fineco_export()
    etoro_file = latest_etoro_export()
    if tr_file:
        events.extend(read_trade_republic_tax_events(tr_file))
    if fineco_file:
        events.extend(read_fineco_friction_events(fineco_file))
    if etoro_file:
        events.extend(read_etoro_friction_events(etoro_file))
    return events


def summarize_frictions(events: list[FrictionEvent], market_value: Decimal | float | int | None) -> dict[str, Any]:
    costs = ZERO
    trade_taxes = ZERO
    dividend_taxes = ZERO
    by_broker: dict[str, dict[str, Decimal]] = {}
    rows = []

    for event in sorted(events, key=lambda item: (item.date, item.broker, item.description), reverse=True):
        by_broker.setdefault(event.broker, {"costs": ZERO, "taxes": ZERO, "dividend_tax": ZERO})
        if event.event_type == "cost":
            costs += event.amount_eur
            by_broker[event.broker]["costs"] += event.amount_eur
            type_label = "Cost"
        elif event.event_type == "dividend_tax":
            dividend_taxes += event.amount_eur
            by_broker[event.broker]["dividend_tax"] += event.amount_eur
            type_label = "Dividend tax"
        else:
            trade_taxes += event.amount_eur
            by_broker[event.broker]["taxes"] += event.amount_eur
            type_label = "Tax"
        rows.append(
            {
                "date": event.date.isoformat(),
                "broker": event.broker,
                "type": event.event_type,
                "type_label": type_label,
                "description": event.description,
                "amount_eur": money(event.amount_eur),
            }
        )

    tax_total = trade_taxes + dividend_taxes
    total_drag = costs + tax_total
    market = Decimal(str(market_value)) if market_value is not None else ZERO
    broker_rows = []
    for broker, values in sorted(by_broker.items()):
        taxes = values["taxes"] + values["dividend_tax"]
        total = values["costs"] + taxes
        if not total:
            continue
        broker_rows.append(
            {
                "broker": broker,
                "costs_eur": money(values["costs"]),
                "taxes_eur": money(taxes),
                "dividend_tax_eur": money(values["dividend_tax"]),
                "total_eur": money(total),
            }
        )

    return {
        "status": "available",
        "message": "",
        "total_costs_eur": money(costs),
        "total_taxes_eur": money(tax_total),
        "trade_taxes_eur": money(trade_taxes),
        "dividend_tax_eur": money(dividend_taxes),
        "total_drag_eur": money(total_drag),
        "net_liquidation_eur": money(market - total_drag),
        "by_broker": broker_rows,
        "rows": rows,
    }


def empty_frictions(status: str = "unavailable", message: str = "") -> dict[str, Any]:
    return {
        "status": status,
        "message": message,
        "total_costs_eur": 0.0,
        "total_taxes_eur": 0.0,
        "trade_taxes_eur": 0.0,
        "dividend_tax_eur": 0.0,
        "total_drag_eur": 0.0,
        "net_liquidation_eur": None,
        "by_broker": [],
        "rows": [],
    }


def latest_family_file(person: str) -> Path | None:
    config = FAMILY_PORTFOLIOS.get(person)
    if not config:
        return None
    files = sorted(ROOT_DIR.glob(config["pattern"]))
    return files[-1] if files else None


def csv_number(value: str | None) -> Decimal | None:
    raw = (value or "").strip()
    if not raw or raw == "#N/A":
        return None
    if "," not in raw and "." in raw:
        parts = raw.split(".")
        if len(parts) > 1 and all(len(part) == 3 for part in parts[1:]):
            raw = "".join(parts)
    return parse_decimal(raw)


def parse_snapshot_date(value: str) -> date:
    return datetime.strptime(value.strip(), "%d/%m/%Y").date()


def read_family_snapshot(person: str) -> dict[str, Any]:
    path = latest_family_file(person)
    if not path:
        raise FileNotFoundError(f"Missing family portfolio CSV for {person}.")

    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    header = rows[0]
    date_cols = [(idx, parse_snapshot_date(label)) for idx, label in enumerate(header) if "/" in label]

    data_rows = []
    for row in rows[2:]:
        name = (row[0] if row else "").strip()
        if name in {"NW", "NW Change", "NW Change %", "USD/EUR"}:
            break
        if len(row) < 7:
            continue
        broker = row[1].strip() if len(row) > 1 else ""
        isin = row[2].strip().upper() if len(row) > 2 else ""
        ticker = row[3].strip() if len(row) > 3 else ""
        if not name or not broker:
            continue
        values = {}
        for idx, dt in date_cols:
            if idx < len(row):
                parsed = csv_number(row[idx])
                if parsed is not None:
                    values[dt] = parsed
        if values:
            data_rows.append(
                {
                    "asset": name,
                    "broker": broker,
                    "isin": isin,
                    "ticker": ticker,
                    "quantity": csv_number(row[4]) or ZERO,
                    "cost_price": csv_number(row[5]) or ZERO,
                    "currency": (row[6].strip() if len(row) > 6 else ""),
                    "values": values,
                }
            )

    totals_by_date: dict[date, Decimal] = {}
    counts_by_date: dict[date, int] = {}
    for _, dt in date_cols:
        values = [row["values"].get(dt, ZERO) for row in data_rows]
        total = sum(values, ZERO)
        count = sum(1 for value in values if value > ZERO)
        if total > ZERO:
            totals_by_date[dt] = total
            counts_by_date[dt] = count

    if not totals_by_date:
        raise ValueError(f"No portfolio values found in {path.name}.")
    max_count = max(counts_by_date.values())
    min_complete_count = max(1, math.ceil(max_count * 0.8))
    complete_totals_by_date = {
        dt: value for dt, value in totals_by_date.items() if counts_by_date.get(dt, 0) >= min_complete_count
    }
    latest_date = max(complete_totals_by_date)
    current_total = totals_by_date[latest_date]
    first_date = min(complete_totals_by_date)
    first_total = totals_by_date[first_date]
    profit = current_total - first_total
    return_pct = (profit / first_total * Decimal("100")) if first_total > ZERO else ZERO

    positions = []
    for row in data_rows:
        value = row["values"].get(latest_date, ZERO)
        if value <= ZERO:
            continue
        cost_basis = row["quantity"] * row["cost_price"] if row["quantity"] and row["cost_price"] else ZERO
        pl = value - cost_basis if cost_basis > ZERO else ZERO
        positions.append(
            {
                "asset": row["asset"],
                "broker": row["broker"],
                "isin": row["isin"],
                "symbol": row["ticker"],
                "quantity": decimal_to_float(row["quantity"]),
                "cost_price": decimal_to_float(row["cost_price"]) if row["cost_price"] else None,
                "price": None,
                "price_currency": row["currency"],
                "market_value_eur": money(value),
                "cost_basis_eur": money(cost_basis),
                "is_open": True,
                "display_pl_eur": money(pl) if cost_basis > ZERO else None,
                "display_pl_pct": float((pl / cost_basis * Decimal("100")).quantize(Decimal("0.01"))) if cost_basis > ZERO else None,
                "pricing_status": "snapshot",
            }
        )

    valuation_series = []
    for dt in sorted(complete_totals_by_date):
        value = totals_by_date[dt]
        point_profit = value - first_total
        point_return = (point_profit / first_total * Decimal("100")) if first_total > ZERO else ZERO
        valuation_series.append(
            {
                "date": dt.isoformat(),
                "market_value": money(value),
                "net_contributions": money(first_total),
                "profit": money(point_profit),
                "return_pct": float(point_return.quantize(Decimal("0.01"))),
                "priced_positions": len(positions),
                "unpriced_positions": 0,
            }
        )

    return {
        "path": path,
        "person_name": FAMILY_PORTFOLIOS[person]["name"],
        "latest_date": latest_date,
        "first_date": first_date,
        "positions": sorted(positions, key=lambda item: item["market_value_eur"], reverse=True),
        "valuation_series": valuation_series,
        "totals": {
            "market_value": money(current_total),
            "return_pct": float(return_pct.quantize(Decimal("0.01"))),
            "open_cost_basis": None,
            "unrealized_pl": None,
            "realized_pl": None,
            "net_contributions": None,
            "historical_profit": money(profit),
            "priced_assets": len(positions),
            "unpriced_assets": 0,
        },
    }


def empty_money_table() -> dict[str, Any]:
    return {
        "rows": [],
        "total_eur": 0.0,
        "tax_eur": 0.0,
        "count": 0,
        "by_asset": [],
        "by_broker": [],
    }


def empty_contributions() -> dict[str, Any]:
    return {
        "total_buys_eur": 0.0,
        "total_sells_eur": 0.0,
        "net_eur": 0.0,
        "by_broker": [],
        "by_date": [],
    }


def normalize_berkshire_mode(value: str | None) -> str:
    return "lookthrough" if value == "lookthrough" else "stock"


def normalize_proxy_mode(value: str | None) -> str:
    return "on" if value == "on" else "off"


def configured_trade(adjustment: Any, history_start: date) -> Trade:
    event_date = history_start if adjustment.use_history_start else adjustment.date
    if event_date is None:
        raise ValueError(f"Configured trade for {adjustment.asset} has no date.")
    total = adjustment.total if adjustment.total is not None else adjustment.price * adjustment.quantity
    quantity_diff = adjustment.quantity if adjustment.action.upper() == "BUY" else -adjustment.quantity
    return Trade(
        asset=adjustment.asset,
        isin=adjustment.isin,
        broker=adjustment.broker,
        action=adjustment.action.upper(),
        currency_hint=adjustment.currency,
        cash_currency=adjustment.currency,
        date=event_date,
        price=adjustment.price,
        quantity=adjustment.quantity,
        quantity_diff=quantity_diff,
        total_spend=total,
        fees=adjustment.fees,
        tax=adjustment.tax,
        grand_total=total,
        grand_total_present=True,
        source=adjustment.source,
    )


def family_dashboard_payload(
    person: str,
    refresh: bool = False,
    berkshire_mode: str = "stock",
    proxy_mode: str = "off",
    broker: str = "all",
    live_only: str = "off",
) -> dict[str, Any]:
    profile = SETTINGS.portfolios.get(person.lower())
    if profile is None:
        raise KeyError(f"Unknown portfolio: {person}")
    proxy_mode = normalize_proxy_mode(proxy_mode)
    broker = (broker or "all").strip().lower()
    snapshot = read_family_snapshot(person)
    mappings = read_mappings()

    # Filter snapshot positions by live_only
    if live_only == "on":
        filtered_positions = []
        for pos in snapshot["positions"]:
            symbol = pos["isin"] or pos["symbol"]
            mapping = mapping_for(pos["asset"], symbol, mappings)
            direct_symbol = yahoo_symbol_from_mapping(mapping.get("ticker", ""), mapping.get("exchange", ""))
            if not direct_symbol and symbol:
                direct_symbol = resolve_isin(symbol).get("symbol", "")
            is_crypto = pos.get("broker", "").lower() == "crypto wallet"
            if direct_symbol or is_crypto:
                filtered_positions.append(pos)
        snapshot = {
            **snapshot,
            "positions": filtered_positions
        }
    first_date = profile.history_start or snapshot["first_date"]

    tr_file = latest_family_trade_republic_export(person)
    has_tr = tr_file is not None

    # Get all brokers first (unfiltered)
    snap_brokers = set()
    for pos in snapshot["positions"]:
        b = pos.get("broker")
        if b:
            snap_brokers.add(b)
    if has_tr:
        snap_brokers.add("Trade Republic")
    all_brokers = sorted(list(snap_brokers))

    # Filter snapshot positions by broker at the start
    if broker != "all":
        snapshot = {
            **snapshot,
            "positions": [pos for pos in snapshot["positions"] if str(pos.get("broker", "")).strip().lower() == broker]
        }

    fake_trades = []
    for pos in snapshot["positions"]:
        if has_tr and str(pos.get("broker", "")).strip().lower() == "trade republic":
            continue

        quantity = Decimal(str(pos["quantity"]))
        cost_basis = Decimal(str(pos["cost_basis_eur"]))
        
        configured_quantity = profile.position_quantity_overrides.get(pos["asset"])
        if configured_quantity is not None:
            quantity = configured_quantity
            cost_basis = quantity * Decimal(str(pos["cost_price"]))
        
        if quantity <= ZERO and pos.get("market_value_eur") is not None:
            snapshot_val = Decimal(str(pos["market_value_eur"]))
            symbol = pos["isin"] or pos["symbol"]
            mapping = mapping_for(pos["asset"], symbol, mappings)
            direct_symbol = yahoo_symbol_from_mapping(mapping.get("ticker", ""), mapping.get("exchange", ""))
            
            if not direct_symbol and symbol:
                direct_symbol = resolve_isin(symbol).get("symbol", "")
            
            computed_quantity = None
            if direct_symbol:
                from datetime import timedelta
                # Use latest_date to find true quantity based on latest snapshot value
                latest_date = snapshot["latest_date"]
                
                # Fetch all history from first_date up to latest_date to find both start and latest prices
                history = fetch_history(direct_symbol, first_date, latest_date + timedelta(days=5))
                prices = history.get("prices", {})
                
                latest_price = previous_price(prices, latest_date)
                
                if latest_price:
                    computed_quantity = snapshot_val / Decimal(str(latest_price))
                    
                    # Find earliest available price on or after first_date
                    sorted_dates = sorted(prices.keys())
                    start_price = None
                    for d in sorted_dates:
                        if d >= first_date.isoformat():
                            start_price = prices[d]
                            break
                            
                    if start_price:
                        # Normalize cost basis to start date (or IPO date) to track true growth
                        cost_basis = computed_quantity * Decimal(str(start_price))
                    else:
                        cost_basis = snapshot_val
            
            if computed_quantity:
                quantity = computed_quantity
            else:
                quantity = Decimal("1")
                cost_basis = snapshot_val
            
        if quantity > ZERO:
            price = cost_basis / quantity
            fake_trades.append(
                Trade(
                    asset=pos["asset"],
                    isin=pos["isin"] or pos["symbol"],
                    broker=pos["broker"],
                    action="BUY",
                    currency_hint=pos.get("price_currency") or "EUR",
                    cash_currency="EUR",
                    date=first_date,
                    price=price,
                    quantity=quantity,
                    quantity_diff=quantity,
                    total_spend=cost_basis,
                    fees=ZERO,
                    tax=ZERO,
                    grand_total=cost_basis,
                    grand_total_present=True,
                    source="snapshot",
                )
            )

    if has_tr and (broker == "all" or broker == "trade republic"):
        fake_trades.extend(read_trade_republic_trades(tr_file))

    fake_trades.extend(configured_trade(adjustment, first_date) for adjustment in profile.extra_trades)

    if broker != "all":
        fake_trades = [t for t in fake_trades if t.broker.lower() == broker]

    summary = summarize_trades(fake_trades)
    
    # Re-inject aggregate snapshot market values so enrich_positions fallback works correctly
    snapshot_values = {}
    for pos in snapshot["positions"]:
        val = pos.get("market_value_eur")
        if val is not None:
            snapshot_values[pos["asset"]] = snapshot_values.get(pos["asset"], ZERO) + Decimal(str(val))
            
    for pos in summary["positions"]:
        if pos["asset"] in snapshot_values:
            pos["market_value_eur"] = float(snapshot_values[pos["asset"]])

    priced = enrich_positions(summary["positions"], mappings, refresh=refresh)
    distribution = calculate_distribution(
        priced["positions"],
        read_exposures(berkshire_mode=berkshire_mode, proxy_mode=proxy_mode),
        berkshire_mode=berkshire_mode,
        proxy_mode=proxy_mode,
    )
    valuation = calculate_valuation_series(fake_trades, mappings, refresh=refresh, person=person, broker=broker)
    
    trade_assets = {trade.asset for trade in fake_trades}
    mapped_assets = set(mappings)
    assets_without_isin = {
        position["asset"]
        for position in summary["positions"]
        if position["is_open"] and not position.get("isin") and not mapping_for(position["asset"], "", mappings).get("isin")
    }

    # Load dividends and cash interests to compute final total returns
    dividends = read_portfolio_dividends(person)
    interests = read_cash_interests(person)
    if broker != "all":
        dividends = [d for d in dividends if d.broker.lower() == broker]
        interests = [i for i in interests if i.get("broker", "").lower() == broker]
    accumulated_dividends = sum((d.amount_eur for d in dividends), ZERO)
    accumulated_interest = sum((Decimal(str(i["net_eur"])) for i in interests), ZERO)
    total_market_value = Decimal(str(priced["market_value"])) + accumulated_dividends + accumulated_interest

    totals = {
        **summary["totals"],
        "market_value": priced["market_value"],
        "total_market_value": float(total_market_value.quantize(Decimal("0.01"))),
        "unrealized_pl": priced["unrealized_pl"],
        "estimated_total_value": money(Decimal(str(priced["market_value"]))),
        "priced_assets": priced["priced_assets"],
        "unpriced_assets": priced["unpriced_assets"],
    }
    if valuation["series"]:
        totals["return_pct"] = valuation["series"][-1]["return_pct"]
        totals["total_return_pct"] = valuation["series"][-1]["total_return_pct"]
        totals["historical_profit"] = valuation["series"][-1]["profit"]
        totals["historical_total_profit"] = valuation["series"][-1]["total_profit"]
        totals["total_market_value"] = valuation["series"][-1]["total_market_value"]
        totals["total_net_contributions"] = valuation["series"][-1]["total_net_contributions"]
    totals["variations"] = valuation.get("variations", {})

    f_events = trade_friction_events(fake_trades)
    f_events.extend(
        FrictionEvent(
            broker=adjustment.broker,
            event_type=adjustment.event_type,
            date=adjustment.date,
            amount_eur=adjustment.amount_eur,
            description=adjustment.description,
        )
        for adjustment in profile.extra_frictions
    )
    if has_tr and (broker == "all" or broker == "trade republic"):
        f_events.extend(read_trade_republic_tax_events(tr_file))

    if has_tr or profile.extra_frictions:
        frictions_payload = summarize_frictions(f_events, priced["market_value"])
    else:
        frictions_payload = empty_frictions(
            "snapshot_unavailable",
            "Taxes, broker costs, and net liquidation are not available from monthly snapshot CSV files.",
        )

    tax_losses = [
        {
            "year": item.year,
            "amount_eur": money(item.amount_eur),
            "expires_year": item.expires_year,
            "broker": item.broker,
        }
        for item in profile.tax_losses
    ]

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "person": person,
        "person_name": snapshot["person_name"],
        "berkshire_mode": berkshire_mode,
        "proxy_mode": proxy_mode,
        "trade_file": configured_path_label(snapshot["path"]),
        "trade_source": f"{snapshot['person_name']} CSV snapshot",
        "mapping_file": configured_path_label(MAPPINGS_CSV),
        "trade_count": len(fake_trades),
        "asset_count": len(trade_assets),
        "date_range": {
            "start": min(trade.date for trade in fake_trades).isoformat() if fake_trades else None,
            "end": valuation["series"][-1]["date"] if valuation["series"] else snapshot["latest_date"].isoformat(),
        },
        "totals": totals,
        "series": summary["series"],
        "valuation_series": valuation["series"],
        "valuation_status": {
            "status": valuation["status"],
            "symbols": valuation.get("symbols", []),
        },
        "positions": priced["positions"],
        "distribution": distribution,
        "dividends": summarize_dividends(dividends) if dividends else empty_money_table(),
        "cash_interests": summarize_cash_interests(read_cash_interests(person)),
        "todo_items": profile.todo_items,
        "expenses": empty_expenses("Expense classification is currently available for the primary portfolio only."),
        "net_contributions": summarize_net_contributions(fake_trades),
        "frictions": frictions_payload,
        "tax_losses": tax_losses,
        "mapping_status": {
            "missing_in_mapping": sorted(assets_without_isin),
            "extra_in_mapping": sorted(mapped_assets - trade_assets),
            "filled_isins": sum(1 for value in mappings.values() if value.get("isin")),
            "total_rows": len(mappings),
        },
        "brokers": all_brokers,
        "stats": calculate_portfolio_statistics(
            fake_trades,
            mappings,
            person=person,
            broker=broker,
            history_context=valuation.get("_history_context", {}),
        ),
    }
    payload["news_symbols"] = news_symbols_from_payload(payload)
    return payload


def read_mappings() -> dict[str, dict[str, str]]:
    mappings: dict[str, dict[str, str]] = {}
    if not MAPPINGS_CSV.exists():
        return mappings

    with MAPPINGS_CSV.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            asset = (row.get("asset_name") or "").strip()
            isin = (row.get("isin") or "").strip().upper()
            if asset:
                mappings[asset] = {
                    "isin": isin,
                    "ticker": (row.get("Ticker") or row.get("ticker") or "").strip(),
                    "exchange": (row.get("Borsa") or row.get("exchange") or "").strip(),
                }
    return mappings


def read_crypto_wallet_positions(person: str = PRIMARY_PORTFOLIO_ID) -> list[dict[str, Any]]:
    if not CRYPTO_WALLET_POSITIONS_CSV.exists():
        return []
    positions: list[dict[str, Any]] = []
    with CRYPTO_WALLET_POSITIONS_CSV.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if (row.get("person") or PRIMARY_PORTFOLIO_ID).strip().lower() != person.lower():
                continue
            quantity = parse_decimal(row.get("quantity"))
            if quantity <= ZERO:
                continue
            market_value = parse_decimal(row.get("market_value_eur"))
            cost_basis = parse_decimal(row.get("cost_basis_eur"))
            position = {
                "asset": (row.get("asset") or row.get("symbol") or "").strip(),
                "isin": "",
                "quantity": float(quantity),
                "cost_basis_eur": float(cost_basis),
                "is_open": True,
                "realized_pl_eur": 0.0,
                "realized_pl_pct": None,
                "broker": (row.get("broker") or "Crypto Wallet").strip(),
                "symbol": (row.get("symbol") or "").strip(),
                "asset_class": (row.get("asset_class") or "Crypto").strip() or "Crypto",
                "sector": (row.get("sector") or "Crypto").strip() or "Crypto",
                "geo": (row.get("geo") or "Global").strip() or "Global",
                "wallet_label": (row.get("wallet_label") or "").strip(),
                "wallet_address": (row.get("wallet_address") or "").strip(),
                "chain": (row.get("chain") or "").strip(),
                "source": (row.get("source") or "crypto_wallet").strip(),
            }
            if market_value > ZERO:
                position["market_value_eur"] = float(market_value)
            positions.append(position)
    return positions


def crypto_wallet_transaction_start_date() -> date:
    path = CRYPTO_WALLET_TRANSACTIONS_CSV
    if not path.exists():
        return date.today() - timedelta(days=30)
    dates: list[date] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            raw = (row.get("created_at") or "").strip()
            if not raw:
                continue
            try:
                dates.append(datetime.fromisoformat(raw.replace("Z", "+00:00")).date())
            except ValueError:
                continue
    return min(dates) if dates else date.today() - timedelta(days=30)


def crypto_wallet_valuation(wallet_positions: list[dict[str, Any]]) -> dict[str, Any]:
    if not wallet_positions:
        return {"series": [], "status": "empty", "symbols": [], "variations": {}}
    events = read_crypto_events()
    today = date.today()
    if not events:
        market_value = sum((Decimal(str(position.get("market_value_eur") or 0)) for position in wallet_positions), ZERO)
        cost_basis = sum((Decimal(str(position.get("cost_basis_eur") or 0)) for position in wallet_positions), ZERO)
        net_contributions = cost_basis if cost_basis > ZERO else market_value
        profit = market_value - net_contributions
        return_pct = profit / net_contributions * Decimal("100") if net_contributions > ZERO else ZERO
        return {
            "series": [
                {
                    "date": today.isoformat(),
                    "market_value": money(market_value),
                    "total_market_value": money(market_value),
                    "net_contributions": money(net_contributions),
                    "total_net_contributions": money(net_contributions),
                    "profit": money(profit),
                    "total_profit": money(profit),
                    "return_pct": float(return_pct.quantize(Decimal("0.01"))),
                    "total_return_pct": float(return_pct.quantize(Decimal("0.01"))),
                    "priced_positions": len(wallet_positions),
                    "unpriced_positions": 0,
                    "msci_return_pct": 0.0,
                    "xeon_return_pct": 0.0,
                    "inflation_return_pct": 0.0,
                }
            ],
            "status": "crypto_wallet_snapshot",
            "symbols": sorted({str(position.get("symbol") or "") for position in wallet_positions if position.get("symbol")}),
            "variations": {"1d": {"amount": 0.0, "pct": 0.0, "past_date": today.isoformat()}, "1w": {"amount": 0.0, "pct": 0.0, "past_date": today.isoformat()}, "1m": {"amount": 0.0, "pct": 0.0, "past_date": today.isoformat()}},
        }

    start = events[0]["date"]
    recent_dates = {today - timedelta(days=i) for i in range(35) if today - timedelta(days=i) >= start}
    dates = sorted(month_end_dates(start, today) | {event["date"] for event in events} | recent_dates)
    histories = {
        asset: fetch_crypto_eur_history(asset, start, today, refresh=False)
        for asset in {"BTC", "ETH", "TON", "USDC"}
        if any(event["asset"] == asset for event in events)
    }
    latest_quantities = {
        str(position.get("asset") or "").upper(): Decimal(str(position.get("quantity") or 0))
        for position in wallet_positions
    }
    latest_values = {
        str(position.get("asset") or "").upper(): Decimal(str(position.get("market_value_eur") or 0))
        for position in wallet_positions
    }
    latest_cost_basis = sum((Decimal(str(position.get("cost_basis_eur") or 0)) for position in wallet_positions), ZERO)

    # Fetch MSCI World prices for comparison
    msci_prices = {}
    for sym in ["SWDA.MI", "EUNL.DE", "URTH"]:
        h = fetch_history(sym, start, today, refresh=False)
        if h and h.get("status") == "priced" and h.get("prices"):
            msci_prices = h["prices"]
            break

    p0 = None
    if msci_prices:
        p0 = previous_price(msci_prices, start)
        if p0 is None:
            sorted_msci_dates = sorted(msci_prices.keys())
            if sorted_msci_dates:
                p0 = msci_prices[sorted_msci_dates[0]]

    # Fetch XEON prices for comparison
    xeon_prices = {}
    h_xeon = fetch_history("XEON.DE", start, today, refresh=False)
    if h_xeon and h_xeon.get("status") == "priced" and h_xeon.get("prices"):
        xeon_prices = h_xeon["prices"]

    xeon_p0 = None
    if xeon_prices:
        xeon_p0 = previous_price(xeon_prices, start)
        if xeon_p0 is None:
            sorted_xeon_dates = sorted(xeon_prices.keys())
            if sorted_xeon_dates:
                xeon_p0 = xeon_prices[sorted_xeon_dates[0]]

    # Load CPI data for inflation comparison
    cpi_data = fetch_eurostat_cpi()
    
    def get_cpi_value(d: date) -> float | None:
        if not cpi_data:
            return None
        label = f"{d.year:04d}-{d.month:02d}"
        if label in cpi_data:
            return cpi_data[label]
        sorted_labels = sorted(cpi_data.keys())
        best_val = None
        for l in sorted_labels:
            if l <= label:
                best_val = cpi_data[l]
            else:
                break
        if best_val is not None:
            return best_val
        return cpi_data[sorted_labels[-1]] if sorted_labels else None

    start_cpi = get_cpi_value(start)

    holdings: dict[str, Decimal] = {}
    cost_basis_by_asset: dict[str, Decimal] = {}
    invested = ZERO
    event_index = 0
    series = []
    for point_date in dates:
        while event_index < len(events) and events[event_index]["date"] <= point_date:
            event = events[event_index]
            asset = event["asset"]
            quantity = event["quantity"]
            cash_eur = event["cash_eur"]
            holdings[asset] = holdings.get(asset, ZERO) + quantity
            if cash_eur > ZERO and quantity > ZERO:
                invested += cash_eur
                cost_basis_by_asset[asset] = cost_basis_by_asset.get(asset, ZERO) + cash_eur
            elif quantity < ZERO:
                previous_quantity = holdings.get(asset, ZERO) - quantity
                if previous_quantity > ZERO:
                    previous_cost = cost_basis_by_asset.get(asset, ZERO)
                    removed_cost = min(previous_cost, previous_cost * abs(quantity) / previous_quantity)
                    cost_basis_by_asset[asset] = max(ZERO, previous_cost - removed_cost)
            event_index += 1

        if point_date == today:
            for asset, quantity in latest_quantities.items():
                holdings[asset] = quantity

        market_value = ZERO
        priced = 0
        unpriced = 0
        for asset, quantity in holdings.items():
            if quantity <= ZERO:
                continue
            if point_date == today and asset in latest_values and latest_values[asset] > ZERO:
                market_value += latest_values[asset]
                priced += 1
                continue
            price = previous_price((histories.get(asset) or {}).get("prices", {}), point_date)
            if price is None:
                unpriced += 1
                market_value += cost_basis_by_asset.get(asset, ZERO)
                continue
            market_value += quantity * Decimal(str(price))
            priced += 1

        net_contributions = sum(cost_basis_by_asset.values(), ZERO)
        if point_date == today and latest_cost_basis > ZERO:
            net_contributions = latest_cost_basis
        profit = market_value - net_contributions
        return_pct = profit / net_contributions * Decimal("100") if net_contributions > ZERO else ZERO
        msci_return = 0.0
        if p0 and p0 > 0:
            price_t = previous_price(msci_prices, point_date)
            if price_t is not None:
                msci_return = float((Decimal(str(price_t)) - Decimal(str(p0))) / Decimal(str(p0)) * 100)

        xeon_return = 0.0
        if xeon_p0 and xeon_p0 > 0:
            price_t = previous_price(xeon_prices, point_date)
            if price_t is not None:
                xeon_return = float((Decimal(str(price_t)) - Decimal(str(xeon_p0))) / Decimal(str(xeon_p0)) * 100)

        inflation_return = 0.0
        if start_cpi and start_cpi > 0:
            cpi_t = get_cpi_value(point_date)
            if cpi_t:
                inflation_return = float((Decimal(str(cpi_t)) - Decimal(str(start_cpi))) / Decimal(str(start_cpi)) * 100)
        else:
            days_since_start = (point_date - start).days
            inflation_return = float(((Decimal("1.02") ** (Decimal(str(days_since_start)) / Decimal("365.25"))) - 1) * 100)

        series.append(
            {
                "date": point_date.isoformat(),
                "market_value": money(market_value),
                "total_market_value": money(market_value),
                "net_contributions": money(net_contributions),
                "total_net_contributions": money(net_contributions),
                "profit": money(profit),
                "total_profit": money(profit),
                "return_pct": float(return_pct.quantize(Decimal("0.01"))),
                "total_return_pct": float(return_pct.quantize(Decimal("0.01"))),
                "priced_positions": priced,
                "unpriced_positions": unpriced,
                "msci_return_pct": round(msci_return, 2),
                "xeon_return_pct": round(xeon_return, 2),
                "inflation_return_pct": round(inflation_return, 2),
            }
        )

    return {
        "series": series,
        "status": "crypto_wallet_reconstructed",
        "symbols": sorted({str(position.get("symbol") or "") for position in wallet_positions if position.get("symbol")}),
        "variations": calculate_variations(series),
    }


def crypto_wallet_contributions(wallet_positions: list[dict[str, Any]]) -> dict[str, Any]:
    net = sum((Decimal(str(position.get("cost_basis_eur") or 0)) for position in wallet_positions), ZERO)
    if net <= ZERO:
        net = sum((Decimal(str(position.get("market_value_eur") or 0)) for position in wallet_positions), ZERO)
    return {
        "total_buys_eur": money(net),
        "total_sells_eur": 0.0,
        "net_eur": money(net),
        "by_broker": [
            {
                "broker": "Crypto Wallet",
                "buys_eur": money(net),
                "sells_eur": 0.0,
                "net_eur": money(net),
                "share_pct": 100.0 if net > ZERO else None,
            }
        ],
        "by_date": [
            {
                "date": crypto_wallet_transaction_start_date().isoformat(),
                "buys_eur": money(net),
                "sells_eur": 0.0,
                "net_eur": money(net),
            }
        ]
        if net > ZERO
        else [],
    }


def parse_binance_time(value: str) -> date:
    raw = (value or "").strip()
    for fmt in ("%y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return date.today()


def latest_binance_transaction_history() -> Path | None:
    if not SETTINGS.scan_downloads:
        return None
    files = sorted(Path.home().joinpath("Downloads").glob("Binance-Transaction-History-*.csv"))
    return files[-1] if files else None


def read_crypto_events() -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    if CRYPTO_WALLET_TRANSACTIONS_CSV.exists():
        with CRYPTO_WALLET_TRANSACTIONS_CSV.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                if (row.get("person") or PRIMARY_PORTFOLIO_ID).strip().lower() != PRIMARY_PORTFOLIO_ID:
                    continue
                asset = (row.get("asset") or "").strip().upper()
                quantity = parse_decimal(row.get("quantity"))
                native = parse_decimal(row.get("native_amount"))
                if not asset or quantity == ZERO:
                    continue
                try:
                    event_date = datetime.fromisoformat((row.get("created_at") or "").replace("Z", "+00:00")).date()
                except ValueError:
                    continue
                events.append(
                    {
                        "date": event_date,
                        "asset": asset,
                        "quantity": quantity,
                        "cash_eur": native,
                        "source": "coinbase_account",
                    }
                )

    binance_file = latest_binance_transaction_history()
    if binance_file:
        rows: list[dict[str, str]] = []
        with binance_file.open(newline="", encoding="utf-8-sig") as handle:
            rows = [{key: (value or "").strip() for key, value in row.items()} for row in csv.DictReader(handle)]
        used_targets: set[int] = set()
        for index, row in enumerate(rows):
            if row.get("Operation") != "Binance Convert" or row.get("Coin") != "EUR":
                continue
            eur_spend = abs(parse_decimal(row.get("Change")))
            if eur_spend <= ZERO:
                continue
            target_index = None
            for candidate_index in (index - 1, index + 1, index - 2, index + 2):
                if candidate_index < 0 or candidate_index >= len(rows) or candidate_index in used_targets:
                    continue
                candidate = rows[candidate_index]
                coin = (candidate.get("Coin") or "").strip().upper()
                if candidate.get("Operation") == "Binance Convert" and coin and coin != "EUR" and parse_decimal(candidate.get("Change")) > ZERO:
                    target_index = candidate_index
                    break
            if target_index is None:
                continue
            used_targets.add(target_index)
            target = rows[target_index]
            events.append(
                {
                    "date": parse_binance_time(target.get("Time") or row.get("Time") or ""),
                    "asset": (target.get("Coin") or "").strip().upper(),
                    "quantity": parse_decimal(target.get("Change")),
                    "cash_eur": eur_spend,
                    "source": "binance_convert",
                }
            )
    return sorted(events, key=lambda item: item["date"])


def calculate_variations(series: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    if not series:
        return {
            "1d": {"amount": 0.0, "pct": 0.0, "msci_pct": None, "vs_msci_pct": None, "past_date": None},
            "1w": {"amount": 0.0, "pct": 0.0, "msci_pct": None, "vs_msci_pct": None, "past_date": None},
            "1m": {"amount": 0.0, "pct": 0.0, "msci_pct": None, "vs_msci_pct": None, "past_date": None},
        }
    latest = series[-1]
    latest_date = date.fromisoformat(latest["date"])

    def find_entry_for_date(target: date):
        best_entry = None
        for entry in series:
            entry_date = date.fromisoformat(entry["date"])
            if entry_date <= target:
                best_entry = entry
            else:
                break
        return best_entry

    variations = {}
    for key, target_date in {
        "1d": latest_date - timedelta(days=1),
        "1w": latest_date - timedelta(days=7),
        "1m": latest_date - timedelta(days=30),
    }.items():
        past = find_entry_for_date(target_date)
        if past and past["date"] != latest["date"]:
            amount = latest["profit"] - past["profit"]
            past_val = past["market_value"]
            pct_val = (amount / past_val * 100) if past_val > 0 else 0.0
            msci_pct = benchmark_delta(latest, past, "msci")
            variations[key] = {
                "amount": round(amount, 2),
                "pct": round(pct_val, 2),
                "msci_pct": round(msci_pct, 2) if msci_pct is not None else None,
                "vs_msci_pct": round(float(pct_val) - msci_pct, 2) if msci_pct is not None else None,
                "past_date": past["date"],
            }
        else:
            variations[key] = {"amount": 0.0, "pct": 0.0, "msci_pct": None, "vs_msci_pct": None, "past_date": latest["date"]}
    return variations


def exposure_key(asset_name: str, isin: str) -> str:
    normalized_isin = (isin or "").strip().upper()
    if normalized_isin:
        return f"isin:{normalized_isin}"
    return f"asset:{(asset_name or '').strip().casefold()}"


def exposure_row_item(row: dict[str, Any], default_asset_name: str = "", default_isin: str = "") -> dict[str, Any] | None:
    asset_name = (row.get("asset_name") or default_asset_name).strip()
    isin = (row.get("isin") or default_isin).strip().upper()
    if not asset_name and not isin:
        return None
    weight = parse_decimal(row.get("weight_pct"))
    if weight <= ZERO:
        return None
    return {
        "asset_name": asset_name,
        "isin": isin,
        "holding_name": (row.get("holding_name") or asset_name or isin).strip(),
        "holding_ticker": (row.get("holding_ticker") or "").strip(),
        "weight_pct": weight,
        "sector": (row.get("sector") or "Unclassified").strip() or "Unclassified",
        "geo": (row.get("geo") or "Unclassified").strip() or "Unclassified",
        "asset_class": (row.get("asset_class") or "Equity").strip() or "Equity",
    }


def load_berkshire_lookthrough_rows() -> list[dict[str, Any]]:
    if not BERKSHIRE_HOLDINGS_CSV.exists():
        return []
    rows: list[dict[str, Any]] = []
    with BERKSHIRE_HOLDINGS_CSV.open(newline="", encoding="utf-8-sig") as handle:
        for raw in csv.DictReader(handle):
            item = exposure_row_item(raw, "Berkshire Hathaway (B)", BERKSHIRE_ISIN)
            if item:
                rows.append(item)
    return rows


def read_berkshire_metadata() -> dict[str, Any]:
    if not BERKSHIRE_HOLDINGS_CSV.exists():
        return {}
    with BERKSHIRE_HOLDINGS_CSV.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        return {}
    weight_sum = sum((parse_decimal(row.get("weight_pct")) for row in rows), ZERO)
    first = rows[0]
    return {
        "issuer": "Berkshire Hathaway",
        "status": "official_sec_13f",
        "fetched_at": first.get("filing_date") or "",
        "holdings_url": first.get("source_url") or "",
        "rows": len(rows),
        "weight_sum": str(weight_sum.quantize(Decimal("0.0001")).normalize()),
        "message": f"SEC 13F-HR public equity holdings, report date {first.get('report_date') or 'unknown'}.",
    }


def load_proxy_exposure_rows() -> dict[str, list[dict[str, Any]]]:
    if not PROXY_EXPOSURES_CSV.exists():
        return {}
    rows_by_isin: dict[str, list[dict[str, Any]]] = {}
    with PROXY_EXPOSURES_CSV.open(newline="", encoding="utf-8-sig") as handle:
        for raw in csv.DictReader(handle):
            item = exposure_row_item(raw)
            if not item or item["isin"] not in PROXY_ISSUERS:
                continue
            rows_by_isin.setdefault(item["isin"], []).append(item)
            
    # Dynamically copy LU1681045370 (Amundi EM proxy) rows to IE00BKM4GZ66 (Core MSCI EM IMI)
    if "LU1681045370" in rows_by_isin and "IE00BKM4GZ66" in PROXY_ISSUERS:
        rows_by_isin["IE00BKM4GZ66"] = [
            {**row, "asset_name": "Core MSCI EM IMI USD (Acc)", "isin": "IE00BKM4GZ66"}
            for row in rows_by_isin["LU1681045370"]
        ]
        
    # Dynamically copy NL0011683594 (VanEck proxy) rows to IE00BYZ2Y955 (MBB Dynamic International Value Opportunity LA EUR)
    value_isin = "IE00BYZ2Y955"
    value_name = "MBB Dynamic International Value Opportunity LA EUR"
    if "NL0011683594" in rows_by_isin and value_isin in PROXY_ISSUERS:
        rows_by_isin[value_isin] = [
            {**row, "asset_name": value_name, "isin": value_isin}
            for row in rows_by_isin["NL0011683594"]
        ]

    # Dynamically copy/construct rows using MSCI World (IE00B4L5Y983) and S&P 500 (IE00B5BMR087)
    world_isin = "IE00B4L5Y983"
    ms_isin = "IE00B2NLMV86"
    ms_name = "MBB Mediolanum Morgan Stanley Global Selection LHA EUR"
    difesa_isin = "IT0005285157"
    difesa_name = "Eurizon Profilo Flessibile Difesa II"
    flexible_eq_isin = "LU0497415702"
    flexible_eq_name = "Eurizon Flexible Equity Strategy R EUR"

    # Configured funds that reuse existing public exposure compositions.
    anima_isin = "IT0004896715"
    anima_name = "Anima Fondo Trading F"
    fidelity_america_isin = "LU0755218046"
    fidelity_name = "Fidelity America Y acc Eur"
    sp500_isin = "IE00B5BMR087"
    
    if EXPOSURES_CSV.exists() and (
        ms_isin in PROXY_ISSUERS or 
        difesa_isin in PROXY_ISSUERS or 
        flexible_eq_isin in PROXY_ISSUERS or 
        anima_isin in PROXY_ISSUERS or 
        fidelity_america_isin in PROXY_ISSUERS
    ):
        msci_rows = []
        sp500_rows = []
        with EXPOSURES_CSV.open(newline="", encoding="utf-8-sig") as handle:
            for raw in csv.DictReader(handle):
                item = exposure_row_item(raw)
                if item:
                    if item["isin"] == world_isin:
                        msci_rows.append(item)
                    elif item["isin"] == sp500_isin:
                        sp500_rows.append(item)
        
        # Populate Morgan Stanley Global Selection (100% MSCI World)
        if msci_rows and ms_isin in PROXY_ISSUERS:
            rows_by_isin[ms_isin] = [
                {**row, "asset_name": ms_name, "isin": ms_isin}
                for row in msci_rows
            ]

        # Anima Fondo Trading F: 60% MSCI World, 40% Euro Cash / Bonds
        if msci_rows and anima_isin in PROXY_ISSUERS:
            anima_rows = [
                {
                    **row,
                    "asset_name": anima_name,
                    "isin": anima_isin,
                    "weight_pct": row["weight_pct"] * Decimal("0.6")
                }
                for row in msci_rows
            ]
            equity_sum = sum(row["weight_pct"] for row in anima_rows)
            anima_rows.append({
                "asset_name": anima_name,
                "isin": anima_isin,
                "holding_name": "Euro Cash / Bonds",
                "holding_ticker": "",
                "weight_pct": Decimal("100") - equity_sum,
                "sector": "Cash / Money Market",
                "geo": "Eurozone",
                "asset_class": "Cash equivalent",
            })
            rows_by_isin[anima_isin] = anima_rows

        # Fidelity America: 100% S&P 500
        if sp500_rows and fidelity_america_isin in PROXY_ISSUERS:
            sp500_sum = sum(row["weight_pct"] for row in sp500_rows)
            fidelity_rows = []
            accumulated_weight = Decimal("0")
            for i, row in enumerate(sp500_rows):
                if i == len(sp500_rows) - 1:
                    weight = Decimal("100") - accumulated_weight
                else:
                    weight = (row["weight_pct"] * Decimal("100") / sp500_sum).quantize(Decimal("0.000001"))
                    accumulated_weight += weight
                fidelity_rows.append({
                    **row,
                    "asset_name": fidelity_name,
                    "isin": fidelity_america_isin,
                    "weight_pct": weight
                })
            rows_by_isin[fidelity_america_isin] = fidelity_rows
                    # Eurizon Profilo Flessibile Difesa II: 10% MSCI World, 60% Eurozone Gov Bonds, rest Cash
        if msci_rows and difesa_isin in PROXY_ISSUERS:
            difesa_rows = [
                {
                    **row,
                    "asset_name": difesa_name,
                    "isin": difesa_isin,
                    "weight_pct": row["weight_pct"] * Decimal("0.1")
                }
                for row in msci_rows
            ]
            equity_sum = sum(row["weight_pct"] for row in difesa_rows)
            difesa_rows.append({
                "asset_name": difesa_name,
                "isin": difesa_isin,
                "holding_name": "Eurozone Gov Bonds",
                "holding_ticker": "",
                "weight_pct": Decimal("60"),
                "sector": "Fixed Income",
                "geo": "Eurozone",
                "asset_class": "Fixed Income",
            })
            difesa_rows.append({
                "asset_name": difesa_name,
                "isin": difesa_isin,
                "holding_name": "Euro Cash / Deposits",
                "holding_ticker": "",
                "weight_pct": Decimal("40") - equity_sum,
                "sector": "Cash / Money Market",
                "geo": "Eurozone",
                "asset_class": "Cash equivalent",
            })
            rows_by_isin[difesa_isin] = difesa_rows
            
        # Eurizon Flexible Equity Strategy: 55% MSCI World, 30% Eurozone Gov Bonds, rest Cash
        if msci_rows and flexible_eq_isin in PROXY_ISSUERS:
            flex_rows = [
                {
                    **row,
                    "asset_name": flexible_eq_name,
                    "isin": flexible_eq_isin,
                    "weight_pct": row["weight_pct"] * Decimal("0.55")
                }
                for row in msci_rows
            ]
            equity_sum = sum(row["weight_pct"] for row in flex_rows)
            flex_rows.append({
                "asset_name": flexible_eq_name,
                "isin": flexible_eq_isin,
                "holding_name": "Eurozone Gov Bonds",
                "holding_ticker": "",
                "weight_pct": Decimal("30"),
                "sector": "Fixed Income",
                "geo": "Eurozone",
                "asset_class": "Fixed Income",
            })
            flex_rows.append({
                "asset_name": flexible_eq_name,
                "isin": flexible_eq_isin,
                "holding_name": "Euro Cash / Deposits",
                "holding_ticker": "",
                "weight_pct": Decimal("70") - equity_sum,
                "sector": "Cash / Money Market",
                "geo": "Eurozone",
                "asset_class": "Cash equivalent",
            })
            rows_by_isin[flexible_eq_isin] = flex_rows

    # Eurizon Obbligazioni Euro High Yield: 80% High Yield Corporate Bonds, 20% Gov Bonds/BOTs
    hy_isin = "IT0001280541"
    hy_name = "Eurizon Obbligazioni Euro High Yield"
    if hy_isin in PROXY_ISSUERS:
        rows_by_isin[hy_isin] = [
            {
                "asset_name": hy_name,
                "isin": hy_isin,
                "holding_name": "Euro Corporate Bonds - Financials",
                "holding_ticker": "",
                "weight_pct": Decimal("30"),
                "sector": "Financials",
                "geo": "Europe",
                "asset_class": "Fixed Income",
            },
            {
                "asset_name": hy_name,
                "isin": hy_isin,
                "holding_name": "Euro Corporate Bonds - Industrials",
                "holding_ticker": "",
                "weight_pct": Decimal("30"),
                "sector": "Industrials",
                "geo": "Europe",
                "asset_class": "Fixed Income",
            },
            {
                "asset_name": hy_name,
                "isin": hy_isin,
                "holding_name": "Euro Corporate Bonds - Utilities",
                "holding_ticker": "",
                "weight_pct": Decimal("20"),
                "sector": "Utilities",
                "geo": "Europe",
                "asset_class": "Fixed Income",
            },
            {
                "asset_name": hy_name,
                "isin": hy_isin,
                "holding_name": "Eurozone Gov Bonds / Short-term debt",
                "holding_ticker": "",
                "weight_pct": Decimal("20"),
                "sector": "Cash / Money Market",
                "geo": "Eurozone",
                "asset_class": "Cash equivalent",
            },
        ]

    # Eurizon Riserva 2 Anni Classe A: 70% Short-term Eurozone Gov Bonds, 30% Cash/Deposits
    riserva_isin = "IT0005104424"
    riserva_name = "Eurizon Riserva 2 Anni Classe A"
    if riserva_isin in PROXY_ISSUERS:
        rows_by_isin[riserva_isin] = [
            {
                "asset_name": riserva_name,
                "isin": riserva_isin,
                "holding_name": "Eurozone Short-term Government Bonds",
                "holding_ticker": "",
                "weight_pct": Decimal("70"),
                "sector": "Fixed Income",
                "geo": "Eurozone",
                "asset_class": "Fixed Income",
            },
            {
                "asset_name": riserva_name,
                "isin": riserva_isin,
                "holding_name": "Euro Cash / Deposits",
                "holding_ticker": "",
                "weight_pct": Decimal("30"),
                "sector": "Cash / Money Market",
                "geo": "Eurozone",
                "asset_class": "Cash equivalent",
            },
        ]

    # Dynamically construct mock PIR-compliant rows for IT0001019329 (SMFI - Mediolanum Flessibile Futuro Italia LA PIR Acc EUR)
    pir_isin = "IT0001019329"
    pir_name = "SMFI - Mediolanum Flessibile Futuro Italia LA PIR Acc EUR"
    if pir_isin in PROXY_ISSUERS:
        rows_by_isin[pir_isin] = [
            {
                "asset_name": pir_name,
                "isin": pir_isin,
                "holding_name": "Enel Spa",
                "holding_ticker": "ENEL.MI",
                "weight_pct": Decimal("15"),
                "sector": "Utilities",
                "geo": "Italy",
                "asset_class": "Equity",
            },
            {
                "asset_name": pir_name,
                "isin": pir_isin,
                "holding_name": "Intesa Sanpaolo Spa",
                "holding_ticker": "ISP.MI",
                "weight_pct": Decimal("15"),
                "sector": "Financials",
                "geo": "Italy",
                "asset_class": "Equity",
            },
            {
                "asset_name": pir_name,
                "isin": pir_isin,
                "holding_name": "UniCredit Spa",
                "holding_ticker": "UCG.MI",
                "weight_pct": Decimal("15"),
                "sector": "Financials",
                "geo": "Italy",
                "asset_class": "Equity",
            },
            {
                "asset_name": pir_name,
                "isin": pir_isin,
                "holding_name": "Ferrari N.V.",
                "holding_ticker": "RACE.MI",
                "weight_pct": Decimal("10"),
                "sector": "Consumer Discretionary",
                "geo": "Italy",
                "asset_class": "Equity",
            },
            {
                "asset_name": pir_name,
                "isin": pir_isin,
                "holding_name": "Stellantis N.V.",
                "holding_ticker": "STLAM.MI",
                "weight_pct": Decimal("10"),
                "sector": "Consumer Discretionary",
                "geo": "Italy",
                "asset_class": "Equity",
            },
            {
                "asset_name": pir_name,
                "isin": pir_isin,
                "holding_name": "Nexi S.p.A.",
                "holding_ticker": "NEXI.MI",
                "weight_pct": Decimal("5"),
                "sector": "Financials",
                "geo": "Italy",
                "asset_class": "Equity",
            },
            {
                "asset_name": pir_name,
                "isin": pir_isin,
                "holding_name": "Italian Gov Bonds",
                "holding_ticker": "",
                "weight_pct": Decimal("30"),
                "sector": "Cash / Money Market",
                "geo": "Italy",
                "asset_class": "Cash equivalent",
            },
        ]

    # JPM Europe Equity (I and C)
    jpm_eur_i_isin = "LU2146152231"
    jpm_eur_i_name = "JPM Europe Equity I acc EUR"
    jpm_eur_c_isin = "LU0129441100"
    jpm_eur_c_name = "JPM Europe Equity C acc EUR"
    jpm_europe_constituents = [
        {"holding_name": "ASML Holding NV", "holding_ticker": "ASML.AS", "weight_pct": Decimal("15"), "sector": "Information Technology", "geo": "Europe", "asset_class": "Equity"},
        {"holding_name": "Novo Nordisk A/S", "holding_ticker": "NOVO-B.CO", "weight_pct": Decimal("15"), "sector": "Health Care", "geo": "Europe", "asset_class": "Equity"},
        {"holding_name": "Nestle SA", "holding_ticker": "NESN.SW", "weight_pct": Decimal("15"), "sector": "Consumer Staples", "geo": "Europe", "asset_class": "Equity"},
        {"holding_name": "Roche Holding AG", "holding_ticker": "ROG.SW", "weight_pct": Decimal("10"), "sector": "Health Care", "geo": "Europe", "asset_class": "Equity"},
        {"holding_name": "Novartis AG", "holding_ticker": "NOVN.SW", "weight_pct": Decimal("10"), "sector": "Health Care", "geo": "Europe", "asset_class": "Equity"},
        {"holding_name": "LVMH Moet Hennessy Louis Vuitton SE", "holding_ticker": "MC.PA", "weight_pct": Decimal("10"), "sector": "Consumer Discretionary", "geo": "Europe", "asset_class": "Equity"},
        {"holding_name": "AstraZeneca PLC", "holding_ticker": "AZN.L", "weight_pct": Decimal("10"), "sector": "Health Care", "geo": "Europe", "asset_class": "Equity"},
        {"holding_name": "Shell PLC", "holding_ticker": "SHEL.L", "weight_pct": Decimal("10"), "sector": "Energy", "geo": "Europe", "asset_class": "Equity"},
        {"holding_name": "SAP SE", "holding_ticker": "SAP.DE", "weight_pct": Decimal("5"), "sector": "Information Technology", "geo": "Europe", "asset_class": "Equity"},
    ]
    if jpm_eur_i_isin in PROXY_ISSUERS:
        rows_by_isin[jpm_eur_i_isin] = [
            {**const, "asset_name": jpm_eur_i_name, "isin": jpm_eur_i_isin}
            for const in jpm_europe_constituents
        ]
    if jpm_eur_c_isin in PROXY_ISSUERS:
        rows_by_isin[jpm_eur_c_isin] = [
            {**const, "asset_name": jpm_eur_c_name, "isin": jpm_eur_c_isin}
            for const in jpm_europe_constituents
        ]

    # Franklin Biotechnology
    franklin_biotech_isin = "LU0109394709"
    franklin_biotech_name = "Franklin Biotechnology Discv A acc USD"
    if franklin_biotech_isin in PROXY_ISSUERS:
        rows_by_isin[franklin_biotech_isin] = [
            {"asset_name": franklin_biotech_name, "isin": franklin_biotech_isin, "holding_name": "Amgen Inc.", "holding_ticker": "AMGN", "weight_pct": Decimal("20"), "sector": "Health Care", "geo": "United States", "asset_class": "Equity"},
            {"asset_name": franklin_biotech_name, "isin": franklin_biotech_isin, "holding_name": "Gilead Sciences Inc.", "holding_ticker": "GILD", "weight_pct": Decimal("20"), "sector": "Health Care", "geo": "United States", "asset_class": "Equity"},
            {"asset_name": franklin_biotech_name, "isin": franklin_biotech_isin, "holding_name": "Regeneron Pharmaceuticals Inc.", "holding_ticker": "REGN", "weight_pct": Decimal("20"), "sector": "Health Care", "geo": "United States", "asset_class": "Equity"},
            {"asset_name": franklin_biotech_name, "isin": franklin_biotech_isin, "holding_name": "Vertex Pharmaceuticals Inc.", "holding_ticker": "VRTX", "weight_pct": Decimal("20"), "sector": "Health Care", "geo": "United States", "asset_class": "Equity"},
            {"asset_name": franklin_biotech_name, "isin": franklin_biotech_isin, "holding_name": "Moderna Inc.", "holding_ticker": "MRNA", "weight_pct": Decimal("20"), "sector": "Health Care", "geo": "United States", "asset_class": "Equity"},
        ]

    # Templeton Global Bond
    templeton_bond_isin = "LU0195953079"
    templeton_bond_name = "Templeton Global Bond I acc EUR"
    if templeton_bond_isin in PROXY_ISSUERS:
        rows_by_isin[templeton_bond_isin] = [
            {"asset_name": templeton_bond_name, "isin": templeton_bond_isin, "holding_name": "US Government Bonds", "holding_ticker": "", "weight_pct": Decimal("40"), "sector": "Fixed Income", "geo": "United States", "asset_class": "Fixed Income"},
            {"asset_name": templeton_bond_name, "isin": templeton_bond_isin, "holding_name": "Eurozone Government Bonds", "holding_ticker": "", "weight_pct": Decimal("40"), "sector": "Fixed Income", "geo": "Eurozone", "asset_class": "Fixed Income"},
            {"asset_name": templeton_bond_name, "isin": templeton_bond_isin, "holding_name": "Japan Government Bonds", "holding_ticker": "", "weight_pct": Decimal("20"), "sector": "Fixed Income", "geo": "Japan", "asset_class": "Fixed Income"},
        ]

    # Schroder ISF Italian Equity
    schroder_italy_isin = "LU0106239527"
    schroder_italy_name = "Schroder ISF Italian Equity C acc Eur"
    if schroder_italy_isin in PROXY_ISSUERS:
        rows_by_isin[schroder_italy_isin] = [
            {"asset_name": schroder_italy_name, "isin": schroder_italy_isin, "holding_name": "Enel Spa", "holding_ticker": "ENEL.MI", "weight_pct": Decimal("20"), "sector": "Utilities", "geo": "Italy", "asset_class": "Equity"},
            {"asset_name": schroder_italy_name, "isin": schroder_italy_isin, "holding_name": "Intesa Sanpaolo Spa", "holding_ticker": "ISP.MI", "weight_pct": Decimal("20"), "sector": "Financials", "geo": "Italy", "asset_class": "Equity"},
            {"asset_name": schroder_italy_name, "isin": schroder_italy_isin, "holding_name": "UniCredit Spa", "holding_ticker": "UCG.MI", "weight_pct": Decimal("20"), "sector": "Financials", "geo": "Italy", "asset_class": "Equity"},
            {"asset_name": schroder_italy_name, "isin": schroder_italy_isin, "holding_name": "Ferrari N.V.", "holding_ticker": "RACE.MI", "weight_pct": Decimal("15"), "sector": "Consumer Discretionary", "geo": "Italy", "asset_class": "Equity"},
            {"asset_name": schroder_italy_name, "isin": schroder_italy_isin, "holding_name": "Stellantis N.V.", "holding_ticker": "STLAM.MI", "weight_pct": Decimal("15"), "sector": "Consumer Discretionary", "geo": "Italy", "asset_class": "Equity"},
            {"asset_name": schroder_italy_name, "isin": schroder_italy_isin, "holding_name": "Nexi S.p.A.", "holding_ticker": "NEXI.MI", "weight_pct": Decimal("10"), "sector": "Financials", "geo": "Italy", "asset_class": "Equity"},
        ]

    return rows_by_isin


def read_proxy_metadata() -> dict[str, dict[str, Any]]:
    rows_by_isin = load_proxy_exposure_rows()
    if not rows_by_isin:
        return {}
    try:
        fetched_at = datetime.fromtimestamp(PROXY_EXPOSURES_CSV.stat().st_mtime).isoformat(timespec="seconds")
    except OSError:
        fetched_at = ""
    metadata: dict[str, dict[str, Any]] = {}
    for isin, rows in rows_by_isin.items():
        weight_sum = sum((row["weight_pct"] for row in rows), ZERO)
        metadata[isin] = {
            "issuer": PROXY_ISSUERS.get(isin, ""),
            "status": "proxy_exposure",
            "fetched_at": fetched_at,
            "holdings_url": PROXY_SOURCE_URLS.get(isin) or configured_path_label(PROXY_EXPOSURES_CSV),
            "rows": len(rows),
            "weight_sum": str(weight_sum.quantize(Decimal("0.0001")).normalize()),
            "message": PROXY_MESSAGES.get(isin) or "Non-official proxy composition for visualization; switch to Official only to hide it.",
        }
    return metadata


def read_exposures(berkshire_mode: str = "stock", proxy_mode: str = "off") -> dict[str, list[dict[str, Any]]]:
    exposures: dict[str, list[dict[str, Any]]] = {}
    documents = read_etf_documents() if proxy_mode == "on" else {}

    # Private look-through rows are optional. Public Berkshire and proxy data must
    # remain available on a clean installation with no asset_exposures.csv.
    if EXPOSURES_CSV.exists():
        with EXPOSURES_CSV.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                item = exposure_row_item(row)
                if not item:
                    continue
                asset_name = item["asset_name"]
                isin = item["isin"]
                exposures.setdefault(exposure_key(asset_name, isin), []).append(item)
                if asset_name and isin:
                    exposures.setdefault(exposure_key(asset_name, ""), []).append(item)

    if berkshire_mode == "lookthrough":
        rows = load_berkshire_lookthrough_rows()
        if rows:
            exposures[exposure_key("Berkshire Hathaway (B)", BERKSHIRE_ISIN)] = rows
            exposures[exposure_key("Berkshire Hathaway (B)", "")] = rows
    if proxy_mode == "on":
        for isin, rows in load_proxy_exposure_rows().items():
            if not rows:
                continue
            if exposure_key("", isin) in exposures and is_full_official_composition(documents.get(isin)):
                continue
            asset_name = rows[0]["asset_name"]
            exposures[exposure_key(asset_name, isin)] = rows
            exposures[exposure_key(asset_name, "")] = rows
    return exposures


def determine_asset_type_from_class(asset_class: str) -> str:
    ac = (asset_class or "").lower()
    if ac in ["single share", "equities", "equity"]:
        return "STOCK"
    if ac in ["cash", "cash equivalent", "cash collateral and margins", "currency", "fx"]:
        return "CUR"
    if ac in ["etf", "etf underlying", "commodity", "commodities", "gold"]:
        return "ETF"
    return "STOCK"


def determine_asset_type(asset: str, isin: str, symbol: str, broker: str, exposures: dict) -> str:
    asset_lower = asset.lower()
    broker_lower = (broker or "").lower()
    symbol_str = (symbol or "")
    
    # 1. Currency / Crypto
    if (
        broker_lower == "crypto wallet" 
        or symbol_str.endswith("-USD") 
        or asset_lower in ["btc", "eth", "ton", "usdc", "usdt"]
    ):
        return "CUR"
        
    # 2. Check exposures
    key1 = exposure_key(asset, isin)
    key2 = exposure_key(asset, "")
    rows = exposures.get(key1) or exposures.get(key2)
    if rows:
        if all(r.get("asset_class") == "Single share" for r in rows):
            return "STOCK"
        if all(r.get("asset_class") in ["Cash", "Cash equivalent", "Cash Collateral and Margins"] for r in rows):
            return "CUR"
        return "ETF"
            
    # 3. Fallback based on name keywords
    etf_keywords = ["etf", "acc", "dist", "swap", "ucits", "index", "msci", "ftse", "stoxx", "leveraged", "lev"]
    if any(k in asset_lower for k in etf_keywords):
        return "ETF"
        
    return "STOCK"


def read_etf_documents() -> dict[str, dict[str, Any]]:
    if not ETF_DOCUMENTS_JSON.exists():
        return {}
    try:
        payload = json.loads(ETF_DOCUMENTS_JSON.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if isinstance(payload, dict) and isinstance(payload.get("funds"), list):
        return {
            str(row.get("isin") or "").strip().upper(): row
            for row in payload["funds"]
            if str(row.get("isin") or "").strip()
        }
    if isinstance(payload, dict):
        return {
            str(isin).strip().upper(): row
            for isin, row in payload.items()
            if isinstance(row, dict) and str(isin).strip()
        }
    return {}


def is_full_official_composition(document: dict[str, Any] | None) -> bool:
    status = str((document or {}).get("status") or "").strip()
    return status in FULL_OFFICIAL_COMPOSITION_STATUSES


def add_distribution_amount(bucket: dict[str, Decimal], key: str, amount: Decimal) -> None:
    bucket[key] = bucket.get(key, ZERO) + amount


def composition_source_label(position: dict[str, Any], is_single_share: bool) -> str:
    if is_single_share:
        return "stock"
    symbol = str(position.get("symbol") or "").strip()
    if symbol:
        return symbol
    asset = str(position.get("asset") or "").strip()
    return asset


HOLDING_NAME_OVERRIDES = {
    "ADVANCED MICRO DEVICES": "Advanced Micro Devices",
    "ADVANCED MICRO DEVICES INC": "Advanced Micro Devices",
    "ALIBABA GROUP HOLDING": "Alibaba Group",
    "ALIBABA GROUP HOLDING LTD": "Alibaba Group",
    "ALIBABA GROUP HOLDING LTD ADR": "Alibaba Group",
    "ALIBABA GROUP HOLDING-SP ADR": "Alibaba Group",
    "ALPHABET": "Alphabet",
    "ALPHABET INC": "Alphabet",
    "ALPHABET INC CLASS A": "Alphabet",
    "ALPHABET INC CLASS C": "Alphabet",
    "AMAZON.COM": "Amazon.com",
    "AMAZON.COM INC": "Amazon.com",
    "AMERICAN EXPRESS": "American Express",
    "AMERICAN EXPRESS CO": "American Express",
    "APPLE": "Apple",
    "APPLE INC": "Apple",
    "BANK AMERICA": "Bank of America",
    "BANK AMERICA CORP": "Bank of America",
    "BERKSHIRE HATHAWAY": "Berkshire Hathaway",
    "BERKSHIRE HATHAWAY INC CLASS B": "Berkshire Hathaway",
    "BROADCOM": "Broadcom",
    "BROADCOM INC": "Broadcom",
    "CHEVRON": "Chevron",
    "CHEVRON CORPORATION": "Chevron",
    "CHUBB LTD SWITZ": "Chubb",
    "COCA COLA": "Coca-Cola",
    "COCA COLA CO": "Coca-Cola",
    "JPMORGAN CHASE": "JPMorgan Chase",
    "JPMORGAN CHASE & CO": "JPMorgan Chase",
    "KRAFT HEINZ": "Kraft Heinz",
    "KRAFT HEINZ CO": "Kraft Heinz",
    "META PLATFORMS": "Meta Platforms",
    "META PLATFORMS INC": "Meta Platforms",
    "META PLATFORMS INC CLASS A": "Meta Platforms",
    "MICROSOFT": "Microsoft",
    "MICROSOFT CORP": "Microsoft",
    "MOODYS": "Moody's",
    "MOODYS CORP": "Moody's",
    "NVIDIA": "NVIDIA",
    "NVIDIA CORP": "NVIDIA",
    "OCCIDENTAL PETE": "Occidental Petroleum",
    "OCCIDENTAL PETE CORP": "Occidental Petroleum",
    "TAIWAN SEMICONDUCTOR MANUFACTURING": "Taiwan Semiconductor Manufacturing",
    "TAIWAN SEMICONDUCTOR MANUFACTURING CO LTD": "Taiwan Semiconductor Manufacturing",
    "TAIWAN SEMICONDUCTOR MANUFACTURING CO LTD ADR": "Taiwan Semiconductor Manufacturing",
    "TESLA": "Tesla",
    "TESLA INC": "Tesla",
}


LEGAL_SUFFIX_RE = re.compile(
    r"\b("
    r"INC|INCORPORATED|CORP|CORPORATION|CO|COMPANY|LTD|LIMITED|PLC|SA|SE|NV|AG|SPA|"
    r"ORD|ORDINARY SHARES?|COMMON STOCK|REGISTERED SHARES?|REG|ADR|ADS|"
    r"CLASS [A-Z]|CL [A-Z]|[A-Z]-SHS|[A-Z] SHS"
    r")\b\.?",
    flags=re.I,
)


def title_holding_name(value: str) -> str:
    words = []
    acronyms = {"AMD", "ASML", "BP", "BYD", "IBM", "ING", "JPMORGAN", "LVMH", "NVIDIA", "SK", "TSMC"}
    for word in re.split(r"(\s+|&|-)", value.lower()):
        upper = word.upper()
        if upper in acronyms:
            words.append(upper if upper != "JPMORGAN" else "JPMorgan")
        elif word in {" ", "-", "&"} or word.isspace():
            words.append(word)
        elif word:
            words.append(word[:1].upper() + word[1:])
        else:
            words.append(word)
    return "".join(words).strip()


def canonical_holding_name(name: str) -> str:
    raw = re.sub(r"\s+", " ", (name or "").strip())
    if not raw:
        return "Unclassified"
    if raw.startswith("Other issuer holdings"):
        return raw
    if raw.startswith("Other proxy holdings"):
        return raw
    if raw.startswith("EUR overnight cash equivalent"):
        return raw

    normalized = raw.upper().replace(".", "")
    normalized = normalized.replace(" AND ", " & ")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    override = HOLDING_NAME_OVERRIDES.get(normalized)
    if override:
        return override

    compact = LEGAL_SUFFIX_RE.sub(" ", normalized)
    compact = re.sub(r"\s+", " ", compact).strip(" -")
    override = HOLDING_NAME_OVERRIDES.get(compact)
    if override:
        return override
    if compact:
        return title_holding_name(compact)
    return title_holding_name(normalized)


COMMON_HOLDING_TICKERS = {
    "ALLIANZ": "ALV",
    "ZURICH INSURANCE GROUP": "ZURN.SW",
    "AXA": "CS",
    "INTEL": "INTC",
    "MICRON TECHNOLOGY": "MU",
    "SAMSUNG ELECTRONICS": "SMSN.IL",
    "MUENCHENER RUECKVER N": "MUV2.DE",
    "APPLE": "AAPL",
    "NVIDIA": "NVDA",
    "VISA": "V",
    "ALPHABET": "GOOGL",
    "SK HYNIX": "000660.KS",
    "MASTERCARD": "MA",
    "ALIBABA GROUP HOLDING REPRESEN": "BABA",
    "GENERALI": "G",
    "SWISS RE": "SREN.SW",
    "JPMORGAN CHASE": "JPM",
    "BANK OF AMERICA": "BAC",
    "MICROSOFT": "MSFT",
    "CROWDSTRIKE HOLDINGS": "CRWD",
    "PRUDENTIAL": "PRU",
    "BERKSHIRE HATHAWAY": "BRK-B",
    "TENCENT HOLDINGS": "TCEHY",
    "CISCO SYSTEMS": "CSCO",
    "META PLATFORMS": "META",
    "AMERICAN EXPRESS": "AXP",
    "DELL TECHNOLOGIES": "DELL",
    "SWISS LIFE HLDG N": "SLHN.SW",
    "CME GROUP": "CME",
    "ORACLE": "ORCL",
    "AMAZON COM": "AMZN",
    "AMAZONCOM": "AMZN",
    "AVIVA": "AV.L",
    "SAMPO": "SAMPO.HE",
    "MERCADOLIBRE": "MELI",
    "NN GROUP": "NN",
    "CHARLES SCHWAB": "SCHW",
    "INTUIT": "INTU",
    "LEGAL & GENERAL GROUP": "LGEN.L",
    "PALANTIR TECHNOLOGIES": "PLTR",
    "INTERNATIONAL BUSINESS MACHINES": "IBM",
    "GOLDMAN SACHS GROUP": "GS",
    "HSBC HOLDINGS": "HSBC",
    "APPLOVIN": "APP",
    "HANNOVER RUECK": "HNR1.DE",
    "SAP": "SAP",
    "ROYAL BANK OF CANADA": "RY",
    "BROADCOM": "AVGO",
}


BLOOMBERG_TO_YAHOO_SUFFIX = {
    "GR": ".DE",
    "US": "",
    "FP": ".PA",
    "SW": ".SW",
    "LN": ".L",
    "CN": ".TO",
    "IM": ".MI",
    "AU": ".AX",
    "SP": ".SI",
    "SM": ".MC",
    "BB": ".BR",
    "HK": ".HK",
    "JP": ".T",
    "JT": ".T",
    "IT": ".MI",
    "NO": ".OL",
    "PW": ".WA",
    "DC": ".CO",
    "AV": ".VI",
    "PL": ".LS",
    "SS": ".ST",
}


def normalize_holding_ticker(ticker: str, holding_name: str = "", geo: str = "") -> str:
    if not ticker:
        return ""
    
    ticker = ticker.strip()
    
    # Handle Bloomberg suffixes (e.g. "ALV GR")
    if " " in ticker:
        parts = ticker.split()
        if len(parts) == 2:
            symbol, suffix = parts[0], parts[1].upper()
            if suffix in BLOOMBERG_TO_YAHOO_SUFFIX:
                ticker = symbol + BLOOMBERG_TO_YAHOO_SUFFIX[suffix]
    
    # Handle trailing slashes
    if ticker.endswith("/"):
        ticker = ticker.rstrip("/")
        
    # Handle slash (e.g. "BRK/B" -> "BRK-B")
    if "/" in ticker:
        ticker = ticker.replace("/", "-")
        
    # Handle numeric tickers (e.g. "005930")
    if ticker.isdigit():
        geo_clean = (geo or "").strip().lower()
        if "korea" in geo_clean:
            ticker = ticker + ".KS"
        elif "taiwan" in geo_clean:
            ticker = ticker + ".TW"
        elif "hong kong" in geo_clean or "china" in geo_clean:
            ticker = ticker + ".HK"
        elif "japan" in geo_clean:
            ticker = ticker + ".T"
            
    return ticker


def clean_company_name(name: str) -> str:
    if not name:
        return ""
    val = re.sub(r"\s+", " ", name.strip()).upper().replace(".", "")
    val = val.replace(" AND ", " & ")
    val = LEGAL_SUFFIX_RE.sub(" ", val)
    val = re.sub(r"\s+", " ", val).strip(" -")
    return val


def resolve_ticker_by_name(holding_name: str, position_symbols: dict[str, str]) -> str:
    cleaned_holding = clean_company_name(holding_name)
    if not cleaned_holding:
        return ""
    # 1. Exact match on cleaned names
    for pos_name, symbol in position_symbols.items():
        if clean_company_name(pos_name) == cleaned_holding:
            return symbol
    # 2. Substring match (e.g. "APPLE" inside "APPLE INC" or vice versa)
    for pos_name, symbol in position_symbols.items():
        cleaned_pos = clean_company_name(pos_name)
        if not cleaned_pos:
            continue
        if cleaned_pos in cleaned_holding or cleaned_holding in cleaned_pos:
            return symbol
    return ""


SECTOR_MAPPING = {
    # Technology / IT
    "technology": "Information Technology",
    "information technology": "Information Technology",
    
    # Communication Services
    "communication": "Communication Services",
    "communication services": "Communication Services",
    "telecommunications": "Communication Services",
    
    # Financials
    "financials": "Financials",
    "diversified banks": "Financials",
    "property & casualty insurance": "Financials",
    "life & health insurance": "Financials",
    "asset management & custody banks": "Financials",
    "financial exchanges & data": "Financials",
    "multi-line insurance": "Financials",
    "transaction & payment processing services": "Financials",
    "investment banking & brokerage": "Financials",
    "regional banks": "Financials",
    "diversified financial services": "Financials",
    "insurance brokers": "Financials",
    "consumer finance": "Financials",
    "reinsurance": "Financials",
    "diversified capital markets": "Financials",
    "mortgage reits": "Financials",
    "commercial & residential mortgage finance": "Financials",
    "specialized finance": "Financials",
    
    # Cash / Currency
    "cash and/or derivatives": "Cash / Money Market",
    "cash / money market": "Cash / Money Market",
    "cash": "Cash / Money Market",
    "cash equivalent": "Cash / Money Market",
    "cash collateral and margins": "Cash / Money Market",
    "currency": "Cash / Money Market",
    "fx": "Cash / Money Market",
    
    # Unknown / Unclassified
    "unknown": "Unclassified",
    "unknown from issuer data": "Unclassified",
    "unclassified": "Unclassified",
}

def normalize_sector(name: str) -> str:
    if not name:
        return "Unclassified"
    key = name.strip().lower()
    return SECTOR_MAPPING.get(key, name.strip())


def distribution_rows(
    bucket: dict[str, Decimal],
    total: Decimal,
    name_key: str,
    source_assets: dict[str, set[str]] | None = None,
    tickers: dict[str, str] | None = None,
    asset_classes: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    rows = []
    for key, value in sorted(bucket.items(), key=lambda item: item[1], reverse=True):
        row = {
            name_key: key,
            "market_value_eur": money(value),
            "weight_pct": float((value / total * Decimal("100")).quantize(Decimal("0.01"))) if total > ZERO else 0.0,
        }
        if source_assets is not None:
            row["source_assets"] = sorted(source_assets.get(key, set()))
        if tickers is not None and key in tickers:
            row["holding_ticker"] = tickers[key]
        if asset_classes is not None and key in asset_classes:
            row["asset_class"] = asset_classes[key]
            row["asset_type"] = determine_asset_type_from_class(asset_classes[key])
        rows.append(row)
    return rows


FUND_DISPLAY_NAMES = {
    "IE00023EZQ82": "iShares Digital Entertainment and Education UCITS ETF",
    "IE000YDOORK7": "Xtrackers MSCI Fintech Innovation UCITS ETF 1C",
    "IE00B4L5Y983": "iShares Core MSCI World UCITS ETF",
    "IE00B5BMR087": "iShares Core S&P 500 UCITS ETF",
    "IE00B5MTXJ97": "Invesco STOXX Europe 600 Optimised Insurance UCITS ETF Acc",
    "IE00BGV5VN51": "Xtrackers Artificial Intelligence and Big Data UCITS ETF 1C",
    "IE00BK5BQT80": "Vanguard FTSE All-World UCITS ETF USD Accumulating",
    "IE00BKM4GZ66": "iShares Core MSCI Emerging Markets IMI UCITS ETF",
    "IE00BM67HL84": "Xtrackers MSCI World Financials UCITS ETF 1C",
    "LU0290358497": "Xtrackers II EUR Overnight Rate Swap UCITS ETF 1C",
    "LU1681045370": "Amundi MSCI Emerging Markets UCITS ETF",
    "NL0011683594": "VanEck Morningstar Developed Markets Dividend Leaders UCITS ETF",
}


def display_fund_name(asset: str, isin: str, document: dict[str, Any]) -> str:
    if isin in FUND_DISPLAY_NAMES:
        return FUND_DISPLAY_NAMES[isin]

    product_url = str(document.get("product_url") or "")
    if product_url:
        slug = urllib.parse.urlparse(product_url).path.rstrip("/").split("/")[-1]
        slug = re.sub(r"^[A-Z0-9]{12}-", "", slug)
        slug = slug.replace("ishares-", "iShares-")
        words = [word for word in slug.split("-") if word and word not in {"fund", "siteEntryPassthrough"}]
        if words:
            acronyms = {"ai", "eur", "usd", "ucits", "etf", "msci", "imi", "s&p", "stoxx", "ftse"}
            pretty_words = []
            for word in words:
                lower = word.lower()
                if lower in acronyms:
                    pretty_words.append(lower.upper())
                elif lower == "ishares":
                    pretty_words.append("iShares")
                elif lower == "xtrackers":
                    pretty_words.append("Xtrackers")
                elif lower == "vaneck":
                    pretty_words.append("VanEck")
                else:
                    pretty_words.append(lower.capitalize())
            return " ".join(pretty_words)

    return str(document.get("asset_name") or asset)


def composition_source_row(
    position: dict[str, Any],
    rows: list[dict[str, Any]] | None,
    document: dict[str, Any] | None,
) -> dict[str, Any] | None:
    asset = str(position.get("asset") or "").strip()
    isin = str(position.get("isin") or "").strip().upper()
    if not rows and not document:
        return None
    if rows and all(row.get("asset_class") == "Single share" for row in rows):
        return None
    if rows and all(row.get("asset_class") == "Cash equivalent" for row in rows) and not document:
        status = "cash_equivalent"
    else:
        status = str((document or {}).get("status") or ("official_metadata_missing" if rows else "official_source_not_found"))
    fund_name = display_fund_name(asset, isin, document or {})
    return {
        "asset": asset,
        "isin": isin,
        "fund_name": fund_name,
        "issuer": str((document or {}).get("issuer") or ""),
        "status": status,
        "fetched_at": str((document or {}).get("fetched_at") or ""),
        "source_url": str((document or {}).get("holdings_url") or (document or {}).get("product_url") or ""),
        "rows": int((document or {}).get("rows") or len(rows or [])),
        "weight_sum": str((document or {}).get("weight_sum") or ""),
        "message": str((document or {}).get("message") or ""),
    }


def calculate_distribution(
    positions: list[dict[str, Any]],
    exposures: dict[str, list[dict[str, Any]]],
    berkshire_mode: str = "stock",
    proxy_mode: str = "off",
) -> dict[str, Any]:
    documents = read_etf_documents()
    if berkshire_mode == "lookthrough":
        metadata = read_berkshire_metadata()
        if metadata:
            documents[BERKSHIRE_ISIN] = metadata
    if proxy_mode == "on":
        for isin, metadata in read_proxy_metadata().items():
            if is_full_official_composition(documents.get(isin)):
                continue
            documents[isin] = metadata
    position_symbols = {p["asset"]: p["symbol"] for p in positions if p.get("symbol")}
    open_positions = [
        position
        for position in positions
        if position.get("is_open") and position.get("market_value_eur") is not None and Decimal(str(position["market_value_eur"])) > ZERO
    ]
    total_value = sum((Decimal(str(position["market_value_eur"])) for position in open_positions), ZERO)
    by_holding: dict[str, Decimal] = {}
    by_holding_sources: dict[str, set[str]] = {}
    holding_tickers: dict[str, str] = {}
    holding_classes: dict[str, str] = {}
    by_sector: dict[str, Decimal] = {}
    by_sector_holdings: dict[str, dict[str, Decimal]] = {}
    by_geo: dict[str, Decimal] = {}
    by_asset_class: dict[str, Decimal] = {}
    missing = []
    covered_assets = 0
    exploded_rows = []
    composition_sources = []

    for position in open_positions:
        value = Decimal(str(position["market_value_eur"]))
        asset = str(position.get("asset") or "").strip()
        isin = str(position.get("isin") or "").strip().upper()
        rows = exposures.get(exposure_key(asset, isin)) or exposures.get(exposure_key(asset, ""))
        document = documents.get(isin)
        source_row = composition_source_row(position, rows, document)
        if source_row:
            composition_sources.append(source_row)
        if not rows:
            missing.append(
                {
                    "asset": asset,
                    "isin": isin,
                    "market_value_eur": money(value),
                    "status": str((document or {}).get("status") or "missing_exposure_row"),
                    "message": str((document or {}).get("message") or "No distribution row is available."),
                }
            )
            fallback_holding = asset or isin or "Unclassified"
            fallback_sector = normalize_sector(str(position.get("sector") or "Unclassified"))
            fallback_geo = str(position.get("geo") or "Unclassified")
            fallback_asset_class = str(position.get("asset_class") or "Unclassified")
            add_distribution_amount(by_holding, fallback_holding, value)
            add_distribution_amount(by_sector, fallback_sector, value)
            by_sector_holdings.setdefault(fallback_sector, {})[fallback_holding] = \
                by_sector_holdings.setdefault(fallback_sector, {}).get(fallback_holding, ZERO) + value
            add_distribution_amount(by_geo, fallback_geo, value)
            add_distribution_amount(by_asset_class, fallback_asset_class, value)
            fallback_ticker = normalize_holding_ticker(str(position.get("symbol") or ""), fallback_holding, fallback_geo)
            holding_tickers[fallback_holding] = fallback_ticker
            holding_classes[fallback_holding] = fallback_asset_class
            exploded_rows.append(
                {
                    "source_asset": asset,
                    "source_isin": isin,
                    "holding_name": fallback_holding,
                    "holding_ticker": fallback_ticker,
                    "market_value_eur": money(value),
                    "weight_pct": float((value / total_value * Decimal("100")).quantize(Decimal("0.01"))) if total_value > ZERO else 0.0,
                    "sector": fallback_sector,
                    "geo": fallback_geo,
                    "asset_class": fallback_asset_class,
                }
            )
            continue

        covered_assets += 1
        is_single_share = all(row.get("asset_class") == "Single share" for row in rows)
        source_label = composition_source_label(position, is_single_share)
        row_weight = sum((row["weight_pct"] for row in rows), ZERO)
        divisor = row_weight if row_weight > Decimal("100") else Decimal("100")
        allocated = ZERO
        for row in rows:
            amount = value * row["weight_pct"] / divisor
            allocated += amount
            raw_holding = row["holding_name"] or asset
            if raw_holding == "Other issuer holdings":
                holding = f"Other issuer holdings - {asset}"
            else:
                holding = canonical_holding_name(raw_holding)
            add_distribution_amount(by_holding, holding, amount)
            if source_label:
                by_holding_sources.setdefault(holding, set()).add(source_label)
            norm_sector = normalize_sector(row["sector"])
            add_distribution_amount(by_sector, norm_sector, amount)
            by_sector_holdings.setdefault(norm_sector, {})[holding] = \
                by_sector_holdings.setdefault(norm_sector, {}).get(holding, ZERO) + amount
            add_distribution_amount(by_geo, row["geo"], amount)
            add_distribution_amount(by_asset_class, row["asset_class"], amount)
            
            ticker = row["holding_ticker"]
            if not ticker:
                ticker = holding_tickers.get(holding, "")
            if not ticker:
                ticker = COMMON_HOLDING_TICKERS.get(clean_company_name(holding), "")
            if not ticker:
                ticker = resolve_ticker_by_name(holding, position_symbols)
                
            ticker = normalize_holding_ticker(ticker, holding, row["geo"])
            holding_tickers[holding] = ticker
            holding_classes[holding] = row["asset_class"]
            exploded_rows.append(
                {
                    "source_asset": asset,
                    "source_isin": isin,
                    "holding_name": holding,
                    "holding_ticker": ticker,
                    "market_value_eur": money(amount),
                    "weight_pct": float((amount / total_value * Decimal("100")).quantize(Decimal("0.01"))) if total_value > ZERO else 0.0,
                    "sector": row["sector"],
                    "geo": row["geo"],
                    "asset_class": row["asset_class"],
                }
            )

        remainder = max(ZERO, value - allocated)
        if remainder > Decimal("0.01"):
            label = f"Other issuer holdings - {asset}"
            add_distribution_amount(by_holding, label, remainder)
            if asset:
                by_holding_sources.setdefault(label, set()).add(asset)
            norm_remainder_sector = normalize_sector("Unknown from issuer data")
            add_distribution_amount(by_sector, norm_remainder_sector, remainder)
            by_sector_holdings.setdefault(norm_remainder_sector, {})[label] = \
                by_sector_holdings.setdefault(norm_remainder_sector, {}).get(label, ZERO) + remainder
            add_distribution_amount(by_geo, "Unknown from issuer data", remainder)
            add_distribution_amount(by_asset_class, "ETF underlying", remainder)
            holding_tickers[label] = ""
            holding_classes[label] = "ETF underlying"

    return {
        "source_file": configured_path_label(EXPOSURES_CSV),
        "documents_file": configured_path_label(ETF_DOCUMENTS_JSON) if ETF_DOCUMENTS_JSON.exists() else "",
        "total_value_eur": money(total_value),
        "covered_assets": covered_assets,
        "open_assets": len(open_positions),
        "missing": sorted(missing, key=lambda item: item["market_value_eur"], reverse=True),
        "composition_sources": sorted(composition_sources, key=lambda item: (item["status"] != "ok", item["asset"])),
        "composition_source_coverage": {
            "resolved": sum(
                1
                for item in composition_sources
                if item["status"]
                in {"ok", "cached", "cash_equivalent", "partial_official_holdings", "official_sec_13f", "proxy_exposure"}
            ),
            "total": len(composition_sources),
        },
        "underlying": distribution_rows(by_holding, total_value, "holding", by_holding_sources, holding_tickers, holding_classes)[:30],
        "sectors": [{
            **s_row,
            "holdings": [{
                "holding": h_name,
                "holding_ticker": holding_tickers.get(h_name, ""),
                "market_value_eur": money(h_val),
                "weight_pct": float((h_val / total_value * Decimal("100")).quantize(Decimal("0.01"))) if total_value > ZERO else 0.0,
            } for h_name, h_val in sorted(by_sector_holdings.get(s_row["sector"], {}).items(), key=lambda x: x[1], reverse=True)]
        } for s_row in distribution_rows(by_sector, total_value, "sector")],
        "geographies": distribution_rows(by_geo, total_value, "geo"),
        "asset_classes": distribution_rows(by_asset_class, total_value, "asset_class"),
        "rows": sorted(exploded_rows, key=lambda item: item["market_value_eur"], reverse=True)[:100],
        "source_rows": sorted(exploded_rows, key=lambda item: (item["source_asset"], -item["market_value_eur"])),
    }


def trade_cash_amount(trade: Trade) -> tuple[Decimal, bool]:
    if trade.grand_total_present:
        return convert_cash_to_eur(abs(trade.grand_total), trade.cash_currency, trade.date)
    if trade.total_spend:
        return convert_cash_to_eur(abs(trade.total_spend), trade.cash_currency, trade.date)
    if trade.price and trade.quantity_diff:
        return convert_cash_to_eur(abs(trade.price * trade.quantity_diff), trade.cash_currency, trade.date)
    return ZERO, False


def normalize_currency_code(currency: Any) -> str:
    raw = str(currency or "EUR").strip()
    if not raw:
        return "EUR"
    upper = raw.upper()
    if raw in {"GBp", "GBX"} or upper in {"GBX", "GBPENCE"}:
        return "GBp"
    return upper


def fx_base_currency(currency: Any) -> str:
    normalized = normalize_currency_code(currency)
    return "GBP" if normalized == "GBp" else normalized


def fx_symbol_for(currency: str) -> tuple[str, bool]:
    currency = fx_base_currency(currency)
    if currency == "USD":
        return "EURUSD=X", True
    return f"{currency}EUR=X", False


def convert_cash_to_eur(amount: Decimal, currency: str, target_date: date) -> tuple[Decimal, bool]:
    currency = fx_base_currency(currency)
    if currency in {"", "E", "EUR"}:
        return amount, False
    symbol, invert = fx_symbol_for(currency)
    history = fetch_history(symbol, target_date - timedelta(days=10), target_date, refresh=False)
    rate = previous_price(history.get("prices", {}), target_date)
    if rate is None:
        return amount, True
    if invert:
        rate = 1 / rate
    return amount * Decimal(str(rate)), False


def summarize_trades(trades: list[Trade]) -> dict[str, Any]:
    quantities: dict[str, Decimal] = {}
    cost_basis: dict[str, Decimal] = {}
    realized_by_asset: dict[str, Decimal] = {}
    realized_cost_by_asset: dict[str, Decimal] = {}
    position_names: dict[str, str] = {}
    position_isins: dict[str, str] = {}
    realized_pl = ZERO
    invested = ZERO
    proceeds = ZERO
    fees = ZERO
    taxes = ZERO
    approximate_cost_basis = ZERO
    series: list[dict[str, Any]] = []

    for trade in trades:
        asset = trade.isin or trade.asset
        position_names.setdefault(asset, trade.asset)
        position_isins.setdefault(asset, trade.isin)
        quantities.setdefault(asset, ZERO)
        cost_basis.setdefault(asset, ZERO)
        realized_by_asset.setdefault(asset, ZERO)
        realized_cost_by_asset.setdefault(asset, ZERO)
        amount, approximate = trade_cash_amount(trade)
        if approximate:
            approximate_cost_basis += amount

        fees += trade.fees
        taxes += trade.tax

        if trade.quantity_diff >= ZERO:
            invested += amount
            quantities[asset] += trade.quantity_diff
            cost_basis[asset] += amount
        else:
            sell_quantity = abs(trade.quantity_diff)
            previous_quantity = quantities[asset]
            previous_cost = cost_basis[asset]
            if previous_quantity > ZERO and sell_quantity > ZERO:
                removed_cost = min(previous_cost, previous_cost * sell_quantity / previous_quantity)
            else:
                removed_cost = ZERO
            realized_delta = amount - removed_cost
            if trade.broker == "Fineco" and trade.tax == ZERO and realized_delta > ZERO:
                taxes += fineco_capital_gain_tax_from_net_gain(realized_delta)
            proceeds += amount
            cost_basis[asset] = max(ZERO, previous_cost - removed_cost)
            quantities[asset] = previous_quantity - sell_quantity
            realized_pl += realized_delta
            realized_by_asset[asset] += realized_delta
            realized_cost_by_asset[asset] += removed_cost

        series.append(
            {
                "date": trade.date.isoformat(),
                "invested": money(invested),
                "proceeds": money(proceeds),
                "net_contributions": money(invested - proceeds),
                "open_cost_basis": money(sum(cost_basis.values(), ZERO)),
                "realized_pl": money(realized_pl),
                "fees": money(fees),
                "taxes": money(taxes),
            }
        )

    positions = []
    for asset in sorted(quantities):
        quantity = quantities[asset]
        if abs(quantity) < Decimal("0.00000001"):
            quantity = ZERO
        realized_cost = realized_cost_by_asset.get(asset, ZERO)
        realized_value = realized_by_asset.get(asset, ZERO)
        positions.append(
            {
                "asset": position_names.get(asset, asset),
                "isin": position_isins.get(asset, ""),
                "quantity": decimal_to_float(quantity),
                "cost_basis_eur": money(cost_basis[asset]),
                "is_open": quantity > ZERO,
                "realized_pl_eur": money(realized_value),
                "realized_pl_pct": float((realized_value / realized_cost * Decimal("100")).quantize(Decimal("0.01"))) if realized_cost > ZERO else None,
            }
        )

    return {
        "positions": positions,
        "series": compress_series(series),
        "totals": {
            "invested": money(invested),
            "proceeds": money(proceeds),
            "net_contributions": money(invested - proceeds),
            "open_cost_basis": money(sum(cost_basis.values(), ZERO)),
            "realized_pl": money(realized_pl),
            "fees": money(fees),
            "taxes": money(taxes),
            "approximate_cost_basis": money(approximate_cost_basis),
        },
    }


def compress_series(series: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_date: dict[str, dict[str, Any]] = {}
    for point in series:
        by_date[point["date"]] = point
    return [by_date[key] for key in sorted(by_date)]


def extract_symbol(search_result: Any) -> dict[str, str] | None:
    quotes = getattr(search_result, "quotes", None)
    if quotes is None and isinstance(search_result, dict):
        quotes = search_result.get("quotes")
    if not quotes:
        return None

    for quote in quotes:
        symbol = quote.get("symbol") if isinstance(quote, dict) else getattr(quote, "symbol", "")
        if not symbol:
            continue
        name = quote.get("shortname") if isinstance(quote, dict) else getattr(quote, "shortname", "")
        exchange = quote.get("exchange") if isinstance(quote, dict) else getattr(quote, "exchange", "")
        return {"symbol": symbol, "name": name or "", "exchange": exchange or ""}
    return None


def yahoo_symbol_from_mapping(ticker: str, exchange: str) -> str:
    ticker = (ticker or "").strip()
    exchange = (exchange or "").strip().upper()
    if not ticker:
        return ""
    if "." in ticker or "=" in ticker or "-" in ticker:
        return ticker
    if exchange in {"BVME", "BIT", "MILAN", "MILANO", "BORSA ITALIANA"}:
        return f"{ticker}.MI"
    return ticker


def mapping_for(asset: str, isin: str, mappings: dict[str, dict[str, str]]) -> dict[str, str]:
    if asset in mappings:
        return mappings[asset]
    if isin:
        for mapping in mappings.values():
            if mapping.get("isin") == isin:
                return mapping
    return {}


_symbol_cache_in_memory = None


def get_symbol_cache():
    global _symbol_cache_in_memory
    if _symbol_cache_in_memory is None:
        _symbol_cache_in_memory = load_json(SYMBOL_CACHE)
    return _symbol_cache_in_memory


_price_cache_in_memory = None


def get_price_cache():
    global _price_cache_in_memory
    if _price_cache_in_memory is None:
        _price_cache_in_memory = load_json(PRICE_CACHE)
    return _price_cache_in_memory


_news_cache_in_memory = None


def get_news_cache():
    global _news_cache_in_memory
    if _news_cache_in_memory is None:
        _news_cache_in_memory = load_json(NEWS_CACHE)
    return _news_cache_in_memory


def resolve_isin(isin: str, refresh: bool = False, direct_symbol: str = "") -> dict[str, Any]:
    if direct_symbol:
        return {"isin": isin, "symbol": direct_symbol, "status": "resolved", "source": "mapping"}
    if not isin:
        return {"status": "missing_isin"}
    if yf is None:
        return {"status": "yfinance_missing"}

    cache = get_symbol_cache()
    cached = cache.get(isin)
    if cached and not refresh:
        return {**cached, "status": cached.get("status", "resolved")}

    try:
        result = yf.Search(
            isin,
            max_results=5,
            news_count=0,
            lists_count=0,
            include_cb=False,
            include_nav_links=False,
            include_research=False,
            include_cultural_assets=False,
            timeout=5,
            raise_errors=False,
        )
        symbol = extract_symbol(result)
        if symbol:
            payload = {**symbol, "isin": isin, "status": "resolved", "resolved_at": int(time.time())}
        else:
            payload = {"isin": isin, "status": "unresolved", "resolved_at": int(time.time())}
    except Exception as exc:
        payload = {"isin": isin, "status": "lookup_error", "error": str(exc), "resolved_at": int(time.time())}

    cache[isin] = payload
    save_json(SYMBOL_CACHE, cache)
    return payload


def fast_info_value(info: Any, key: str) -> Any:
    try:
        if isinstance(info, dict):
            return info.get(key)
        return info[key]
    except Exception:
        return getattr(info, key, None)


def infer_currency(symbol: str) -> str:
    symbol = symbol.upper()
    if symbol.endswith((".DE", ".MI", ".AS", ".PA")):
        return "EUR"
    if symbol.endswith(".L"):
        return "GBp"
    if symbol.endswith("=X"):
        return "EUR"
    return "USD"


def fetch_price(symbol: str, refresh: bool = False) -> dict[str, Any]:
    if not symbol:
        return {"status": "missing_symbol"}
    if yf is None:
        return {"status": "yfinance_missing"}

    cache = get_price_cache()
    cached = cache.get(symbol)
    now = int(time.time())
    if cached and cached.get("status") == "priced" and not refresh:
        return {**cached, "cache_stale": now - int(cached.get("fetched_at", 0)) >= PRICE_TTL_SECONDS}

    try:
        ticker = yf.Ticker(symbol)
        currency = infer_currency(symbol)
        try:
            info = ticker.fast_info
            price = fast_info_value(info, "last_price") or fast_info_value(info, "lastPrice")
            currency = fast_info_value(info, "currency") or currency
        except Exception:
            price = None
        if price is None or (isinstance(price, float) and math.isnan(price)):
            history = ticker.history(period="10d")
            if history.empty:
                raise ValueError("No last price returned.")
            price = float(history["Close"].dropna().iloc[-1])
        payload = {
            "symbol": symbol,
            "price": float(price),
            "currency": normalize_currency_code(currency),
            "status": "priced",
            "fetched_at": now,
            "price_date": date.today().isoformat(),
        }
    except Exception as exc:
        fallback = latest_cached_history_price(symbol)
        if fallback:
            cache[symbol] = fallback
            save_json(PRICE_CACHE, cache)
            return fallback
        payload = {"symbol": symbol, "status": "price_error", "error": str(exc), "fetched_at": now}

    cache[symbol] = payload
    save_json(PRICE_CACHE, cache)
    return payload


_history_store: HistoryStore | None = None


def get_history_store() -> HistoryStore:
    global _history_store
    if _history_store is None:
        _history_store = HistoryStore(HISTORY_CACHE_DIR, legacy_file=HISTORY_CACHE)
    return _history_store


def fetch_history(symbol: str, start: date, end: date, refresh: bool = False) -> dict[str, Any]:
    if not symbol:
        return {"status": "missing_symbol", "prices": {}}
    if yf is None:
        return {"status": "yfinance_missing", "prices": {}}

    store = get_history_store()
    now = int(time.time())
    cached = None if refresh else store.get_cached(symbol, start, end)
    if cached:
        return cached

    try:
        ticker = yf.Ticker(symbol)
        history = ticker.history(start=start.isoformat(), end=(end + timedelta(days=1)).isoformat(), auto_adjust=False)
        if history.empty:
            raise ValueError("No historical prices returned.")
        prices = {
            idx.date().isoformat(): float(value)
            for idx, value in history["Close"].dropna().items()
            if value is not None and not math.isnan(float(value))
        }
        try:
            info = ticker.fast_info
            currency = normalize_currency_code(fast_info_value(info, "currency") or infer_currency(symbol))
        except Exception:
            currency = infer_currency(symbol)
        payload = {
            "symbol": symbol,
            "currency": currency,
            "prices": prices,
            "status": "priced",
            "fetched_at": now,
            "range_start": start.isoformat(),
            "range_end": end.isoformat(),
        }
    except Exception as exc:
        payload = {"symbol": symbol, "status": "history_error", "error": str(exc), "prices": {}, "fetched_at": now}

    return store.merge(symbol, payload, start, end)


_PRICE_LOOKUP_CACHE: dict[int, tuple[dict[str, float], int, tuple[date, ...], tuple[float, ...]]] = {}


def price_lookup_index(prices: dict[str, float]) -> tuple[tuple[date, ...], tuple[float, ...]]:
    cache_key = id(prices)
    cached = _PRICE_LOOKUP_CACHE.get(cache_key)
    if cached and cached[0] is prices and cached[1] == len(prices):
        return cached[2], cached[3]

    pairs: list[tuple[date, float]] = []
    for raw_date, value in prices.items():
        if value is None:
            continue
        try:
            parsed_date = date.fromisoformat(str(raw_date))
            parsed_value = float(value)
        except (TypeError, ValueError):
            continue
        if not math.isnan(parsed_value):
            pairs.append((parsed_date, parsed_value))
    pairs.sort(key=lambda item: item[0])
    dates = tuple(item[0] for item in pairs)
    values = tuple(item[1] for item in pairs)

    if len(_PRICE_LOOKUP_CACHE) > 256:
        _PRICE_LOOKUP_CACHE.clear()
    _PRICE_LOOKUP_CACHE[cache_key] = (prices, len(prices), dates, values)
    return dates, values


def previous_price_point(prices: dict[str, float], target: date) -> tuple[float, date] | None:
    dates, values = price_lookup_index(prices)
    if not dates:
        return None
    idx = bisect_right(dates, target) - 1
    if idx < 0:
        return None
    return values[idx], dates[idx]


def previous_price(prices: dict[str, float], target: date) -> float | None:
    point = previous_price_point(prices, target)
    if point is None:
        return None
    return point[0]


def price_payload_for_target(payload: dict[str, Any] | None, target: date) -> float | None:
    if not payload or payload.get("status") != "priced":
        return None
    source = payload.get("source")
    if source == "history_cache" and payload.get("price_date") != target.isoformat():
        return None
    try:
        price = float(payload["price"])
    except (KeyError, TypeError, ValueError):
        return None
    if math.isnan(price):
        return None
    return price


def price_point_for_date(
    prices: dict[str, float],
    target: date,
    current_price: dict[str, Any] | None = None,
) -> tuple[float, date] | None:
    current = price_payload_for_target(current_price, target)
    if current is not None and target == date.today():
        return current, target
    return previous_price_point(prices, target)


def current_price_overrides(
    symbol_histories: dict[str, dict[str, Any]],
    target: date,
    refresh: bool = False,
) -> dict[str, dict[str, Any]]:
    if target != date.today():
        return {}
    overrides: dict[str, dict[str, Any]] = {}
    for symbol in sorted(symbol_histories):
        history = symbol_histories.get(symbol) or {}
        if history.get("status") != "priced" or not history.get("prices"):
            continue
        price_data = fetch_price(symbol, refresh=refresh)
        if price_payload_for_target(price_data, target) is not None:
            overrides[symbol] = price_data
    return overrides


def benchmark_delta(latest: dict[str, Any], past: dict[str, Any], key: str) -> float | None:
    latest_value = latest.get(f"{key}_return_pct")
    past_value = past.get(f"{key}_return_pct")
    if latest_value is None or past_value is None:
        return None
    latest_price_date = latest.get(f"{key}_price_date")
    past_price_date = past.get(f"{key}_price_date")
    if latest_price_date is not None and past_price_date is not None and latest_price_date <= past_price_date:
        return None
    denom = 100.0 + float(past_value)
    if denom <= 0:
        return None
    return ((100.0 + float(latest_value)) / denom - 1.0) * 100.0



def fetch_crypto_eur_history(asset: str, start: date, end: date, refresh: bool = False) -> dict[str, Any]:
    asset = (asset or "").upper()
    store = get_history_store()
    key = f"crypto-eur:{asset}"
    now = int(time.time())
    cached = None if refresh else store.get_cached(key, start, end)
    if cached:
        return cached

    import urllib.parse
    import urllib.request

    prices: dict[str, float] = {}
    try:
        if asset == "USDC":
            fx_history = fetch_history("EURUSD=X", start, end, refresh=refresh)
            if fx_history.get("status") == "priced" and fx_history.get("prices"):
                for d_str, val in fx_history["prices"].items():
                    if val > 0:
                        prices[d_str] = float(Decimal("1") / Decimal(str(val)))
        elif asset in {"BTC", "ETH", "TON"}:
            quote_currency = "USD" if asset == "TON" else "EUR"
            fx_history = fetch_history("EURUSD=X", start, end, refresh=refresh) if quote_currency == "USD" else {"prices": {}}
            cursor = start
            while cursor <= end:
                chunk_end = min(end, cursor + timedelta(days=299))
                params = urllib.parse.urlencode(
                    {
                        "start": datetime.combine(cursor, datetime.min.time()).replace(tzinfo=timezone.utc).isoformat(),
                        "end": datetime.combine(chunk_end, datetime.min.time()).replace(tzinfo=timezone.utc).isoformat(),
                        "granularity": "86400",
                    }
                )
                url = f"https://api.exchange.coinbase.com/products/{asset}-{quote_currency}/candles?{params}"
                req = urllib.request.Request(url, headers={"User-Agent": "portfolio-dashboard/1.0"})
                with urllib.request.urlopen(req, timeout=20) as response:
                    candles = json.loads(response.read().decode("utf-8"))
                if isinstance(candles, list):
                    for candle in candles:
                        if not isinstance(candle, list) or len(candle) < 5:
                            continue
                        candle_date = datetime.fromtimestamp(int(candle[0]), tz=timezone.utc).date()
                        close = float(candle[4])
                        if quote_currency == "USD":
                            eurusd = previous_price(fx_history.get("prices", {}), candle_date)
                            if eurusd:
                                close = close / eurusd
                            else:
                                continue
                        prices[candle_date.isoformat()] = close
                cursor = chunk_end + timedelta(days=1)
        payload = {
            "symbol": f"{asset}-EUR",
            "currency": "EUR",
            "prices": prices,
            "status": "priced" if prices else "history_error",
            "fetched_at": now,
        }
    except Exception as exc:
        payload = {"symbol": f"{asset}-EUR", "status": "history_error", "error": str(exc), "prices": {}, "fetched_at": now}

    return store.merge(key, payload, start, end)


def latest_cached_history_price(symbol: str) -> dict[str, Any] | None:
    cached = get_history_store().latest_price(symbol)
    if cached is None:
        return None
    return {
        "symbol": symbol,
        "price": float(cached["price"]),
        "currency": normalize_currency_code(cached.get("currency") or infer_currency(symbol)),
        "status": "priced",
        "fetched_at": int(time.time()),
        "source": "history_cache",
        "price_date": cached["price_date"],
    }


def month_end_dates(start: date, end: date) -> set[date]:
    dates: set[date] = set()
    cursor = date(start.year, start.month, 1)
    while cursor <= end:
        if cursor.month == 12:
            next_month = date(cursor.year + 1, 1, 1)
        else:
            next_month = date(cursor.year, cursor.month + 1, 1)
        month_end = next_month - timedelta(days=1)
        if start <= month_end <= end:
            dates.add(month_end)
        cursor = next_month
    dates.add(end)
    return dates


def fx_to_eur(currency: str, refresh: bool = False) -> dict[str, Any]:
    currency = fx_base_currency(currency)
    if currency == "EUR":
        return {"rate": 1.0, "status": "native"}
    symbol, invert = fx_symbol_for(currency)
    price = fetch_price(symbol, refresh=refresh)
    if price.get("status") == "priced":
        rate = float(price["price"])
        return {"rate": 1 / rate if invert else rate, "status": "priced", "symbol": symbol}
    return {"rate": None, "status": price.get("status", "fx_error"), "symbol": symbol}


def build_instrument_refs(
    trades: list[Trade], mappings: dict[str, dict[str, str]], refresh: bool = False
) -> dict[str, dict[str, str]]:
    refs: dict[str, dict[str, str]] = {}
    for trade in trades:
        key = trade.isin or trade.asset
        if key in refs:
            continue
        mapping = mapping_for(trade.asset, trade.isin, mappings)
        direct_symbol = yahoo_symbol_from_mapping(mapping.get("ticker", ""), mapping.get("exchange", ""))
        symbol_data = resolve_isin(trade.isin or mapping.get("isin", ""), refresh=refresh, direct_symbol=direct_symbol)
        refs[key] = {
            "asset": trade.asset,
            "isin": trade.isin or mapping.get("isin", ""),
            "symbol": symbol_data.get("symbol", ""),
            "status": symbol_data.get("status", "missing_isin"),
        }
    return refs


def historical_fx_rate(
    currency: str,
    fx_histories: dict[str, dict[str, Any]],
    target: date,
    live_fx_prices: dict[str, dict[str, Any]] | None = None,
) -> float | None:
    currency = normalize_currency_code(currency)
    if currency == "EUR":
        return 1.0
    if currency == "GBp":
        currency = "GBP"
    symbol, invert = fx_symbol_for(currency)
    history = fx_histories.get(symbol, {})
    price_point = price_point_for_date(history.get("prices", {}), target, (live_fx_prices or {}).get(symbol))
    rate = price_point[0] if price_point else None
    if rate is None:
        return None
    return 1 / rate if invert else rate


EUROSTAT_CACHE_PATH = SETTINGS.cache_path("eurostat-cpi.json")

def fetch_eurostat_cpi() -> dict[str, float]:
    import json
    import urllib.request
    now = int(time.time())
    if EUROSTAT_CACHE_PATH.exists():
        try:
            with open(EUROSTAT_CACHE_PATH, "r") as f:
                cached = json.load(f)
            if now - cached.get("fetched_at", 0) < 86400:  # 24 hours TTL
                return cached.get("data", {})
        except Exception:
            pass

    try:
        url = 'https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/prc_hicp_minr?format=JSON&lang=EN&geo=EA&coicop18=TOTAL&freq=M&unit=I15'
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as res:
            raw = json.loads(res.read().decode("utf-8"))
        
        time_dim = raw['dimension']['time']['category']
        labels = sorted(time_dim['index'].keys(), key=lambda k: time_dim['index'][k])
        values = raw['value']
        
        cpi_data = {}
        for label in labels:
            idx = str(time_dim['index'][label])
            val = values.get(idx)
            if val is not None:
                cpi_data[label] = float(val)
                
        if cpi_data:
            payload = {"fetched_at": now, "data": cpi_data}
            EUROSTAT_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
            save_json(EUROSTAT_CACHE_PATH, payload)
            return cpi_data
    except Exception as e:
        print("Error fetching Eurostat HICP CPI:", e)
        
    if EUROSTAT_CACHE_PATH.exists():
        try:
            with open(EUROSTAT_CACHE_PATH, "r") as f:
                return json.load(f).get("data", {})
        except Exception:
            pass
            
    return {}


def load_cash_histories(person: str) -> tuple[list[tuple[date, Decimal, Decimal]], list[tuple[date, Decimal, Decimal]], list[dict[str, Any]]]:
    trade_republic = read_ledger_cash_events(person, "Trade Republic")
    bbva = read_ledger_cash_events(person, "BBVA")
    revolut = read_ledger_cash_events(person, "Revolut")
    return (
        [(item["date"], item["cash_change"], item["contrib_change"]) for item in trade_republic],
        [(item["date"], item["cash_change"], item["contrib_change"]) for item in bbva],
        revolut,
    )


def load_cash_histories_from_raw_files(person: str) -> tuple[list[tuple[date, Decimal, Decimal]], list[tuple[date, Decimal, Decimal]], list[dict[str, Any]]]:
    """Legacy statement parser retained for migration diagnostics only."""

    tr_cash_history: list[tuple[date, Decimal, Decimal]] = []
    tr_file = latest_trade_republic_export() if person == PRIMARY_PORTFOLIO_ID else latest_family_trade_republic_export(person)

    if tr_file and tr_file.exists():
        try:
            with tr_file.open(newline="", encoding="utf-8-sig") as handle:
                for row in csv.DictReader(handle):
                    dt_str = row.get("date")
                    if not dt_str:
                        continue
                    t_date = datetime.strptime(dt_str, "%Y-%m-%d").date()
                    amount = parse_decimal(row.get("amount"))
                    fee = parse_decimal(row.get("fee"))
                    tax = parse_decimal(row.get("tax"))
                    cash_change = amount + fee + tax

                    t_type = row.get("type", "")
                    contrib_change = ZERO
                    if t_type in {"CUSTOMER_INBOUND", "CUSTOMER_INPAYMENT", "TRANSFER_INSTANT_INBOUND", "VIBAN_TRANSFER_INBOUND"}:
                        contrib_change = cash_change
                    elif t_type in {"CUSTOMER_OUTBOUND_REQUEST", "TRANSFER_INSTANT_OUTBOUND", "CARD_TRANSACTION", "CARD_TRANSACTION_INTERNATIONAL", "CARD_ORDERING_FEE"}:
                        contrib_change = cash_change

                    tr_cash_history.append((t_date, cash_change, contrib_change))
        except Exception:
            pass
    tr_cash_history.sort(key=lambda item: item[0])

    bbva_cash_history: list[tuple[date, Decimal, Decimal]] = []
    if person == PRIMARY_PORTFOLIO_ID:
        bbva_files = bbva_statement_files()
        if bbva_files:
            bbva_file = bbva_files[-1]
            import shutil
            import tempfile

            temp_xlsx = Path(tempfile.gettempdir()) / f"bbva_temp_{os.getpid()}_{threading.get_ident()}.xlsx"
            try:
                shutil.copy(bbva_file, temp_xlsx)
                wb = load_workbook(temp_xlsx, read_only=True)
                sheet = wb.active
                for row in sheet.iter_rows(values_only=True):
                    if len(row) < 7:
                        continue
                    _, op_date_str, causale, _, _, imp = row[1:7]
                    if causale and causale != "Causale" and imp:
                        try:
                            op_date = datetime.strptime(str(op_date_str).strip(), "%d/%m/%Y").date()
                        except Exception:
                            continue
                        try:
                            clean_imp = str(imp).replace(" EUR", "").replace(",", ".").strip()
                            amount = Decimal(clean_imp)
                        except Exception:
                            continue

                        if amount != ZERO:
                            cash_change = amount
                            contrib_change = ZERO if "INTERESSI" in str(causale).upper() else amount
                            bbva_cash_history.append((op_date, cash_change, contrib_change))
            except Exception:
                pass
            finally:
                if temp_xlsx.exists():
                    try:
                        temp_xlsx.unlink()
                    except OSError:
                        pass
    bbva_cash_history.sort(key=lambda item: item[0])
    revolut_cash_history = read_revolut_cash_events() if person == PRIMARY_PORTFOLIO_ID else []
    revolut_cash_history.sort(key=lambda item: item["datetime"])
    return tr_cash_history, bbva_cash_history, revolut_cash_history


def calculate_valuation_series(
    trades: list[Trade],
    mappings: dict[str, dict[str, str]],
    refresh: bool = False,
    person: str = PRIMARY_PORTFOLIO_ID,
    broker: str = "all",
) -> dict[str, Any]:
    if not trades:
        return {"series": [], "status": "empty"}

    start = min(trade.date for trade in trades)
    
    # Load dividends and cash interests to calculate Total Return series
    dividends = read_portfolio_dividends(person)
    interests = read_cash_interests(person)
    if broker != "all":
        dividends = [d for d in dividends if d.broker.lower() == broker]
        interests = [i for i in interests if i.get("broker", "").lower() == broker]

    tr_cash_history, bbva_cash_history, revolut_cash_history = load_cash_histories(person)

    end = date.today()
    refs = build_instrument_refs(trades, mappings, refresh=refresh)
    recent_dates = {end - timedelta(days=i) for i in range(35) if end - timedelta(days=i) >= start}
    valuation_dates = month_end_dates(start, end) | {trade.date for trade in trades} | recent_dates
    sorted_dates = sorted(valuation_dates)

    histories: dict[str, dict[str, Any]] = {}
    currencies: set[str] = set()
    for ref in refs.values():
        symbol = ref.get("symbol", "")
        if not symbol:
            continue
        history = fetch_history(symbol, start, end, refresh=refresh)
        histories[symbol] = history
        if history.get("status") == "priced":
            currencies.add(normalize_currency_code(history.get("currency") or "EUR"))

    fx_histories: dict[str, dict[str, Any]] = {}
    for currency in currencies:
        normalized = fx_base_currency(currency)
        if normalized and normalized != "EUR":
            symbol, _ = fx_symbol_for(normalized)
            fx_histories[symbol] = fetch_history(symbol, start, end, refresh=refresh)
    add_revolut_fx_histories(revolut_cash_history, start, end, fx_histories, refresh=refresh)

    holdings: dict[str, Decimal] = {}
    invested_by_asset: dict[str, Decimal] = {}
    invested = ZERO
    proceeds = ZERO
    invested_tr = ZERO
    proceeds_tr = ZERO
    invested_other = ZERO
    proceeds_other = ZERO
    trade_index = 0
    ordered_trades = sorted(trades, key=lambda item: (item.date, item.broker, item.asset, item.action))
    series: list[dict[str, Any]] = []

    # Fetch MSCI World prices
    msci_prices = {}
    msci_symbol = ""
    for sym in ["SWDA.MI", "EUNL.DE", "URTH"]:
        h = fetch_history(sym, start, end, refresh=refresh)
        if h and h.get("status") == "priced" and h.get("prices"):
            msci_prices = h["prices"]
            msci_symbol = sym
            break

    p0 = None
    if msci_prices:
        p0 = previous_price(msci_prices, start)
        if p0 is None:
            sorted_msci_dates = sorted(msci_prices.keys())
            if sorted_msci_dates:
                p0 = msci_prices[sorted_msci_dates[0]]

    # Fetch XEON prices
    xeon_prices = {}
    h_xeon = fetch_history("XEON.DE", start, end, refresh=refresh)
    if h_xeon and h_xeon.get("status") == "priced" and h_xeon.get("prices"):
        xeon_prices = h_xeon["prices"]

    xeon_p0 = None
    if xeon_prices:
        xeon_p0 = previous_price(xeon_prices, start)
        if xeon_p0 is None:
            sorted_xeon_dates = sorted(xeon_prices.keys())
            if sorted_xeon_dates:
                xeon_p0 = xeon_prices[sorted_xeon_dates[0]]

    # Load Eurostat CPI
    cpi_data = fetch_eurostat_cpi()
    
    def get_cpi_value(d: date) -> float | None:
        if not cpi_data:
            return None
        label = f"{d.year:04d}-{d.month:02d}"
        if label in cpi_data:
            return cpi_data[label]
        sorted_labels = sorted(cpi_data.keys())
        best_val = None
        for l in sorted_labels:
            if l <= label:
                best_val = cpi_data[l]
            else:
                break
        if best_val is not None:
            return best_val
        return cpi_data[sorted_labels[-1]] if sorted_labels else None

    start_cpi = get_cpi_value(start)
    live_prices = current_price_overrides(histories, end, refresh=refresh)
    live_fx_prices = current_price_overrides(fx_histories, end, refresh=refresh)
    live_msci_price = {}
    if msci_symbol and msci_prices:
        live_msci_price = current_price_overrides({msci_symbol: {"status": "priced", "prices": msci_prices}}, end, refresh=refresh).get(msci_symbol, {})
    live_xeon_price = {}
    if xeon_prices:
        live_xeon_price = current_price_overrides({"XEON.DE": {"status": "priced", "prices": xeon_prices}}, end, refresh=refresh).get("XEON.DE", {})

    for valuation_date in sorted_dates:
        while trade_index < len(ordered_trades) and ordered_trades[trade_index].date <= valuation_date:
            trade = ordered_trades[trade_index]
            key = trade.isin or trade.asset
            amount, _ = trade_cash_amount(trade)
            holdings[key] = holdings.get(key, ZERO) + trade.quantity_diff
            
            is_tr = (trade.broker.lower() == "trade republic")
            if trade.quantity_diff >= ZERO:
                invested += amount
                if is_tr:
                    invested_tr += amount
                else:
                    invested_other += amount
                invested_by_asset[key] = invested_by_asset.get(key, ZERO) + amount
            else:
                proceeds += amount
                if is_tr:
                    proceeds_tr += amount
                else:
                    proceeds_other += amount
                invested_by_asset[key] = invested_by_asset.get(key, ZERO) - amount
            trade_index += 1

        market_value = ZERO
        priced_positions = 0
        unpriced_positions = 0
        for key, quantity in holdings.items():
            if quantity <= ZERO:
                continue
            ref = refs.get(key, {})
            symbol = ref.get("symbol", "")
            history = histories.get(symbol, {})
            price_point = price_point_for_date(history.get("prices", {}), valuation_date, live_prices.get(symbol))
            price = price_point[0] if price_point else None
            currency = normalize_currency_code(history.get("currency") or "EUR")
            fx_rate = historical_fx_rate(currency, fx_histories, valuation_date, live_fx_prices=live_fx_prices)
            if price is None or fx_rate is None:
                unpriced_positions += 1
                fallback_val = invested_by_asset.get(key, ZERO)
                if fallback_val > ZERO:
                    market_value += fallback_val
                continue
            if currency == "GBp":
                price = price / 100
            market_value += quantity * Decimal(str(price)) * Decimal(str(fx_rate))
            priced_positions += 1

        net_contributions = invested - proceeds
        profit = market_value - net_contributions
        return_pct = (profit / net_contributions * Decimal("100")) if net_contributions > ZERO else ZERO
        
        # Accumulate dividends and cash interest up to valuation_date by broker
        accumulated_dividends_tr = sum((d.amount_eur for d in dividends if d.date <= valuation_date and d.broker.lower() == "trade republic"), ZERO)
        accumulated_dividends_other = sum((d.amount_eur for d in dividends if d.date <= valuation_date and d.broker.lower() != "trade republic"), ZERO)
        
        accumulated_interest_tr = sum((Decimal(str(i["net_eur"])) for i in interests if i["date"] <= valuation_date and i.get("broker", "").lower() == "trade republic"), ZERO)
        accumulated_interest_bbva = sum((Decimal(str(i["net_eur"])) for i in interests if i["date"] <= valuation_date and i.get("broker", "").lower() == "bbva"), ZERO)
        accumulated_interest_other = sum((Decimal(str(i["net_eur"])) for i in interests if i["date"] <= valuation_date and i.get("broker", "").lower() not in {"trade republic", "bbva"}), ZERO)

        # Get historical cash balances
        tr_cash_bal = sum(item[1] for item in tr_cash_history if item[0] <= valuation_date)
        tr_contrib = sum(item[2] for item in tr_cash_history if item[0] <= valuation_date)
        
        bbva_cash_bal = sum(item[1] for item in bbva_cash_history if item[0] <= valuation_date)
        bbva_contrib = sum(item[2] for item in bbva_cash_history if item[0] <= valuation_date)
        revolut_cash_bal = revolut_cash_balance_eur(revolut_cash_history, valuation_date, fx_histories, live_fx_prices)
        revolut_contrib = revolut_contributions_eur(revolut_cash_history, valuation_date, fx_histories)

        broker_lower = broker.lower()
        include_tr_cash = (broker_lower == "all" or broker_lower == "trade republic")
        include_bbva_cash = (broker_lower == "all" or broker_lower == "bbva")
        include_revolut_cash = broker_lower == "all"

        # Reconstruct total market value including cash balance under total return
        total_market_value = market_value + accumulated_dividends_other + accumulated_interest_other
        if include_tr_cash:
            total_market_value += tr_cash_bal
        else:
            total_market_value += accumulated_dividends_tr + accumulated_interest_tr
            
        if include_bbva_cash:
            total_market_value += bbva_cash_bal
        else:
            total_market_value += accumulated_interest_bbva

        if include_revolut_cash:
            total_market_value += revolut_cash_bal

        # Reconstruct total net contributions (Capital at Work) under total return
        total_net_contributions = (invested_other - proceeds_other)
        if include_tr_cash:
            total_net_contributions += tr_contrib
        else:
            total_net_contributions += (invested_tr - proceeds_tr)
            
        if include_bbva_cash:
            total_net_contributions += bbva_contrib

        if include_revolut_cash:
            total_net_contributions += revolut_contrib

        total_profit = total_market_value - total_net_contributions
        total_return_pct = (total_profit / total_net_contributions * Decimal("100")) if total_net_contributions > ZERO else ZERO

        msci_return = None
        msci_price_date = None
        if p0 and p0 > 0:
            price_point = price_point_for_date(msci_prices, valuation_date, live_msci_price)
            if price_point is not None:
                price_t, price_date = price_point
                msci_price_date = price_date.isoformat()
                msci_return = float((Decimal(str(price_t)) - Decimal(str(p0))) / Decimal(str(p0)) * 100)

        xeon_return = None
        xeon_price_date = None
        if xeon_p0 and xeon_p0 > 0:
            price_point = price_point_for_date(xeon_prices, valuation_date, live_xeon_price)
            if price_point is not None:
                price_t, price_date = price_point
                xeon_price_date = price_date.isoformat()
                xeon_return = float((Decimal(str(price_t)) - Decimal(str(xeon_p0))) / Decimal(str(xeon_p0)) * 100)

        # Real Eurozone HICP inflation index
        inflation_return = 0.0
        if start_cpi and start_cpi > 0:
            cpi_t = get_cpi_value(valuation_date)
            if cpi_t:
                inflation_return = float((Decimal(str(cpi_t)) - Decimal(str(start_cpi))) / Decimal(str(start_cpi)) * 100)
        else:
            # Fallback if Eurostat is completely offline
            days_since_start = (valuation_date - start).days
            inflation_return = float(((Decimal("1.02") ** (Decimal(str(days_since_start)) / Decimal("365.25"))) - 1) * 100)

        if market_value > ZERO and (net_contributions > ZERO or market_value > ZERO):
            series.append(
                {
                    "date": valuation_date.isoformat(),
                    "market_value": money(market_value),
                    "total_market_value": money(total_market_value),
                    "net_contributions": money(net_contributions),
                    "total_net_contributions": money(total_net_contributions),
                    "profit": money(profit),
                    "total_profit": money(total_profit),
                    "return_pct": float(return_pct.quantize(Decimal("0.01"))),
                    "total_return_pct": float(total_return_pct.quantize(Decimal("0.01"))),
                    "priced_positions": priced_positions,
                    "unpriced_positions": unpriced_positions,
                    "msci_return_pct": round(msci_return, 2) if msci_return is not None else None,
                    "msci_price_date": msci_price_date,
                    "xeon_return_pct": round(xeon_return, 2) if xeon_return is not None else None,
                    "xeon_price_date": xeon_price_date,
                    "inflation_return_pct": round(inflation_return, 2),
                }
            )

    return {
        "series": series,
        "status": "priced",
        "symbols": sorted({ref.get("symbol", "") for ref in refs.values() if ref.get("symbol")}),
        "variations": calculate_variations(series),
        "_history_context": {
            "start": start,
            "end": end,
            "histories": histories,
            "fx_histories": fx_histories,
            "msci_prices": msci_prices,
            "msci_symbol": msci_symbol,
            "xeon_prices": xeon_prices,
            "live_prices": live_prices,
            "live_fx_prices": live_fx_prices,
            "live_msci_price": live_msci_price,
            "live_xeon_price": live_xeon_price,
        },
    }


def get_historical_price_in_eur(prices: dict[str, float], fx_prices: dict[str, float], target_date: date, invert: bool, currency: str) -> float | None:
    asset_price = previous_price(prices, target_date)
    if asset_price is None:
        return None
    currency = normalize_currency_code(currency)
    if currency == "GBp":
        asset_price = asset_price / 100
    if fx_prices:
        fx_rate = previous_price(fx_prices, target_date)
        if fx_rate is not None:
            rate = 1 / fx_rate if invert else fx_rate
            return float(Decimal(str(asset_price)) * Decimal(str(rate)))
        return None
    return asset_price


def compute_position_variations(
    asset: str,
    symbol: str,
    currency: str,
    quantity: Decimal,
    current_value_eur: Decimal,
    is_crypto_wallet: bool,
    refresh: bool = False
) -> dict[str, dict[str, float]]:
    res = {
        "1d": {"pct": 0.0, "amount": 0.0},
        "1w": {"pct": 0.0, "amount": 0.0},
        "1m": {"pct": 0.0, "amount": 0.0},
    }
    if quantity <= ZERO:
        return res

    today = date.today()
    start_date = today - timedelta(days=45)

    try:
        if is_crypto_wallet:
            h = fetch_crypto_eur_history(asset, start_date, today, refresh=refresh)
            prices = h.get("prices", {})
            fx_prices = {}
            invert = False
            currency = "EUR"
        else:
            if not symbol:
                return res
            h = fetch_history(symbol, start_date, today, refresh=refresh)
            prices = h.get("prices", {})
            currency = normalize_currency_code(currency or h.get("currency") or "EUR")
            if currency != "EUR":
                fx_sym, invert = fx_symbol_for(currency)
                fx_h = fetch_history(fx_sym, start_date, today, refresh=refresh)
                fx_prices = fx_h.get("prices", {})
            else:
                fx_prices = {}
                invert = False

        if not prices:
            return res

        for period, days in [("1d", 1), ("1w", 7), ("1m", 30)]:
            target_date = today - timedelta(days=days)
            past_price = get_historical_price_in_eur(prices, fx_prices, target_date, invert, currency)
            if past_price is not None and past_price > 0:
                past_val = Decimal(str(past_price)) * quantity
                amount = current_value_eur - past_val
                pct_val = (amount / past_val * Decimal("100")) if past_val > 0 else ZERO
                res[period] = {
                    "pct": float(pct_val.quantize(Decimal("0.01"))),
                    "amount": float(amount.quantize(Decimal("0.01"))),
                }
    except Exception:
        pass
    return res


def enrich_positions(positions: list[dict[str, Any]], mappings: dict[str, dict[str, str]], refresh: bool = False) -> dict[str, Any]:
    exposures = read_exposures()
    enriched = []
    market_value = ZERO
    unrealized_pl = ZERO
    priced_assets = 0
    unpriced_assets = 0

    for position in positions:
        asset = position["asset"]
        quantity = Decimal(str(position["quantity"]))
        cost = Decimal(str(position["cost_basis_eur"]))
        position_isin = position.get("isin", "")
        mapping = mapping_for(asset, position_isin, mappings)
        isin = position_isin or mapping.get("isin", "")
        
        mapping_ticker = mapping.get("ticker", "")
        mapping_exchange = mapping.get("exchange", "")
        position_symbol = position.get("symbol", "")
        
        direct_symbol = yahoo_symbol_from_mapping(mapping_ticker, mapping_exchange)
        if not direct_symbol and position_symbol and not isin:
            direct_symbol = position_symbol

        row = {**position, "isin": isin, "symbol": "", "price": None, "price_currency": "", "market_value_eur": None}
        
        # Resolve symbol upfront to support logo mappings for both open and closed positions
        search_query = isin or position_symbol
        symbol_data = {}
        if (search_query or direct_symbol) and str(position.get("broker") or "").lower() != "crypto wallet":
            symbol_data = resolve_isin(search_query, refresh=refresh, direct_symbol=direct_symbol)
            row["symbol"] = symbol_data.get("symbol", "")

        snapshot_val = position.get("market_value_eur")
        if (
            str(position.get("broker") or "").lower() == "crypto wallet"
            and snapshot_val is not None
            and Decimal(str(snapshot_val)) > ZERO
        ):
            value = Decimal(str(snapshot_val))
            pricing_status = "crypto_wallet"
            price_currency = "EUR"
            fetched_at = None

            # Try to fetch live price dynamically
            if position_symbol:
                try:
                    price_data = fetch_price(position_symbol, refresh=refresh)
                    if price_data.get("status") == "priced":
                        fx = fx_to_eur(price_data.get("currency", "EUR"), refresh=refresh)
                        if fx.get("rate") is not None:
                            price_val = Decimal(str(price_data["price"]))
                            rate_val = Decimal(str(fx["rate"]))
                            value = quantity * price_val * rate_val
                            pricing_status = "priced"
                            price_currency = price_data.get("currency", "EUR")
                            fetched_at = price_data.get("fetched_at")
                except Exception:
                    pass

            price = value / quantity if quantity > ZERO else ZERO
            row.update(
                {
                    "symbol": position_symbol,
                    "price": float(price) if price > ZERO else None,
                    "price_currency": price_currency,
                    "market_value_eur": money(value),
                    "unrealized_pl_eur": money(value - cost),
                    "unrealized_pl_pct": float(((value - cost) / cost * Decimal("100")).quantize(Decimal("0.01"))) if cost > ZERO else None,
                    "display_pl_eur": money(value - cost),
                    "display_pl_pct": float(((value - cost) / cost * Decimal("100")).quantize(Decimal("0.01"))) if cost > ZERO else None,
                    "pricing_status": pricing_status,
                }
            )
            if fetched_at:
                row["fetched_at"] = fetched_at

            # Compute variations
            pos_vars = compute_position_variations(
                asset=asset,
                symbol=position_symbol,
                currency=price_currency,
                quantity=quantity,
                current_value_eur=value,
                is_crypto_wallet=True,
                refresh=refresh
            )
            row["variations"] = pos_vars

            market_value += value
            unrealized_pl += value - cost
            priced_assets += 1
            row["asset_type"] = determine_asset_type(asset, isin, row["symbol"], row.get("broker", ""), exposures)
            enriched.append(row)
            continue

        def apply_fallback_if_possible(err_status: str):
            nonlocal market_value, unrealized_pl, unpriced_assets
            snapshot_val = position.get("market_value_eur")
            if snapshot_val is not None and snapshot_val > 0:
                row["pricing_status"] = "snapshot"
                row["market_value_eur"] = snapshot_val
                market_value += Decimal(str(snapshot_val))
                snapshot_pl = position.get("display_pl_eur")
                if snapshot_pl is not None:
                    unrealized_pl += Decimal(str(snapshot_pl))
                unpriced_assets += 1

                # Compute variations for fallback snapshot
                is_crypto = str(position.get("broker") or "").lower() == "crypto wallet"
                pos_vars = compute_position_variations(
                    asset=asset,
                    symbol=position_symbol or direct_symbol,
                    currency=position.get("price_currency", "EUR") or "EUR",
                    quantity=quantity,
                    current_value_eur=Decimal(str(snapshot_val)),
                    is_crypto_wallet=is_crypto,
                    refresh=refresh
                )
                row["variations"] = pos_vars
            else:
                row["pricing_status"] = err_status
                if err_status == "closed":
                    row["display_pl_eur"] = position.get("realized_pl_eur")
                    row["display_pl_pct"] = position.get("realized_pl_pct")
                else:
                    unpriced_assets += 1
            row["asset_type"] = determine_asset_type(asset, isin, row.get("symbol", ""), row.get("broker", ""), exposures)
            enriched.append(row)

        if quantity <= ZERO:
            apply_fallback_if_possible("closed")
            continue
            
        if not search_query and not direct_symbol:
            apply_fallback_if_possible("missing_isin")
            continue

        if symbol_data.get("status") != "resolved":
            row["pricing_error"] = symbol_data.get("error", "")
            apply_fallback_if_possible(symbol_data.get("status", "unresolved"))
            continue

        price_data = fetch_price(row["symbol"], refresh=refresh)
        if price_data.get("status") != "priced":
            row["pricing_error"] = price_data.get("error", "")
            apply_fallback_if_possible(price_data.get("status", "price_error"))
            continue

        fx = fx_to_eur(price_data.get("currency", "EUR"), refresh=refresh)
        if fx.get("rate") is None:
            row["pricing_error"] = f"Could not convert {price_data.get('currency')} to EUR."
            apply_fallback_if_possible(fx.get("status", "fx_error"))
            continue

        price = Decimal(str(price_data["price"]))
        rate = Decimal(str(fx["rate"]))
        value = quantity * price * rate
        row.update(
            {
                "price": float(price),
                "price_currency": price_data.get("currency", "EUR"),
                "market_value_eur": money(value),
                "unrealized_pl_eur": money(value - cost),
                "unrealized_pl_pct": float(((value - cost) / cost * Decimal("100")).quantize(Decimal("0.01"))) if cost > ZERO else None,
                "display_pl_eur": money(value - cost),
                "display_pl_pct": float(((value - cost) / cost * Decimal("100")).quantize(Decimal("0.01"))) if cost > ZERO else None,
                "pricing_status": "priced",
                "fetched_at": price_data.get("fetched_at"),
            }
        )

        # Compute variations
        pos_vars = compute_position_variations(
            asset=asset,
            symbol=row["symbol"],
            currency=price_data.get("currency", "EUR"),
            quantity=quantity,
            current_value_eur=value,
            is_crypto_wallet=False,
            refresh=refresh
        )
        row["variations"] = pos_vars

        market_value += value
        unrealized_pl += value - cost
        priced_assets += 1
        row["asset_type"] = determine_asset_type(asset, isin, row["symbol"], row.get("broker", ""), exposures)
        enriched.append(row)

    return {
        "positions": sorted(enriched, key=lambda item: (not item["is_open"], item["asset"])),
        "market_value": money(market_value),
        "unrealized_pl": money(unrealized_pl),
        "priced_assets": priced_assets,
        "unpriced_assets": unpriced_assets,
    }


def calculate_portfolio_statistics(trades, mappings, person=PRIMARY_PORTFOLIO_ID, broker="all", history_context=None):
    import numpy as np

    if not trades:
        return None
    start_date = min(trade.date for trade in trades)
    end_date = date.today()
    if start_date >= end_date:
        return None
        
    dates_list = []
    curr = start_date
    while curr <= end_date:
        if curr.weekday() < 5:
            dates_list.append(curr)
        curr += timedelta(days=1)
        
    if not dates_list:
        return None
        
    tr_cash_events, bbva_cash_events, revolut_cash_events = load_cash_histories(person)
                
    dividends = read_portfolio_dividends(person)
    interests = read_cash_interests(person)
    
    refs = build_instrument_refs(trades, mappings, refresh=False)
    history_context = history_context or {}
    context_histories = history_context.get("histories", {}) if isinstance(history_context, dict) else {}
    histories = {}
    currencies = set()
    for ref in refs.values():
        symbol = ref.get("symbol", "")
        if not symbol: continue
        history = context_histories.get(symbol) or fetch_history(symbol, start_date - timedelta(days=30), end_date, refresh=False)
        histories[symbol] = history
        if history.get("status") == "priced":
            currencies.add(normalize_currency_code(history.get("currency") or "EUR"))
            
    fx_histories = dict(history_context.get("fx_histories", {}) if isinstance(history_context, dict) else {})
    for currency in currencies:
        normalized = fx_base_currency(currency)
        if normalized and normalized != "EUR":
            symbol, _ = fx_symbol_for(normalized)
            if symbol not in fx_histories:
                fx_histories[symbol] = fetch_history(symbol, start_date - timedelta(days=30), end_date, refresh=False)
    add_revolut_fx_histories(revolut_cash_events, start_date, end_date, fx_histories, refresh=False)
            
    live_prices = dict(history_context.get("live_prices", {}) if isinstance(history_context, dict) else {})
    live_fx_prices = dict(history_context.get("live_fx_prices", {}) if isinstance(history_context, dict) else {})
    live_msci_price = dict(history_context.get("live_msci_price", {}) if isinstance(history_context, dict) else {})
    live_xeon_price = dict(history_context.get("live_xeon_price", {}) if isinstance(history_context, dict) else {})

    msci_prices = history_context.get("msci_prices", {}) if isinstance(history_context, dict) else {}
    msci_symbol = history_context.get("msci_symbol", "") if isinstance(history_context, dict) else ""
    if not msci_prices:
        for sym in ["SWDA.MI", "EUNL.DE", "URTH"]:
            h = fetch_history(sym, start_date - timedelta(days=30), end_date, refresh=False)
            if h and h.get("status") == "priced" and h.get("prices"):
                msci_prices = h["prices"]
                msci_symbol = sym
                break

    xeon_prices = history_context.get("xeon_prices", {}) if isinstance(history_context, dict) else {}
    if not xeon_prices:
        h_xeon = fetch_history("XEON.DE", start_date - timedelta(days=30), end_date, refresh=False)
        if h_xeon and h_xeon.get("status") == "priced" and h_xeon.get("prices"):
            xeon_prices = h_xeon["prices"]
            
    broker_lower = broker.lower()
    include_tr_cash = (broker_lower == "all" or broker_lower == "trade republic")
    include_bbva_cash = (broker_lower == "all" or broker_lower == "bbva")
    include_revolut_cash = broker_lower == "all"
    
    filtered_trades = trades
    if broker_lower != "all":
        filtered_trades = [t for t in trades if t.broker.lower() == broker_lower]
        
    ordered_trades = sorted(filtered_trades, key=lambda item: item.date)
    
    daily_metrics = []
    holdings = {}
    cost_basis = {}
    invested_other = ZERO
    proceeds_other = ZERO
    invested_tr = ZERO
    proceeds_tr = ZERO
    trade_idx = 0
    
    for d in dates_list:
        while trade_idx < len(ordered_trades) and ordered_trades[trade_idx].date <= d:
            t = ordered_trades[trade_idx]
            asset = t.isin or t.asset
            amount, _ = trade_cash_amount(t)
            is_tr = (t.broker.lower() == "trade republic")
            
            if t.quantity_diff >= ZERO:
                holdings[asset] = holdings.get(asset, ZERO) + t.quantity_diff
                cost_basis[asset] = cost_basis.get(asset, ZERO) + amount
                if is_tr:
                    invested_tr += amount
                else:
                    invested_other += amount
            else:
                sell_qty = abs(t.quantity_diff)
                prev_qty = holdings.get(asset, ZERO)
                prev_cost = cost_basis.get(asset, ZERO)
                if prev_qty > ZERO:
                    removed_cost = min(prev_cost, prev_cost * sell_qty / prev_qty)
                else:
                    removed_cost = ZERO
                cost_basis[asset] = max(ZERO, prev_cost - removed_cost)
                holdings[asset] = max(ZERO, prev_qty - sell_qty)
                if is_tr:
                    proceeds_tr += amount
                else:
                    proceeds_other += amount
            trade_idx += 1
            
        sec_mv = ZERO
        for asset, qty in holdings.items():
            if qty <= ZERO: continue
            ref = refs.get(asset)
            if not ref:
                ref_found = next((r_val for r_asset, r_val in refs.items() if r_val.get("isin") == asset or r_asset == asset), None)
                if ref_found: ref = ref_found
                else: continue
            symbol = ref.get("symbol", "")
            if not symbol: continue
            h = histories.get(symbol)
            if not h or h.get("status") != "priced": continue
            
            price_point = price_point_for_date(h.get("prices", {}), d, live_prices.get(symbol))
            if price_point is None: continue
            price = price_point[0]
            
            val_in_curr = Decimal(str(price)) * qty
            currency = normalize_currency_code(h.get("currency") or "EUR")
            
            if currency != "EUR":
                normalized = fx_base_currency(currency)
                fx_sym, is_div = fx_symbol_for(normalized)
                fx_h = fx_histories.get(fx_sym)
                if fx_h and fx_h.get("prices"):
                    fx_point = price_point_for_date(fx_h["prices"], d, live_fx_prices.get(fx_sym))
                    fx_rate = fx_point[0] if fx_point else None
                    if fx_rate is not None and fx_rate > 0:
                        rate_dec = Decimal(str(fx_rate))
                        if currency == "GBp": val_in_curr = val_in_curr / Decimal('100')
                        multiplier = Decimal('1') / rate_dec if is_div else rate_dec
                        val_in_curr = val_in_curr * multiplier
            sec_mv += val_in_curr
            
        tr_cash_bal = sum(item[1] for item in tr_cash_events if item[0] <= d) if include_tr_cash else ZERO
        tr_contrib = sum(item[2] for item in tr_cash_events if item[0] <= d) if include_tr_cash else ZERO
        bbva_cash_bal = sum(item[1] for item in bbva_cash_events if item[0] <= d) if include_bbva_cash else ZERO
        bbva_contrib = sum(item[2] for item in bbva_cash_events if item[0] <= d) if include_bbva_cash else ZERO
        revolut_cash_bal = revolut_cash_balance_eur(revolut_cash_events, d, fx_histories, live_fx_prices) if include_revolut_cash else ZERO
        revolut_contrib = revolut_contributions_eur(revolut_cash_events, d, fx_histories) if include_revolut_cash else ZERO
        
        accumulated_dividends_other = sum((div.amount_eur for div in dividends if div.date <= d and div.broker.lower() != "trade republic"), ZERO)
        accumulated_interest_other = sum((Decimal(str(i["net_eur"])) for i in interests if i["date"] <= d and i.get("broker", "").lower() not in {"trade republic", "bbva"}), ZERO)
        
        # Total Return metrics
        total_mv = sec_mv + accumulated_dividends_other + accumulated_interest_other + tr_cash_bal + bbva_cash_bal + revolut_cash_bal
        total_nc = (invested_other - proceeds_other) + tr_contrib + bbva_contrib + revolut_contrib
        
        # Price Return metrics
        price_mv = sec_mv
        price_nc = (invested_other - proceeds_other) + (invested_tr - proceeds_tr)
        
        msci_point = price_point_for_date(msci_prices, d, live_msci_price if msci_symbol else None)
        msci_price = msci_point[0] if msci_point else None
        msci_price_date = msci_point[1] if msci_point else None

        xeon_point = price_point_for_date(xeon_prices, d, live_xeon_price)
        xeon_price = xeon_point[0] if xeon_point else None
        xeon_price_date = xeon_point[1] if xeon_point else None
            
        daily_metrics.append({
            'date': d,
            'total_mv': float(total_mv),
            'total_nc': float(total_nc),
            'price_mv': float(price_mv),
            'price_nc': float(price_nc),
            'msci': float(msci_price) if msci_price else None,
            'msci_date': msci_price_date,
            'xeon': float(xeon_price) if xeon_price else None,
            'xeon_date': xeon_price_date,
        })
        
    # Compute returns for both
    total_returns = []
    price_returns = []
    msci_returns = []
    xeon_returns = []
    daily_returns = []
    
    for i in range(1, len(daily_metrics)):
        prev = daily_metrics[i-1]
        curr = daily_metrics[i]
        
        # Total Return
        total_cf = curr['total_nc'] - prev['total_nc']
        total_denom = prev['total_mv'] + max(total_cf, 0)
        total_ret = None
        if total_denom > 0:
            total_ret = (curr['total_mv'] - prev['total_mv'] - total_cf) / total_denom
            total_returns.append(total_ret)
            
        # Price Return
        price_cf = curr['price_nc'] - prev['price_nc']
        price_denom = prev['price_mv'] + max(price_cf, 0)
        price_ret = None
        if price_denom > 0:
            price_ret = (curr['price_mv'] - prev['price_mv'] - price_cf) / price_denom
            price_returns.append(price_ret)
            
        # MSCI World
        p_prev = prev['msci']
        p_curr = curr['msci']
        msci_ret = None
        if p_prev and p_curr and p_prev > 0 and curr.get("msci_date") != prev.get("msci_date"):
            msci_ret = (p_curr - p_prev) / p_prev
            msci_returns.append(msci_ret)

        # XEON
        x_prev = prev['xeon']
        x_curr = curr['xeon']
        xeon_ret = None
        if x_prev and x_curr and x_prev > 0 and curr.get("xeon_date") != prev.get("xeon_date"):
            xeon_ret = (x_curr - x_prev) / x_prev
            xeon_returns.append(xeon_ret)

        daily_returns.append({
            "date": curr["date"].isoformat(),
            "total_return": float(total_ret) if total_ret is not None else None,
            "price_return": float(price_ret) if price_ret is not None else None,
            "msci_return": float(msci_ret) if msci_ret is not None else None,
            "xeon_return": float(xeon_ret) if xeon_ret is not None else None,
        })
            
    if not total_returns or not price_returns:
        return None
        
    tot_rets = np.array(total_returns)
    prc_rets = np.array(price_returns)
    msci_rets_only = np.array(msci_returns)
    xeon_rets_only = np.array(xeon_returns)
    
    rf_annual = SETTINGS.annual_risk_free_rate
    rf_daily = rf_annual / 252
    
    def calculate_stats(rets, m_rets, x_rets):
        mean = np.mean(rets)
        std = np.std(rets)
        sharpe = ((mean - rf_daily) / std * np.sqrt(252)) if std > 0 else 0.0
        
        m_mean = np.mean(m_rets) if len(m_rets) > 0 else 0.0
        m_std = np.std(m_rets) if len(m_rets) > 0 else 0.0
        m_sharpe = ((m_mean - rf_daily) / m_std * np.sqrt(252)) if m_std > 0 else 0.0

        x_mean = np.mean(x_rets) if len(x_rets) > 0 else 0.0
        x_std = np.std(x_rets) if len(x_rets) > 0 else 0.0
        x_sharpe = ((x_mean - rf_daily) / x_std * np.sqrt(252)) if x_std > 0 else 0.0
        
        return {
            "portfolio": {
                "daily_variance_pct": float(np.var(rets) * 10000),
                "daily_volatility_pct": float(std * 100),
                "annualized_volatility_pct": float(std * np.sqrt(252) * 100),
                "sharpe_ratio": float(sharpe),
                "mean_daily_return_pct": float(mean * 100),
            },
            "msci": {
                "daily_variance_pct": float(np.var(m_rets) * 10000) if len(m_rets) > 0 else 0.0,
                "daily_volatility_pct": float(m_std * 100) if len(m_rets) > 0 else 0.0,
                "annualized_volatility_pct": float(m_std * np.sqrt(252) * 100) if len(m_rets) > 0 else 0.0,
                "sharpe_ratio": float(m_sharpe),
                "mean_daily_return_pct": float(m_mean * 100) if len(m_rets) > 0 else 0.0,
            },
            "xeon": {
                "daily_variance_pct": float(np.var(x_rets) * 10000) if len(x_rets) > 0 else 0.0,
                "daily_volatility_pct": float(x_std * 100) if len(x_rets) > 0 else 0.0,
                "annualized_volatility_pct": float(x_std * np.sqrt(252) * 100) if len(x_rets) > 0 else 0.0,
                "sharpe_ratio": float(x_sharpe),
                "mean_daily_return_pct": float(x_mean * 100) if len(x_rets) > 0 else 0.0,
            }
        }
        
    return {
        "total_return": calculate_stats(tot_rets, msci_rets_only, xeon_rets_only),
        "price_return": calculate_stats(prc_rets, msci_rets_only, xeon_rets_only),
        "days_evaluated": len(tot_rets),
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "daily_returns": daily_returns,
    }


def dashboard_payload(
    refresh: bool = False,
    person: str = PRIMARY_PORTFOLIO_ID,
    berkshire_mode: str = "stock",
    proxy_mode: str = "off",
    broker: str = "all",
    live_only: str = "off",
) -> dict[str, Any]:
    person = configured_portfolio_id(person)
    berkshire_mode = normalize_berkshire_mode(berkshire_mode)
    proxy_mode = normalize_proxy_mode(proxy_mode)
    broker = (broker or "all").strip().lower()
    ensure_legacy_statements_imported(person)
    if person != PRIMARY_PORTFOLIO_ID and not get_movement_store().movements(person, ("trade",)):
        return family_dashboard_payload(person, refresh=refresh, berkshire_mode=berkshire_mode, proxy_mode=proxy_mode, broker=broker, live_only=live_only)

    trades, source = read_ledger_trades(person)
    dividends = read_ledger_dividends(person)
    wallet_positions = read_crypto_wallet_positions(person)

    # Get all brokers first (unfiltered)
    all_brokers = sorted(list(set(t.broker for t in trades if t.broker) | set(d.broker for d in dividends if d.broker)))
    if wallet_positions:
        all_brokers = sorted(set(all_brokers) | {str(position.get("broker") or "Crypto Wallet") for position in wallet_positions})

    # Filter by broker
    if broker != "all":
        trades = [t for t in trades if t.broker.lower() == broker]
        dividends = [d for d in dividends if d.broker.lower() == broker]
        wallet_positions = [p for p in wallet_positions if str(p.get("broker") or "").lower() == broker]

    mappings = read_mappings()

    # Filter by live_only
    if live_only == "on":
        live_trades = []
        for t in trades:
            mapping = mapping_for(t.asset, t.isin, mappings)
            direct_symbol = yahoo_symbol_from_mapping(mapping.get("ticker", ""), mapping.get("exchange", ""))
            if not direct_symbol and t.isin:
                direct_symbol = resolve_isin(t.isin).get("symbol", "")
            is_crypto = t.broker.lower() == "crypto wallet"
            if direct_symbol or is_crypto:
                live_trades.append(t)
        trades = live_trades
        dividends = [d for d in dividends if any(t.asset == d.asset for t in trades)]
        if wallet_positions:
            wallet_positions = [p for p in wallet_positions if p.get("pricing_status") != "snapshot"]
    summary = summarize_trades(trades)
    if wallet_positions:
        summary["positions"].extend(wallet_positions)
        if broker == "crypto wallet" and not summary["series"]:
            wallet_net = sum((Decimal(str(position.get("cost_basis_eur") or 0)) for position in wallet_positions), ZERO)
            if wallet_net <= ZERO:
                wallet_net = sum((Decimal(str(position.get("market_value_eur") or 0)) for position in wallet_positions), ZERO)
            summary["series"] = [
                {
                    "date": crypto_wallet_transaction_start_date().isoformat(),
                    "invested": money(wallet_net),
                    "proceeds": 0.0,
                    "net_contributions": money(wallet_net),
                    "open_cost_basis": money(wallet_net),
                    "realized_pl": 0.0,
                    "fees": 0.0,
                    "taxes": 0.0,
                },
                {
                    "date": date.today().isoformat(),
                    "invested": money(wallet_net),
                    "proceeds": 0.0,
                    "net_contributions": money(wallet_net),
                    "open_cost_basis": money(wallet_net),
                    "realized_pl": 0.0,
                    "fees": 0.0,
                    "taxes": 0.0,
                }
            ]
            summary["totals"]["invested"] = money(wallet_net)
            summary["totals"]["net_contributions"] = money(wallet_net)
            summary["totals"]["open_cost_basis"] = money(wallet_net)
    dividend_summary = summarize_dividends(dividends)
    contribution_summary = (
        crypto_wallet_contributions(wallet_positions)
        if broker == "crypto wallet" and wallet_positions
        else summarize_net_contributions(trades)
    )
    priced = enrich_positions(summary["positions"], mappings, refresh=refresh)
    distribution = calculate_distribution(
        priced["positions"],
        read_exposures(berkshire_mode=berkshire_mode, proxy_mode=proxy_mode),
        berkshire_mode=berkshire_mode,
        proxy_mode=proxy_mode,
    )
    valuation = (
        crypto_wallet_valuation(wallet_positions)
        if broker == "crypto wallet" and wallet_positions
        else calculate_valuation_series(trades, mappings, refresh=refresh, person=person, broker=broker)
    )

    extra_frictions = read_ledger_frictions(person)
    if broker != "all":
        extra_frictions = [f for f in extra_frictions if f.broker.lower() == broker]

    frictions = summarize_frictions(
        trade_friction_events(trades) + inferred_fineco_sell_tax_events(trades) + extra_frictions,
        priced["market_value"],
    )
    expenses = summarize_expense_events(read_ledger_expenses(person))
    mapped_assets = set(mappings)
    trade_assets = {trade.asset for trade in trades} | {str(position.get("asset") or "") for position in wallet_positions}
    assets_without_isin = {
        position["asset"]
        for position in summary["positions"]
        if position["is_open"] and not position.get("isin") and not mapping_for(position["asset"], "", mappings).get("isin")
    }

    # Load dividends and cash interests to compute final total returns
    dividends = read_ledger_dividends(person)
    interests = read_ledger_interests(person)
    if broker != "all":
        dividends = [d for d in dividends if d.broker.lower() == broker]
        interests = [i for i in interests if i.get("broker", "").lower() == broker]
    accumulated_dividends = sum((d.amount_eur for d in dividends), ZERO)
    accumulated_interest = sum((Decimal(str(i["net_eur"])) for i in interests), ZERO)
    total_market_value = Decimal(str(priced["market_value"])) + accumulated_dividends + accumulated_interest

    totals = {
        **summary["totals"],
        "market_value": priced["market_value"],
        "total_market_value": float(total_market_value.quantize(Decimal("0.01"))),
        "unrealized_pl": priced["unrealized_pl"],
        "estimated_total_value": money(Decimal(str(priced["market_value"]))),
        "priced_assets": priced["priced_assets"],
        "unpriced_assets": priced["unpriced_assets"],
    }
    if valuation["series"]:
        totals["return_pct"] = valuation["series"][-1]["return_pct"]
        totals["total_return_pct"] = valuation["series"][-1]["total_return_pct"]
        totals["historical_profit"] = valuation["series"][-1]["profit"]
        totals["historical_total_profit"] = valuation["series"][-1]["total_profit"]
        totals["total_market_value"] = valuation["series"][-1]["total_market_value"]
        totals["total_net_contributions"] = valuation["series"][-1]["total_net_contributions"]
    totals["variations"] = valuation.get("variations", {})

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "person": person,
        "person_name": PRIMARY_PORTFOLIO_NAME if person == PRIMARY_PORTFOLIO_ID else SETTINGS.portfolios[person].display_name,
        "berkshire_mode": berkshire_mode,
        "proxy_mode": proxy_mode,
        "trade_file": source["relative_path"],
        "trade_source": source["kind"],
        "mapping_file": configured_path_label(MAPPINGS_CSV),
        "trade_count": len(trades),
        "asset_count": len(trade_assets),
        "date_range": {
            "start": min(trade.date for trade in trades).isoformat() if trades else None,
            "end": max(trade.date for trade in trades).isoformat() if trades else None,
        },
        "totals": totals,
        "series": summary["series"],
        "valuation_series": valuation["series"],
        "valuation_status": {
            "status": valuation["status"],
            "symbols": valuation.get("symbols", []),
        },
        "positions": priced["positions"],
        "distribution": distribution,
        "dividends": dividend_summary,
        "cash_interests": summarize_cash_interests(read_ledger_interests(person)),
        "expenses": expenses,
        "net_contributions": contribution_summary,
        "frictions": frictions,
        "mapping_status": {
            "missing_in_mapping": sorted(assets_without_isin),
            "extra_in_mapping": sorted(mapped_assets - trade_assets),
            "filled_isins": sum(1 for value in mappings.values() if value.get("isin")),
            "total_rows": len(mappings),
        },
        "brokers": all_brokers,
        "stats": calculate_portfolio_statistics(
            trades,
            mappings,
            person=person,
            broker=broker,
            history_context=valuation.get("_history_context", {}),
        ),
    }
    payload["news_symbols"] = news_symbols_from_payload(payload)
    return payload


NEWS_SYMBOL_EXCLUDE = {"", "USD", "EUR", "GBP", "USDC-USD", "BTC-USD", "ETH-USD", "TON-USD", "TON11419-USD"}


def normalize_news_symbol(raw: str) -> str:
    symbol = str(raw or "").strip().upper()
    if not symbol:
        return ""
    symbol = symbol.replace("/", "-")
    if symbol in NEWS_SYMBOL_EXCLUDE or " " in symbol:
        return ""
    if not re.fullmatch(r"[A-Z0-9.-]{1,15}", symbol):
        return ""
    return symbol


def news_symbols_from_payload(payload: dict[str, Any], limit: int = 14) -> list[str]:
    scores: dict[str, Decimal] = {}

    for position in payload.get("positions", []):
        if not position.get("is_open"):
            continue
        if str(position.get("asset_class") or "").lower() == "crypto":
            continue
        isin = str(position.get("isin") or "").upper()
        asset = str(position.get("asset") or "")
        if isin.startswith(("IE", "LU", "NL")) or " acc" in asset.lower() or "etf" in asset.lower():
            continue
        symbol = normalize_news_symbol(position.get("symbol") or "")
        if symbol:
            scores[symbol] = scores.get(symbol, ZERO) + Decimal(str(position.get("market_value_eur") or 0)) + Decimal("1000000")

    for row in payload.get("distribution", {}).get("rows", []):
        symbol = normalize_news_symbol(row.get("holding_ticker") or "")
        if not symbol or symbol.isdigit():
            continue
        scores[symbol] = scores.get(symbol, ZERO) + Decimal(str(row.get("market_value_eur") or 0))

    return [symbol for symbol, _ in sorted(scores.items(), key=lambda item: item[1], reverse=True)[:limit]]


def normalize_news_symbols(symbols: list[str] | tuple[str, ...] | None, limit: int = 14) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for raw in symbols or []:
        symbol = normalize_news_symbol(raw)
        if not symbol or symbol in seen:
            continue
        seen.add(symbol)
        normalized.append(symbol)
        if len(normalized) >= limit:
            break
    return normalized


def parse_rss_date(raw: str) -> str:
    if not raw:
        return ""
    try:
        parsed = email.utils.parsedate_to_datetime(raw)
    except (TypeError, ValueError):
        return raw
    if parsed.tzinfo:
        parsed = parsed.astimezone(timezone.utc)
    return parsed.strftime("%Y-%m-%d %H:%M UTC")


def fetch_symbol_news(symbol: str, limit: int = 6) -> list[dict[str, Any]]:
    query = urllib.parse.urlencode({"s": symbol, "region": "US", "lang": "en-US"})
    url = f"https://feeds.finance.yahoo.com/rss/2.0/headline?{query}"
    req = urllib.request.Request(url, headers={"User-Agent": "portfolio-dashboard/1.0"})
    with urllib.request.urlopen(req, timeout=12) as response:
        raw = response.read()

    root = ET.fromstring(raw)
    items: list[dict[str, Any]] = []
    for item in root.findall(".//item")[:limit]:
        title = " ".join((item.findtext("title", default="") or "").split())
        link = (item.findtext("link", default="") or "").strip()
        if not title or not link:
            continue
        items.append(
            {
                "symbol": symbol,
                "title": title,
                "link": link,
                "published": parse_rss_date(item.findtext("pubDate", default="") or ""),
                "source": item.findtext("source", default="") or "Yahoo Finance",
            }
        )
    return items


def portfolio_news_payload(
    person: str = PRIMARY_PORTFOLIO_ID,
    berkshire_mode: str = "stock",
    proxy_mode: str = "off",
    broker: str = "all",
    refresh: bool = False,
    live_only: str = "off",
    symbols: list[str] | None = None,
) -> dict[str, Any]:
    symbols = normalize_news_symbols(symbols)
    if not symbols:
        payload = dashboard_payload(refresh=False, person=person, berkshire_mode=berkshire_mode, proxy_mode=proxy_mode, broker=broker, live_only=live_only)
        symbols = normalize_news_symbols(payload.get("news_symbols") or news_symbols_from_payload(payload))
    cache = get_news_cache()
    cache_key = "|".join([person, berkshire_mode, proxy_mode, broker, live_only, ",".join(symbols)])
    cached = cache.get(cache_key) if isinstance(cache, dict) else None
    now = int(time.time())
    if not refresh and cached and now - int(cached.get("fetched_at_ts", 0)) < NEWS_TTL_SECONDS:
        return cached.get("payload", {})

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    seen: set[str] = set()
    for symbol in symbols:
        try:
            for item in fetch_symbol_news(symbol):
                dedupe_key = item["link"] or f"{item['symbol']}:{item['title']}"
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                rows.append(item)
        except Exception as exc:
            errors.append({"symbol": symbol, "error": str(exc)})

    rows.sort(key=lambda item: item.get("published") or "", reverse=True)
    result = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": "Yahoo Finance RSS",
        "symbols": symbols,
        "count": len(rows),
        "errors": errors,
        "items": rows[:40],
        "status": "available" if rows else "unavailable",
    }
    if isinstance(cache, dict):
        cache[cache_key] = {"fetched_at_ts": now, "payload": result}
        save_json(NEWS_CACHE, cache)
    return result


EXPORT_PERIOD_LABELS = {
    "1w": "Last 1 week",
    "1m": "Last 1 month",
    "ytd": "Year to date",
    "1y": "Last 1 year",
    "since24": "Since Jan 2024",
    "all": "All time",
}


def export_period_label(period: str) -> str:
    return EXPORT_PERIOD_LABELS.get((period or "all").lower(), "All time")


def export_value(value: Any) -> Any:
    if isinstance(value, (str, int, float)) or value is None:
        return value
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=True)
    return str(value)


def export_sheet_name(name: str) -> str:
    clean = re.sub(r"[\[\]\*\?/\\:]", " ", name).strip()[:31]
    return clean or "Sheet"


def export_rows_from_dicts(rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> list[list[Any]]:
    output = [[label for _, label in columns]]
    for row in rows:
        output.append([export_value(row.get(key)) for key, _ in columns])
    return output


def sorted_by_market_value(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(rows or [], key=lambda row: float(row.get("market_value_eur") or 0), reverse=True)


def dividend_yearly_export_rows(dividends: dict[str, Any]) -> list[dict[str, Any]]:
    yearly: dict[str, dict[str, Any]] = {}
    for row in dividends.get("rows", []) or []:
        year = str(row.get("date") or "")[:4]
        if not year:
            continue
        bucket = yearly.setdefault(year, {"year": year, "amount_eur": 0.0, "tax_eur": 0.0, "gross_eur": 0.0, "count": 0})
        bucket["amount_eur"] += float(row.get("amount_eur") or 0)
        bucket["tax_eur"] += float(row.get("tax_eur") or 0)
        bucket["gross_eur"] += float(row.get("gross_eur") or 0)
        bucket["count"] += 1
    return [
        {
            "year": year,
            "count": values["count"],
            "amount_eur": round(values["amount_eur"], 2),
            "tax_eur": round(values["tax_eur"], 2),
            "gross_eur": round(values["gross_eur"], 2),
        }
        for year, values in sorted(yearly.items(), reverse=True)
    ]


def add_xlsx_sheet(workbook: Workbook, title: str, rows: list[list[Any]]) -> None:
    ws = workbook.create_sheet(export_sheet_name(title))
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        width = min(42, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        ws.column_dimensions[column_cells[0].column_letter].width = width
    ws.freeze_panes = "A2"


def dashboard_export_context(
    payload: dict[str, Any],
    person: str,
    berkshire_mode: str,
    proxy_mode: str,
    broker: str,
    live_only: str,
    period: str,
) -> list[list[Any]]:
    totals = payload.get("totals", {})
    date_range = payload.get("date_range", {})
    toggles = [
        f"Person: {payload.get('person_name') or person}",
        f"Window: {export_period_label(period)}",
        f"Broker: {broker}",
        f"Berkshire: {'13F look-through' if berkshire_mode == 'lookthrough' else 'stock'}",
        f"Composition: {'proxy gaps on' if proxy_mode == 'on' else 'official only'}",
        f"Assets: {'live only' if live_only == 'on' else 'all assets'}",
        "Vs MSCI World: shown",
        "Vs inflation: shown",
    ]
    return [
        ["Generated at", payload.get("generated_at")],
        ["Person", payload.get("person_name") or person],
        ["Window", export_period_label(period)],
        ["Broker", broker],
        ["Berkshire mode", berkshire_mode],
        ["Proxy mode", proxy_mode],
        ["Live filter", "Live only" if live_only == "on" else "All assets"],
        ["Vs MSCI World", "Shown"],
        ["Vs inflation", "Shown"],
        ["Toggles", " | ".join(toggles)],
        ["Date range", f"{date_range.get('start') or '-'} to {date_range.get('end') or '-'}"],
        ["Market value EUR", totals.get("market_value")],
        ["Return %", totals.get("return_pct")],
        ["Unrealized P/L EUR", totals.get("unrealized_pl")],
        ["Realized P/L EUR", totals.get("realized_pl")],
        ["Net contributions EUR", totals.get("net_contributions")],
        ["Priced assets", totals.get("priced_assets")],
        ["Unpriced assets", totals.get("unpriced_assets")],
    ]


def export_period_series(series: list[dict[str, Any]], period: str) -> list[dict[str, Any]]:
    if not series:
        return []
    period = (period or "all").lower()
    if period == "all":
        return series
    end = date.fromisoformat(series[-1]["date"])
    if period == "ytd":
        start = date(end.year, 1, 1)
    elif period == "1w":
        start = end - timedelta(days=7)
    elif period == "1m":
        start = end - timedelta(days=30)
    elif period == "1y":
        start = end.replace(year=end.year - 1)
    elif period == "since24":
        start = date(2024, 1, 11)
    else:
        start = date.min
    return [row for row in series if date.fromisoformat(row["date"]) >= start]


def export_period_rows_by_date(rows: list[dict[str, Any]], period: str) -> list[dict[str, Any]]:
    if not rows:
        return []
    period = (period or "all").lower()
    if period == "all":
        return rows
    parsed_dates: list[date] = []
    for row in rows:
        try:
            parsed_dates.append(date.fromisoformat(str(row.get("date") or "")))
        except ValueError:
            continue
    if not parsed_dates:
        return []
    end = max(parsed_dates)
    if period == "ytd":
        start = date(end.year, 1, 1)
    elif period == "1w":
        start = end - timedelta(days=7)
    elif period == "1m":
        start = end - timedelta(days=30)
    elif period == "1y":
        start = end.replace(year=end.year - 1)
    elif period == "since24":
        start = date(2024, 1, 11)
    else:
        start = date.min
    output: list[dict[str, Any]] = []
    for row in rows:
        try:
            row_date = date.fromisoformat(str(row.get("date") or ""))
        except ValueError:
            continue
        if row_date >= start:
            output.append(row)
    return output


def expense_category_export_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    total_outflow = ZERO
    for row in rows:
        amount = Decimal(str(row.get("amount_eur") or 0))
        category = normalize_expense_category(str(row.get("category") or ""))
        bucket = grouped.setdefault(category, {"category": category, "amount_eur": ZERO, "count": 0})
        bucket["amount_eur"] += amount
        bucket["count"] += 1
        if expense_row_kind(row) not in {"income", "credits"}:
            total_outflow += amount
    output = []
    for values in grouped.values():
        amount = values["amount_eur"]
        is_non_outflow = values["category"] in {"Income", "Credits"}
        output.append(
            {
                "category": values["category"],
                "amount_eur": money(amount),
                "count": values["count"],
                "share_pct": float((amount / total_outflow * Decimal("100")).quantize(Decimal("0.01"))) if total_outflow and not is_non_outflow else None,
            }
        )
    return sorted(output, key=lambda item: float(item.get("amount_eur") or 0), reverse=True)


def normalized_return_delta(first: dict[str, Any], last: dict[str, Any], key: str) -> float:
    first_val = float(first.get(key) or 0)
    last_val = float(last.get(key) or 0)
    base = 1 + first_val / 100
    if abs(base) <= 1e-6:
        return 0.0
    return round(((1 + last_val / 100) / base - 1) * 100, 2)


def benchmark_export_rows(payload: dict[str, Any], period: str) -> list[dict[str, Any]]:
    series = export_period_series(payload.get("valuation_series", []) or [], period)
    if not series:
        return []
    first = series[0]
    last = series[-1]
    portfolio = normalized_return_delta(first, last, "return_pct")
    msci = normalized_return_delta(first, last, "msci_return_pct")
    xeon = normalized_return_delta(first, last, "xeon_return_pct")
    inflation = normalized_return_delta(first, last, "inflation_return_pct")
    return [
        {"metric": "Portfolio return", "value_pct": portfolio, "comparison": ""},
        {"metric": "MSCI World return", "value_pct": msci, "comparison": ""},
        {"metric": "XEON return", "value_pct": xeon, "comparison": ""},
        {"metric": "Eurozone inflation", "value_pct": inflation, "comparison": ""},
        {"metric": "Portfolio vs MSCI World", "value_pct": round(portfolio - msci, 2), "comparison": "Outperformance" if portfolio >= msci else "Underperformance"},
        {"metric": "Portfolio vs XEON", "value_pct": round(portfolio - xeon, 2), "comparison": "Outperformance" if portfolio >= xeon else "Underperformance"},
        {"metric": "Portfolio vs inflation", "value_pct": round(portfolio - inflation, 2), "comparison": "Real gain" if portfolio >= inflation else "Real loss"},
    ]


def build_export_tables(payload: dict[str, Any], context_rows: list[list[Any]], period: str) -> dict[str, list[list[Any]]]:
    distribution = payload.get("distribution", {})
    dividends = payload.get("dividends", {})
    contributions = payload.get("net_contributions", {})
    frictions = payload.get("frictions", {})
    expenses = payload.get("expenses", {})
    expense_rows = export_period_rows_by_date(expenses.get("rows", []) or [], period)
    expense_credit_rows = [row for row in expense_rows if expense_row_kind(row) == "credits"]
    return {
        "Overview": [["Field", "Value"], *context_rows],
        "Benchmark Comparison": export_rows_from_dicts(benchmark_export_rows(payload, period), [
            ("metric", "Metric"),
            ("value_pct", "Value %"),
            ("comparison", "Comparison"),
        ]),
        "Holdings": export_rows_from_dicts(sorted_by_market_value(payload.get("positions", [])), [
            ("asset", "Asset"),
            ("asset_type", "Type"),
            ("isin", "ISIN"),
            ("symbol", "Symbol"),
            ("quantity", "Quantity"),
            ("price", "Price"),
            ("price_currency", "Currency"),
            ("market_value_eur", "Market value EUR"),
            ("cost_basis_eur", "Cost EUR"),
            ("display_pl_eur", "P/L EUR"),
            ("display_pl_pct", "P/L %"),
            ("pricing_status", "Status"),
        ]),
        "Distribution": export_rows_from_dicts(distribution.get("underlying", []), [
            ("holding", "Asset"),
            ("holding_ticker", "Ticker"),
            ("market_value_eur", "Value EUR"),
            ("weight_pct", "Weight %"),
        ]),
        "Sectors": export_rows_from_dicts(distribution.get("sectors", []), [
            ("sector", "Sector"),
            ("market_value_eur", "Value EUR"),
            ("weight_pct", "Weight %"),
        ]),
        "Geographies": export_rows_from_dicts(distribution.get("geographies", []), [
            ("geo", "Geo area"),
            ("market_value_eur", "Value EUR"),
            ("weight_pct", "Weight %"),
        ]),
        "Asset Classes": export_rows_from_dicts(distribution.get("asset_classes", []), [
            ("asset_class", "Asset class"),
            ("market_value_eur", "Value EUR"),
            ("weight_pct", "Weight %"),
        ]),
        "Missing Composition": export_rows_from_dicts(distribution.get("missing", []), [
            ("asset", "Asset"),
            ("isin", "ISIN"),
            ("market_value_eur", "Value EUR"),
            ("status", "Status"),
        ]),
        "Dividends Yearly": export_rows_from_dicts(dividend_yearly_export_rows(dividends), [
            ("year", "Year"),
            ("count", "Payments"),
            ("amount_eur", "Net EUR"),
            ("tax_eur", "Tax EUR"),
            ("gross_eur", "Gross EUR"),
        ]),
        "Dividends Payments": export_rows_from_dicts(dividends.get("rows", []), [
            ("date", "Date"),
            ("broker", "Broker"),
            ("asset", "Asset"),
            ("isin", "ISIN"),
            ("amount_eur", "Net EUR"),
            ("tax_eur", "Tax EUR"),
            ("gross_eur", "Gross EUR"),
        ]),
        "Contributions": export_rows_from_dicts(contributions.get("by_date", []), [
            ("date", "Date"),
            ("buys_eur", "Buys EUR"),
            ("sells_eur", "Sells EUR"),
            ("net_eur", "Net EUR"),
        ]),
        "Frictions": export_rows_from_dicts(frictions.get("rows", []), [
            ("date", "Date"),
            ("broker", "Broker"),
            ("type", "Type"),
            ("description", "Description"),
            ("amount_eur", "Amount EUR"),
        ]),
        "Expense Categories": export_rows_from_dicts(expense_category_export_rows(expense_rows), [
            ("category", "Category"),
            ("count", "Rows"),
            ("amount_eur", "Amount EUR"),
            ("share_pct", "Outflow share %"),
        ]),
        "Expense Rows": export_rows_from_dicts(expense_rows, [
            ("date", "Date"),
            ("source_label", "Source"),
            ("merchant", "Merchant"),
            ("description", "Description"),
            ("flow_kind", "Flow kind"),
            ("category", "Category"),
            ("subcategory", "Subcategory"),
            ("amount_eur", "Amount EUR"),
            ("currency", "Currency"),
            ("native_amount", "Native amount"),
            ("confidence", "Confidence"),
        ]),
        "Expense Credits": export_rows_from_dicts(expense_credit_rows, [
            ("date", "Date"),
            ("source_label", "Source"),
            ("merchant", "Merchant"),
            ("description", "Description"),
            ("category", "Category"),
            ("subcategory", "Subcategory"),
            ("amount_eur", "Amount EUR"),
            ("currency", "Currency"),
            ("native_amount", "Native amount"),
        ]),
    }


def build_xlsx_export(tables: dict[str, list[list[Any]]]) -> io.BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in tables.items():
        add_xlsx_sheet(workbook, title, rows)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_pdf_export(tables: dict[str, list[list[Any]]], title: str) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    body = ParagraphStyle("ExportBody", parent=styles["BodyText"], fontSize=7, leading=8.5)
    header = ParagraphStyle("ExportHeader", parent=styles["BodyText"], fontSize=7, leading=8.5, textColor=colors.white, fontName="Helvetica-Bold")
    positive = ParagraphStyle("ExportPositive", parent=body, textColor=colors.HexColor("#047857"), fontName="Helvetica-Bold")
    negative = ParagraphStyle("ExportNegative", parent=body, textColor=colors.HexColor("#B91C1C"), fontName="Helvetica-Bold")
    small = ParagraphStyle("ExportSmall", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.HexColor("#4B5563"))
    card_label = ParagraphStyle("ExportCardLabel", parent=styles["BodyText"], fontSize=7, leading=8.5, textColor=colors.HexColor("#6B7280"))
    card_value = ParagraphStyle("ExportCardValue", parent=styles["BodyText"], fontSize=13, leading=15, textColor=colors.HexColor("#111827"), fontName="Helvetica-Bold")

    def text(value: Any) -> str:
        raw = "" if value is None else str(value)
        return raw.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def overview_dict() -> dict[str, Any]:
        rows = tables.get("Overview", [])
        return {str(row[0]): row[1] for row in rows[1:] if len(row) >= 2}

    def is_positive_negative_column(label: str) -> bool:
        label_l = label.lower()
        return any(token in label_l for token in ["p/l", "return", "net eur", "amount eur"])

    def numeric_value(value: Any) -> float | None:
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = re.sub(r"[^0-9.\-]", "", value)
            if cleaned and cleaned not in {"-", ".", "-."}:
                try:
                    return float(cleaned)
                except ValueError:
                    return None
        return None

    def table_for(section: str, rows: list[list[Any]]) -> Table | None:
        if not rows:
            return None
        headers = [str(item) for item in rows[0]]
        data_rows = rows[1:]
        if not data_rows:
            data_rows = [["No data available for this selection.", *["" for _ in headers[1:]]]]
        display_rows = [headers, *data_rows[:30]]
        if section in {"Holdings", "Distribution"} and display_rows and headers and headers[0] == "Asset":
            display_rows = [display_rows[0], *[[asset_cell_html(row), *row[1:]] for row in display_rows[1:]]]
        cell_rows = []
        for ridx, row in enumerate(display_rows):
            cells = []
            for cidx, cell in enumerate(row):
                style = header if ridx == 0 else body
                if ridx > 0 and is_positive_negative_column(headers[cidx]):
                    value = numeric_value(cell)
                    if value is not None and value > 0:
                        style = positive
                    elif value is not None and value < 0:
                        style = negative
                rendered = str(cell) if ridx > 0 and section in {"Holdings", "Distribution"} and cidx == 0 else text(cell)
                cells.append(Paragraph(rendered, style))
            cell_rows.append(cells)
        table = Table(cell_rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        for cidx, label in enumerate(headers):
            if is_positive_negative_column(label):
                for ridx, row in enumerate(display_rows[1:], start=1):
                    value = numeric_value(row[cidx]) if cidx < len(row) else None
                    if value is None or value == 0:
                        continue
                    table.setStyle(TableStyle([("TEXTCOLOR", (cidx, ridx), (cidx, ridx), colors.HexColor("#047857" if value > 0 else "#B91C1C"))]))
        return table

    def asset_cell_html(row: list[Any]) -> str:
        ticker = ""
        if len(row) > 3 and str(row[3] or "").strip():
            ticker = str(row[3]).strip()
        elif len(row) > 1 and str(row[1] or "").strip():
            ticker = str(row[1]).strip()
        source = ticker or (str(row[0]) if row else "")
        letters = re.sub(r"[^A-Za-z0-9]", "", source.upper())[:3]
        badge = letters or "ETF"
        asset = text(row[0] if row else "")
        return f'<font color="#2563EB"><b>[{text(badge)}]</b></font> {asset}'

    overview = overview_dict()
    card_fields = [
        ("Market value EUR", overview.get("Market value EUR")),
        ("Return %", overview.get("Return %")),
        ("Unrealized P/L EUR", overview.get("Unrealized P/L EUR")),
        ("Net contributions EUR", overview.get("Net contributions EUR")),
    ]
    story = [Paragraph(title, styles["Title"]), Spacer(1, 8)]
    story.append(Paragraph(f"Generated at {text(overview.get('Generated at', '-'))}", small))
    story.append(Spacer(1, 12))
    cards = Table([[
        [Paragraph(text(label), card_label), Paragraph(text(value), card_value)]
        for label, value in card_fields
    ]], colWidths=[doc.width / 4 - 8] * 4)
    cards.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F3F4F6")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(cards)
    story.append(Spacer(1, 14))
    toggles = [
        ["Person", overview.get("Person")],
        ["Window", overview.get("Window")],
        ["Broker", overview.get("Broker")],
        ["Berkshire", overview.get("Berkshire mode")],
        ["Composition", overview.get("Proxy mode")],
        ["Assets", overview.get("Live filter")],
        ["Vs MSCI World", overview.get("Vs MSCI World")],
        ["Vs inflation", overview.get("Vs inflation")],
    ]
    toggle_table = Table([[Paragraph(text(k), header), Paragraph(text(v), body)] for k, v in toggles], colWidths=[90, 220])
    toggle_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#EEF2FF")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(Paragraph("Active dashboard toggles", styles["Heading2"]))
    story.append(toggle_table)
    story.append(Spacer(1, 12))
    overview_table = table_for("Overview", tables.get("Overview", []))
    if overview_table:
        story.append(Paragraph("Report overview", styles["Heading2"]))
        story.append(overview_table)

    for section, rows in tables.items():
        if section == "Overview":
            continue
        story.append(PageBreak())
        story.append(Paragraph(section, styles["Heading1"]))
        story.append(Spacer(1, 8))
        table = table_for(section, rows)
        if table:
            story.append(table)
        if len(rows) > 31:
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Showing first 30 rows of {len(rows) - 1}. Use XLSX for the full table.", styles["Italic"]))
    doc.build(story)
    output.seek(0)
    return output


def export_filename(person: str, period: str, fmt: str) -> str:
    safe_person = re.sub(r"[^A-Za-z0-9_-]+", "-", person or "portfolio").strip("-").lower()
    safe_period = re.sub(r"[^A-Za-z0-9_-]+", "-", period or "all").strip("-").lower()
    return f"portfolio-{safe_person}-{safe_period}-{date.today().isoformat()}.{fmt}"


# ─── Local statement imports and normalized movement ledger ───
_movement_store: MovementStore | None = None
_legacy_import_lock = threading.Lock()
_legacy_imported_portfolios: set[str] = set()
_legacy_import_errors: dict[str, list[str]] = {}


def get_movement_store() -> MovementStore:
    global _movement_store
    if _movement_store is None:
        _movement_store = MovementStore(MOVEMENT_DATABASE, default_portfolio_id=PRIMARY_PORTFOLIO_ID)
    return _movement_store


def configured_portfolio_id(value: str | None) -> str:
    portfolio_id = (value or PRIMARY_PORTFOLIO_ID).strip().lower()
    configured = {PRIMARY_PORTFOLIO_ID, *(key.lower() for key in SETTINGS.portfolios)}
    if portfolio_id not in configured:
        raise ValueError(f"Unknown portfolio: {value}")
    return portfolio_id


def normalize_trade_movement(trade: Trade) -> dict[str, Any]:
    return {
        "occurred_on": trade.date,
        "event_type": "trade",
        "broker": trade.broker,
        "asset": trade.asset,
        "isin": trade.isin,
        "description": trade.action,
        "currency": trade.cash_currency or trade.currency_hint,
        "amount": trade.grand_total,
        "quantity": trade.quantity_diff,
        "price": trade.price,
        "fees": trade.fees,
        "tax": trade.tax,
        "metadata": {"action": trade.action, "source": trade.source},
    }


def normalize_dividend_movement(dividend: Dividend) -> dict[str, Any]:
    return {
        "occurred_on": dividend.date,
        "event_type": "dividend",
        "broker": dividend.broker,
        "asset": dividend.asset,
        "isin": dividend.isin,
        "description": "Dividend",
        "currency": "EUR",
        "amount": dividend.amount_eur,
        "tax": dividend.tax_eur,
    }


def normalize_friction_movement(event: FrictionEvent) -> dict[str, Any]:
    return {
        "occurred_on": event.date,
        "event_type": event.event_type,
        "broker": event.broker,
        "description": event.description,
        "currency": "EUR",
        "amount": event.amount_eur,
    }


def normalize_expense_movement(event: dict[str, Any]) -> dict[str, Any]:
    return {
        "occurred_on": event.get("date"),
        "event_type": event.get("flow_kind") or "cash_movement",
        "account": event.get("source"),
        "description": event.get("description") or event.get("merchant"),
        "currency": event.get("currency") or "EUR",
        "amount": event.get("native_amount") or event.get("amount_eur"),
        "metadata": {
            "amount_eur": event.get("amount_eur"),
            "category": event.get("category"),
            "subcategory": event.get("subcategory"),
            "merchant": event.get("merchant"),
            "confidence": event.get("confidence"),
            "source_category": event.get("source_category"),
            "value_date": event.get("value_date"),
            "source_state": event.get("source_state"),
        },
    }


def normalize_interest_movement(event: dict[str, Any]) -> dict[str, Any]:
    return {
        "occurred_on": event.get("date"),
        "event_type": "interest",
        "account": event.get("broker"),
        "description": event.get("description") or "Cash Interest",
        "currency": "EUR",
        "amount": event.get("net_eur"),
        "tax": event.get("tax_eur"),
        "metadata": {"gross_eur": event.get("gross_eur")},
    }


def normalize_cash_flow_movement(event: dict[str, Any], account: str) -> dict[str, Any]:
    return {
        "occurred_on": event.get("datetime") or event.get("date"),
        "event_type": "cash_flow",
        "account": account,
        "description": event.get("description") or "Cash movement",
        "currency": event.get("currency") or "EUR",
        "amount": event.get("cash_change"),
        "metadata": {"contribution_change": event.get("contrib_change") or ZERO},
    }


def read_trade_republic_interests_from_path(path: Path) -> list[dict[str, Any]]:
    interests: list[dict[str, Any]] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if row.get("type") != "INTEREST_PAYMENT" or not row.get("date"):
                continue
            amount = parse_decimal(row.get("amount"))
            tax = abs(parse_decimal(row.get("tax") or "0"))
            interests.append(
                {
                    "broker": "Trade Republic",
                    "date": datetime.strptime(row["date"], "%Y-%m-%d").date(),
                    "net_eur": amount - tax,
                    "tax_eur": tax,
                    "gross_eur": amount,
                    "description": (row.get("description") or "Cash Interest").strip(),
                }
            )
    return interests


def read_bbva_interests_from_path(path: Path) -> list[dict[str, Any]]:
    interests: list[dict[str, Any]] = []
    temp_xlsx = Path(tempfile.gettempdir()) / f"bbva-interest-{os.getpid()}-{threading.get_ident()}.xlsx"
    try:
        shutil.copy(path, temp_xlsx)
        workbook = load_workbook(temp_xlsx, read_only=True)
        try:
            for row in workbook.active.iter_rows(values_only=True):
                if len(row) < 7:
                    continue
                op_date_raw, description, amount_raw = row[2], row[3], row[6]
                if not description or "INTERESSI" not in str(description).upper() or not amount_raw:
                    continue
                try:
                    event_date = datetime.strptime(str(op_date_raw).strip(), "%d/%m/%Y").date()
                    net = Decimal(str(amount_raw).replace(" EUR", "").replace(",", ".").strip())
                except (ValueError, TypeError):
                    continue
                gross = net / BBVA_INTEREST_NET_RATE if net > ZERO else ZERO
                interests.append(
                    {
                        "broker": "BBVA",
                        "date": event_date,
                        "net_eur": net,
                        "tax_eur": gross - net,
                        "gross_eur": gross,
                        "description": str(description).strip(),
                    }
                )
        finally:
            workbook.close()
    finally:
        temp_xlsx.unlink(missing_ok=True)
    return interests


def read_trade_republic_cash_flows(path: Path) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    contribution_types = {
        "CUSTOMER_INBOUND",
        "CUSTOMER_INPAYMENT",
        "TRANSFER_INSTANT_INBOUND",
        "VIBAN_TRANSFER_INBOUND",
        "CUSTOMER_OUTBOUND_REQUEST",
        "TRANSFER_INSTANT_OUTBOUND",
        "CARD_TRANSACTION",
        "CARD_TRANSACTION_INTERNATIONAL",
        "CARD_ORDERING_FEE",
    }
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if not row.get("date"):
                continue
            cash_change = parse_decimal(row.get("amount")) + parse_decimal(row.get("fee")) + parse_decimal(row.get("tax"))
            if cash_change == ZERO:
                continue
            events.append(
                {
                    "date": datetime.strptime(row["date"], "%Y-%m-%d").date(),
                    "cash_change": cash_change,
                    "contrib_change": cash_change if row.get("type") in contribution_types else ZERO,
                    "currency": row.get("currency") or "EUR",
                    "description": row.get("description") or row.get("type") or "Cash movement",
                }
            )
    return events


def read_bbva_cash_flows(path: Path) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    temp_xlsx = Path(tempfile.gettempdir()) / f"bbva-cash-{os.getpid()}-{threading.get_ident()}.xlsx"
    try:
        shutil.copy(path, temp_xlsx)
        workbook = load_workbook(temp_xlsx, read_only=True)
        try:
            for row in workbook.active.iter_rows(values_only=True):
                if len(row) < 7:
                    continue
                op_date_raw, description, amount_raw = row[2], row[3], row[6]
                if not description or description == "Causale" or not amount_raw:
                    continue
                try:
                    event_date = datetime.strptime(str(op_date_raw).strip(), "%d/%m/%Y").date()
                    amount = Decimal(str(amount_raw).replace(" EUR", "").replace(",", ".").strip())
                except (ValueError, TypeError):
                    continue
                if amount == ZERO:
                    continue
                events.append(
                    {
                        "date": event_date,
                        "cash_change": amount,
                        "contrib_change": ZERO if "INTERESSI" in str(description).upper() else amount,
                        "currency": "EUR",
                        "description": str(description).strip(),
                    }
                )
        finally:
            workbook.close()
    finally:
        temp_xlsx.unlink(missing_ok=True)
    return events


def parse_statement_movements(source: str, path: Path) -> list[dict[str, Any]]:
    trades: list[Trade] = []
    dividends: list[Dividend] = []
    frictions: list[FrictionEvent] = []
    expenses: list[dict[str, Any]] = []
    interests: list[dict[str, Any]] = []
    cash_flows: list[dict[str, Any]] = []
    if source == "trade_republic":
        trades = read_trade_republic_trades(path)
        dividends = read_trade_republic_dividends(path)
        frictions = read_trade_republic_tax_events(path)
        expenses = read_trade_republic_expense_events(path)
        interests = read_trade_republic_interests_from_path(path)
        cash_flows = [normalize_cash_flow_movement(item, "Trade Republic") for item in read_trade_republic_cash_flows(path)]
    elif source == "fineco":
        fineco_kind = fineco_statement_kind(path)
        if fineco_kind == "bank":
            expenses = read_fineco_bank_expense_events(path)
        elif fineco_kind == "securities":
            trades = read_fineco_trades(path)
            dividends = read_fineco_dividends(path)
            frictions = read_fineco_friction_events(path)
    elif source == "interactive_brokers":
        trades = read_interactive_brokers_trades(path)
    elif source == "etoro":
        trades = read_etoro_trades(path)
        dividends = read_etoro_dividends(path)
        frictions = read_etoro_friction_events(path)
    elif source == "revolut":
        expenses = read_revolut_expense_events([path])
        cash_flows = [normalize_cash_flow_movement(item, "Revolut") for item in read_revolut_cash_events([path])]
    elif source == "intesa":
        expenses = read_intesa_expense_events(path)
    elif source == "bbva":
        expenses = read_bbva_expense_events([path])
        interests = read_bbva_interests_from_path(path)
        cash_flows = [normalize_cash_flow_movement(item, "BBVA") for item in read_bbva_cash_flows(path)]
    elif source == "manual":
        trades = read_manual_trades(path)
    else:
        raise ValueError(f"Unsupported source: {source}")

    return [
        *(normalize_trade_movement(item) for item in trades),
        *(normalize_dividend_movement(item) for item in dividends),
        *(normalize_friction_movement(item) for item in frictions),
        *(normalize_expense_movement(item) for item in expenses),
        *(normalize_interest_movement(item) for item in interests),
        *cash_flows,
    ]


def import_statement_path(
    path: Path,
    original_name: str | None = None,
    requested_source: str = "auto",
    portfolio_id: str = PRIMARY_PORTFOLIO_ID,
    archive: bool = True,
) -> dict[str, Any]:
    if not path.is_file():
        raise ValueError("The selected statement file does not exist")
    if path.stat().st_size > SETTINGS.import_max_bytes:
        raise ValueError(f"Statement files must be smaller than {SETTINGS.import_max_bytes // (1024 * 1024)} MB")
    owner = configured_portfolio_id(portfolio_id)
    source = detect_statement_source(path, requested_source)
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    store = get_movement_store()
    existing = store.import_by_hash(digest, owner)
    if existing is not None and existing.parser_version == APP_VERSION:
        return {
            "status": "duplicate",
            "duplicate": True,
            "source": existing.source_kind,
            "source_label": SOURCE_LABELS.get(existing.source_kind, existing.source_kind),
            "portfolio_id": existing.portfolio_id,
            "movements": existing.movement_count,
            "duplicates": existing.duplicate_count,
            "stored_path": existing.stored_path,
        }

    safe_original_name = secure_filename(original_name or path.name) or f"statement{path.suffix.casefold()}"
    movements = parse_statement_movements(source, path)
    if not movements:
        raise ValueError(f"No supported movements were found in this {SOURCE_LABELS[source]} statement")
    if existing is not None:
        inserted, duplicates = store.enrich_import(existing, parser_version=APP_VERSION, movements=movements)
        return {
            "status": "upgraded",
            "duplicate": False,
            "source": existing.source_kind,
            "source_label": SOURCE_LABELS.get(existing.source_kind, existing.source_kind),
            "portfolio_id": existing.portfolio_id,
            "movements": inserted,
            "duplicates": duplicates,
            "stored_path": existing.stored_path,
        }

    archive_root = ROOT_DIR if owner == PRIMARY_PORTFOLIO_ID else ROOT_DIR / "portfolios" / owner
    destination = import_destination(archive_root, source, digest, safe_original_name) if archive else path
    if archive:
        destination.parent.mkdir(parents=True, exist_ok=True)
    if archive and destination.exists() and destination.resolve() != path.resolve():
        raise ValueError(f"The destination already exists: {configured_path_label(destination)}")
    copied = archive and destination.resolve() != path.resolve()
    if copied:
        shutil.copy2(path, destination)
        destination.chmod(0o600)
    try:
        record = store.record_import(
            sha256=digest,
            source_kind=source,
            original_name=safe_original_name,
            stored_path=configured_path_label(destination),
            parser_version=APP_VERSION,
            movements=movements,
            portfolio_id=owner,
        )
    except Exception:
        if copied:
            destination.unlink(missing_ok=True)
        raise
    return {
        "status": "imported",
        "duplicate": False,
        "source": source,
        "source_label": SOURCE_LABELS[source],
        "portfolio_id": owner,
        "movements": record.movement_count,
        "duplicates": record.duplicate_count,
        "stored_path": record.stored_path,
    }


def legacy_statement_candidates(portfolio_id: str) -> list[tuple[Path, str]]:
    """Find files once for migration; normal dashboard reads never use these paths."""

    owner = configured_portfolio_id(portfolio_id)
    candidates: list[tuple[Path, str]] = []
    if owner == PRIMARY_PORTFOLIO_ID:
        broker_paths = [
            (latest_trade_republic_export(), "trade_republic"),
            (latest_fineco_export(), "fineco"),
            (latest_ib_export(), "interactive_brokers"),
            (latest_etoro_export(), "etoro"),
        ]
        candidates.extend((path, source) for path, source in broker_paths if path is not None)
        archived_manual = sorted((ROOT_DIR / "broker_exports" / "manual").glob("*.csv")) + sorted(
            (ROOT_DIR / "broker_exports" / "manual").glob("*.xlsx")
        )
        candidates.extend((path, "manual") for path in archived_manual)
        legacy_manual = ROOT_DIR / SETTINGS.manual_trades_file
        if not any(path for path, _source in broker_paths) and legacy_manual.exists():
            candidates.append((legacy_manual, "manual"))
        revolut_paths = set(ROOT_DIR.glob(REVOLUT_PATTERN)) | set((ROOT_DIR / "cash_exports" / "revolut").glob("*.csv"))
        intesa_paths = set(ROOT_DIR.glob(INTESA_OPERATIONS_PATTERN)) | set((ROOT_DIR / "cash_exports" / "intesa").glob("*.xlsx"))
        candidates.extend((path, "revolut") for path in sorted(revolut_paths))
        candidates.extend((path, "intesa") for path in sorted(intesa_paths))
        candidates.extend((path, "bbva") for path in bbva_statement_files())
    else:
        family_trade_republic = latest_family_trade_republic_export(owner)
        if family_trade_republic:
            candidates.append((family_trade_republic, "trade_republic"))

    archive_root = ROOT_DIR / "portfolios" / owner
    for source, extensions in SOURCE_EXTENSIONS.items():
        category = "cash_exports" if source in {"revolut", "intesa", "bbva"} else "broker_exports"
        for extension in extensions:
            candidates.extend((path, source) for path in (archive_root / category / source).glob(f"*{extension}"))

    unique: dict[Path, tuple[Path, str]] = {}
    for path, source in candidates:
        try:
            unique[path.resolve()] = (path, source)
        except OSError:
            continue
    return sorted(unique.values(), key=lambda item: (item[1], str(item[0])))


def ensure_legacy_statements_imported(portfolio_id: str = PRIMARY_PORTFOLIO_ID) -> None:
    """One-time compatibility bridge from raw-file installs to the SQLite ledger."""

    owner = configured_portfolio_id(portfolio_id)
    if owner in _legacy_imported_portfolios:
        return
    with _legacy_import_lock:
        if owner in _legacy_imported_portfolios:
            return
        errors: list[str] = []
        for path, source in legacy_statement_candidates(owner):
            try:
                import_statement_path(
                    path,
                    original_name=path.name,
                    requested_source=source,
                    portfolio_id=owner,
                    archive=False,
                )
            except Exception as exc:  # A broken historical file must not block other sources.
                errors.append(f"{configured_path_label(path)}: {exc}")
        _legacy_import_errors[owner] = errors
        _legacy_imported_portfolios.add(owner)


def movement_decimal(row: dict[str, Any], field: str) -> Decimal:
    raw = row.get(field)
    return Decimal(str(raw)) if raw not in {None, ""} else ZERO


def movement_date(row: dict[str, Any]) -> date:
    return date.fromisoformat(str(row["occurred_on"])[:10])


def read_ledger_trades(portfolio_id: str = PRIMARY_PORTFOLIO_ID) -> tuple[list[Trade], dict[str, Any]]:
    owner = configured_portfolio_id(portfolio_id)
    ensure_legacy_statements_imported(owner)
    rows = get_movement_store().movements(owner, ("trade",))
    trades: list[Trade] = []
    for row in rows:
        metadata = row.get("metadata") or {}
        quantity_diff = movement_decimal(row, "quantity")
        amount = movement_decimal(row, "amount")
        currency = str(row.get("currency") or "EUR")
        trades.append(
            Trade(
                asset=str(row.get("asset") or "Unknown"),
                isin=str(row.get("isin") or ""),
                broker=str(row.get("account") or SOURCE_LABELS.get(str(row.get("source_kind")), "Personal")),
                action=str(metadata.get("action") or row.get("description") or ("Acquisto" if quantity_diff >= ZERO else "Vendita")),
                currency_hint=currency,
                cash_currency=currency,
                date=movement_date(row),
                price=movement_decimal(row, "price"),
                quantity=abs(quantity_diff),
                quantity_diff=quantity_diff,
                total_spend=amount,
                fees=movement_decimal(row, "fees"),
                tax=movement_decimal(row, "tax"),
                grand_total=amount,
                grand_total_present=bool(row.get("amount")),
                source=str(metadata.get("source") or row.get("source_kind") or "sqlite"),
            )
        )
    ledger = get_movement_store().summary(owner)
    trade_source_ids = {str(row.get("source_kind") or "") for row in rows}
    trade_sources = [item for item in ledger["sources"] if item["source"] in trade_source_ids]
    labels = [SOURCE_LABELS.get(item["source"], item["source"]) for item in trade_sources]
    return trades, {
        "kind": " + ".join(labels) if labels else "No imported statements",
        "relative_path": configured_path_label(MOVEMENT_DATABASE),
        "sources": trade_sources,
    }


def read_ledger_dividends(portfolio_id: str = PRIMARY_PORTFOLIO_ID) -> list[Dividend]:
    owner = configured_portfolio_id(portfolio_id)
    ensure_legacy_statements_imported(owner)
    return [
        Dividend(
            broker=str(row.get("account") or SOURCE_LABELS.get(str(row.get("source_kind")), "")),
            asset=str(row.get("asset") or "Unknown"),
            isin=str(row.get("isin") or ""),
            date=movement_date(row),
            amount_eur=movement_decimal(row, "amount"),
            tax_eur=movement_decimal(row, "tax"),
        )
        for row in get_movement_store().movements(owner, ("dividend",))
    ]


def read_ledger_frictions(portfolio_id: str = PRIMARY_PORTFOLIO_ID) -> list[FrictionEvent]:
    owner = configured_portfolio_id(portfolio_id)
    ensure_legacy_statements_imported(owner)
    return [
        FrictionEvent(
            broker=str(row.get("account") or SOURCE_LABELS.get(str(row.get("source_kind")), "")),
            event_type=str(row.get("event_type") or "cost"),
            date=movement_date(row),
            amount_eur=abs(movement_decimal(row, "amount")),
            description=str(row.get("description") or "Cost or tax"),
        )
        for row in get_movement_store().movements(owner, ("cost", "tax", "dividend_tax"))
    ]


def read_ledger_expenses(portfolio_id: str = PRIMARY_PORTFOLIO_ID) -> list[dict[str, Any]]:
    owner = configured_portfolio_id(portfolio_id)
    ensure_legacy_statements_imported(owner)
    excluded = {"trade", "dividend", "cost", "tax", "dividend_tax", "interest", "cash_flow"}
    events: list[dict[str, Any]] = []
    for row in get_movement_store().movements(owner):
        if row.get("event_type") in excluded:
            continue
        metadata = row.get("metadata") or {}
        events.append(
            {
                "date": movement_date(row),
                "source": str(row.get("account") or row.get("source_kind") or ""),
                "merchant": str(metadata.get("merchant") or row.get("description") or "Unknown"),
                "description": str(row.get("description") or ""),
                "flow_kind": str(row.get("event_type") or "spend"),
                "category": str(metadata.get("category") or "Uncategorized"),
                "subcategory": str(metadata.get("subcategory") or ""),
                "amount_eur": Decimal(str(metadata.get("amount_eur") or abs(movement_decimal(row, "amount")))),
                "currency": str(row.get("currency") or "EUR"),
                "native_amount": movement_decimal(row, "amount"),
                "confidence": metadata.get("confidence"),
                "source_category": metadata.get("source_category"),
            }
        )
    return sorted(events, key=lambda item: (item["date"], item["source"], item["merchant"]))


def read_ledger_interests(portfolio_id: str = PRIMARY_PORTFOLIO_ID) -> list[dict[str, Any]]:
    owner = configured_portfolio_id(portfolio_id)
    ensure_legacy_statements_imported(owner)
    interests: list[dict[str, Any]] = []
    for row in get_movement_store().movements(owner, ("interest",)):
        metadata = row.get("metadata") or {}
        net = movement_decimal(row, "amount")
        tax = movement_decimal(row, "tax")
        interests.append(
            {
                "broker": str(row.get("account") or SOURCE_LABELS.get(str(row.get("source_kind")), "")),
                "date": movement_date(row),
                "net_eur": net,
                "tax_eur": tax,
                "gross_eur": Decimal(str(metadata.get("gross_eur") or net + tax)),
                "description": str(row.get("description") or "Cash Interest"),
            }
        )
    return sorted(interests, key=lambda item: item["date"], reverse=True)


def read_ledger_cash_events(portfolio_id: str, account: str) -> list[dict[str, Any]]:
    owner = configured_portfolio_id(portfolio_id)
    ensure_legacy_statements_imported(owner)
    events: list[dict[str, Any]] = []
    for row in get_movement_store().movements(owner, ("cash_flow",)):
        if str(row.get("account") or "").casefold() != account.casefold():
            continue
        occurred_on = str(row["occurred_on"])
        try:
            event_time = datetime.fromisoformat(occurred_on)
        except ValueError:
            event_time = datetime.combine(movement_date(row), datetime.min.time())
        metadata = row.get("metadata") or {}
        events.append(
            {
                "datetime": event_time,
                "date": event_time.date(),
                "cash_change": movement_decimal(row, "amount"),
                "contrib_change": Decimal(str(metadata.get("contribution_change") or 0)),
                "currency": str(row.get("currency") or "EUR"),
            }
        )
    return events


def import_status_payload(portfolio_id: str = PRIMARY_PORTFOLIO_ID) -> dict[str, Any]:
    owner = configured_portfolio_id(portfolio_id)
    ensure_legacy_statements_imported(owner)
    store = get_movement_store()
    ledger = store.summary(owner)
    has_trade_source = bool(store.movements(owner, ("trade",)))
    return {
        **ledger,
        "ready": has_trade_source,
        "portfolio_id": owner,
        "migration_errors": _legacy_import_errors.get(owner, []),
        "source_dir": "sources",
        "database": configured_path_label(MOVEMENT_DATABASE),
        "supported_sources": [
            {
                "id": key,
                "label": value,
                "format": source_format_label(key),
                "extensions": list(SOURCE_EXTENSIONS[key]),
            }
            for key, value in SOURCE_LABELS.items()
        ],
        "max_upload_mb": SETTINGS.import_max_bytes // (1024 * 1024),
    }


@app.get("/api/imports/status")
def api_import_status():
    try:
        return jsonify(import_status_payload(request.args.get("portfolio_id", PRIMARY_PORTFOLIO_ID)))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/imports/template")
def api_personal_trade_template():
    template_format = (request.args.get("format") or "csv").strip().casefold()
    if template_format == "csv":
        text_output = io.StringIO()
        csv.writer(text_output).writerow(PERSONAL_TRADE_COLUMNS)
        output = io.BytesIO(text_output.getvalue().encode("utf-8-sig"))
        return send_file(output, mimetype="text/csv", as_attachment=True, download_name="personal-trades-template.csv")
    if template_format == "xlsx":
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation

        workbook = Workbook()
        trades_sheet = workbook.active
        trades_sheet.title = "Trades"
        trades_sheet.append(PERSONAL_TRADE_COLUMNS)
        trades_sheet.freeze_panes = "A2"
        trades_sheet.auto_filter.ref = "A1:K1"
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        for cell in trades_sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        widths = (14, 12, 28, 18, 18, 12, 14, 14, 12, 12, 16)
        for index, width in enumerate(widths, start=1):
            trades_sheet.column_dimensions[chr(64 + index)].width = width
        action_validation = DataValidation(type="list", formula1='"BUY,SELL"', allow_blank=False)
        trades_sheet.add_data_validation(action_validation)
        action_validation.add("B2:B10000")

        instructions = workbook.create_sheet("Instructions")
        instructions.append(("Personal trades template", "Enter one transaction per row on the Trades sheet."))
        instructions.append(("Required", "date, action, asset, quantity, price"))
        instructions.append(("Date", "Use YYYY-MM-DD; Excel date cells are also accepted."))
        instructions.append(("Action", "BUY or SELL"))
        instructions.append(("Optional", "isin, broker, currency, fees, tax, total"))
        instructions.append(("Total", "Absolute cash value. If blank, it is derived from price, quantity, fees, and tax."))
        instructions.column_dimensions["A"].width = 20
        instructions.column_dimensions["B"].width = 88
        instructions["A1"].font = Font(bold=True)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="personal-trades-template.xlsx",
        )
    return jsonify({"error": "Template format must be csv or xlsx"}), 400


@app.post("/api/imports")
def api_import_statement():
    uploaded = request.files.get("file")
    if uploaded is None or not uploaded.filename:
        return jsonify({"error": "Choose a CSV, XLS, XLSX, or PDF statement"}), 400
    suffix = Path(uploaded.filename).suffix.casefold()
    if suffix not in {".csv", ".xls", ".xlsx", ".pdf"}:
        return jsonify({"error": "Supported statement formats are CSV, XLS, XLSX, and PDF"}), 400
    requested_source = request.form.get("source", "auto")
    portfolio_id = request.form.get("portfolio_id", PRIMARY_PORTFOLIO_ID)
    temp_directory = SETTINGS.cache_dir / "imports"
    temp_directory.mkdir(parents=True, exist_ok=True)
    descriptor, temp_name = tempfile.mkstemp(prefix="upload-", suffix=suffix, dir=temp_directory)
    os.close(descriptor)
    temporary = Path(temp_name)
    try:
        uploaded.save(temporary)
        result = import_statement_path(
            temporary,
            original_name=uploaded.filename,
            requested_source=requested_source,
            portfolio_id=portfolio_id,
        )
        return jsonify(result), 200
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"Import failed: {exc}"}), 500
    finally:
        temporary.unlink(missing_ok=True)


# ─── Watchlist Feature ───
WATCHLIST_FILE = SETTINGS.data_path("watchlist.json")
WATCHLIST_CACHE_FILE = SETTINGS.cache_path("watchlist.json")

KNOWN_ISINS = {
    "IWQU.L": "IE00BP3QZ601",
    "FUSD.L": "IE00BYXV8M31",
    "TDIV.AS": "NL0011683594",
    "TDIV.L": "NL0011683594",
}

def get_watchlist_tickers() -> list[str]:
    if not WATCHLIST_FILE.exists():
        default_list = ["IWQU.L", "FUSD.L"]
        WATCHLIST_FILE.parent.mkdir(parents=True, exist_ok=True)
        try:
            save_json(WATCHLIST_FILE, default_list)
        except Exception:
            pass
        return default_list
    try:
        with open(WATCHLIST_FILE, "r") as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
    except Exception:
        pass
    return ["IWQU.L", "FUSD.L"]

def fetch_ticker_quote(symbol: str) -> dict[str, Any]:
    try:
        ticker = yf.Ticker(symbol)
        hist = ticker.history(period="2d")
        if hist.empty:
            return {"ticker": symbol, "error": "No historical data found"}
        
        closes = hist["Close"].dropna().tolist()
        if not closes:
            return {"ticker": symbol, "error": "No Close prices available"}
            
        current_price = closes[-1]
        prev_close = closes[-2] if len(closes) > 1 else current_price
        change = current_price - prev_close
        change_pct = (change / prev_close * 100) if prev_close > 0 else 0.0
        
        try:
            info = ticker.fast_info
            currency = normalize_currency_code(fast_info_value(info, "currency") or infer_currency(symbol))
        except Exception:
            currency = infer_currency(symbol)
            
        long_name = symbol
        try:
            long_name = ticker.info.get("longName") or ticker.info.get("shortName") or symbol
        except Exception:
            pass
            
        isin = KNOWN_ISINS.get(symbol.upper())
        just_etf_url = f"https://www.justetf.com/en/etf-profile.html?isin={isin}" if isin else f"https://www.justetf.com/en/find-etf.html?query={symbol.split('.')[0]}"
        
        return {
            "ticker": symbol,
            "name": long_name,
            "price": round(float(current_price), 3),
            "prev_close": round(float(prev_close), 3),
            "change": round(float(change), 3),
            "change_pct": round(float(change_pct), 2),
            "currency": currency,
            "yfinance_url": f"https://finance.yahoo.com/quote/{symbol}",
            "justetf_url": just_etf_url,
        }
    except Exception as e:
        return {"ticker": symbol, "error": str(e)}

def fetch_watchlist_data(refresh: bool = False) -> list[dict[str, Any]]:
    tickers = get_watchlist_tickers()
    now = int(time.time())
    cache = {}
    if WATCHLIST_CACHE_FILE.exists() and not refresh:
        try:
            with open(WATCHLIST_CACHE_FILE, "r") as f:
                cache = json.load(f)
        except Exception:
            pass

    results = []
    cache_dirty = False
    for ticker in tickers:
        cached_item = cache.get(ticker)
        if cached_item and now - cached_item.get("fetched_at", 0) < 900:
            results.append(cached_item["data"])
        else:
            data = fetch_ticker_quote(ticker)
            results.append(data)
            cache[ticker] = {"fetched_at": now, "data": data}
            cache_dirty = True

    if cache_dirty:
        try:
            save_json(WATCHLIST_CACHE_FILE, cache)
        except Exception:
            pass
    return results

@app.get("/api/watchlist")
def api_get_watchlist():
    refresh = request.args.get("refresh", "false").lower() == "true"
    try:
        data = fetch_watchlist_data(refresh=refresh)
        return jsonify({"watchlist": data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.post("/api/watchlist")
def api_modify_watchlist():
    payload = request.get_json() or {}
    ticker = str(payload.get("ticker", "")).strip().upper()
    action = str(payload.get("action", "")).strip().lower()
    if not ticker:
        return jsonify({"error": "Ticker is required"}), 400
    
    tickers = get_watchlist_tickers()
    if action == "add":
        if ticker not in tickers:
            tickers.append(ticker)
    elif action == "remove":
        if ticker in tickers:
            tickers.remove(ticker)
    else:
        return jsonify({"error": "Invalid action, must be 'add' or 'remove'"}), 400
        
    try:
        save_json(WATCHLIST_FILE, tickers)
        if WATCHLIST_CACHE_FILE.exists():
            try:
                with open(WATCHLIST_CACHE_FILE, "r") as f:
                    cache = json.load(f)
                if ticker in cache:
                    del cache[ticker]
                save_json(WATCHLIST_CACHE_FILE, cache)
            except Exception:
                pass
        return jsonify({"status": "success", "tickers": tickers, "watchlist": fetch_watchlist_data(refresh=True)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.get("/api/portfolio")
def api_portfolio():
    refresh = request.args.get("refresh") == "1"
    person = request.args.get("person", PRIMARY_PORTFOLIO_ID)
    berkshire_mode = request.args.get("berkshire", "stock")
    proxy_mode = request.args.get("proxy", DEFAULT_PROXY_MODE)
    broker = request.args.get("broker", "all")
    live_only = request.args.get("live_only", "off")
    try:
        return jsonify(dashboard_payload(refresh=refresh, person=person, berkshire_mode=berkshire_mode, proxy_mode=proxy_mode, broker=broker, live_only=live_only))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.get("/api/news")
def api_news():
    refresh = request.args.get("refresh") == "1"
    person = request.args.get("person", PRIMARY_PORTFOLIO_ID)
    berkshire_mode = request.args.get("berkshire", "stock")
    proxy_mode = request.args.get("proxy", DEFAULT_PROXY_MODE)
    broker = request.args.get("broker", "all")
    live_only = request.args.get("live_only", "off")
    raw_symbols = request.args.get("symbols", "")
    symbols = [item for item in raw_symbols.split(",") if item] if raw_symbols else None
    try:
        return jsonify(
            portfolio_news_payload(
                person=person,
                berkshire_mode=berkshire_mode,
                proxy_mode=proxy_mode,
                broker=broker,
                refresh=refresh,
                live_only=live_only,
                symbols=symbols,
            )
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

@app.get("/api/rankings")
def api_rankings():
    refresh = request.args.get("refresh") == "1"
    berkshire_mode = request.args.get("berkshire", "stock")
    proxy_mode = request.args.get("proxy", DEFAULT_PROXY_MODE)
    broker = request.args.get("broker", "all")
    live_only = request.args.get("live_only", "off")
    try:
        users = [PRIMARY_PORTFOLIO_ID] + list(FAMILY_PORTFOLIOS.keys())
        user_series = {}
        for user in users:
            payload = dashboard_payload(
                refresh=refresh,
                person=user,
                berkshire_mode=berkshire_mode,
                proxy_mode=proxy_mode,
                broker=broker,
                live_only=live_only
            )
            user_series[user] = payload.get("valuation_series", [])
            
        start_dates = {}
        for user, series in user_series.items():
            if series:
                start_dates[user] = series[0]["date"]
                
        common_start_date = None
        if start_dates:
            common_start_date = max(start_dates.values())
            
        current_year_start = f"{datetime.now().year}-01-01"
        
        rankings = []
        for user in users:
            series = user_series[user]
            if not series:
                rankings.append({
                    "person": user,
                    "name": FAMILY_PORTFOLIOS.get(user, {}).get("name", user.capitalize()),
                    "start_date": None,
                    "returns": {
                        "price": {"start": 0.0, "common": 0.0, "ytd": 0.0},
                        "total": {"start": 0.0, "common": 0.0, "ytd": 0.0}
                    }
                })
                continue
                
            p_first = series[0]
            p_last = series[-1]
            
            def find_point_on_or_after(target_date):
                for p in series:
                    if p["date"] >= target_date:
                        return p
                return series[-1]
                
            p_common = find_point_on_or_after(common_start_date) if common_start_date else p_first
            p_ytd = find_point_on_or_after(current_year_start)
            
            def point_number(point, key):
                try:
                    return float(point.get(key) or 0.0)
                except (TypeError, ValueError):
                    return 0.0

            def calc_period_return(p_t, p_0, total=False):
                profit_key = "total_profit" if total else "profit"
                contrib_key = "total_net_contributions" if total else "net_contributions"
                value_key = "total_market_value" if total else "market_value"
                period_profit = point_number(p_t, profit_key) - point_number(p_0, profit_key)
                period_contributions = point_number(p_t, contrib_key) - point_number(p_0, contrib_key)
                start_value = point_number(p_0, value_key)
                capital_at_work = max(0.01, start_value + max(0.0, period_contributions))
                return round(period_profit / capital_at_work * 100.0, 2)
                
            ret_start_price = calc_period_return(p_last, p_first, total=False)
            ret_start_total = calc_period_return(p_last, p_first, total=True)
            
            ret_common_price = calc_period_return(p_last, p_common, total=False)
            ret_common_total = calc_period_return(p_last, p_common, total=True)
            
            ret_ytd_price = calc_period_return(p_last, p_ytd, total=False)
            ret_ytd_total = calc_period_return(p_last, p_ytd, total=True)
            
            rankings.append({
                "person": user,
                "name": FAMILY_PORTFOLIOS.get(user, {}).get("name", user.capitalize()),
                "start_date": p_first["date"],
                "returns": {
                    "price": {
                        "start": ret_start_price,
                        "common": ret_common_price,
                        "ytd": ret_ytd_price
                    },
                    "total": {
                        "start": ret_start_total,
                        "common": ret_common_total,
                        "ytd": ret_ytd_total
                    }
                }
            })
            
        return jsonify({
            "common_start_date": common_start_date,
            "ytd_start_date": current_year_start,
            "rankings": rankings
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.get("/api/export")
def api_export():
    fmt = (request.args.get("format") or "xlsx").strip().lower()
    if fmt not in {"xlsx", "pdf"}:
        return jsonify({"error": "Unsupported export format. Use xlsx or pdf."}), 400
    person = request.args.get("person", PRIMARY_PORTFOLIO_ID)
    berkshire_mode = normalize_berkshire_mode(request.args.get("berkshire", "stock"))
    proxy_mode = normalize_proxy_mode(request.args.get("proxy", DEFAULT_PROXY_MODE))
    broker = request.args.get("broker", "all")
    live_only = request.args.get("live_only", "off")
    period = (request.args.get("period") or "all").lower()
    try:
        payload = dashboard_payload(
            refresh=False,
            person=person,
            berkshire_mode=berkshire_mode,
            proxy_mode=proxy_mode,
            broker=broker,
            live_only=live_only,
        )
        context_rows = dashboard_export_context(payload, person, berkshire_mode, proxy_mode, broker, live_only, period)
        tables = build_export_tables(payload, context_rows, period)
        filename = export_filename(person, period, fmt)
        if fmt == "xlsx":
            output = build_xlsx_export(tables)
            return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)
        output = build_pdf_export(tables, f"Portfolio Report - {payload.get('person_name') or person} - {export_period_label(period)}")
        return send_file(output, mimetype="application/pdf", as_attachment=True, download_name=filename)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.get("/")
def index():
    profiles = [(PRIMARY_PORTFOLIO_ID, PRIMARY_PORTFOLIO_NAME)] + [
        (portfolio_id, profile.display_name)
        for portfolio_id, profile in SETTINGS.portfolios.items()
        if portfolio_id.lower() != PRIMARY_PORTFOLIO_ID
    ]
    icon = '<svg viewBox="0 0 24 24"><path d="M20 21a8 8 0 0 0-16 0"/><circle cx="12" cy="7" r="4"/></svg>'
    buttons = []
    for index, (portfolio_id, display_name) in enumerate(profiles):
        active = ' class="active"' if index == 0 else ""
        safe_id = html.escape(portfolio_id, quote=True)
        safe_name = html.escape(display_name, quote=True)
        buttons.append(
            f'<button type="button" data-person="{safe_id}" data-label="{safe_name}"{active}>'
            f'<span class="user-avatar" aria-hidden="true">{icon}</span>'
            f'<span class="selector-label">{safe_name}</span></button>'
        )
    since_2024_button = ""
    if SINCE_2024_PORTFOLIO_IDS:
        active = ' class="active"' if PRIMARY_PORTFOLIO_ID in SINCE_2024_PORTFOLIO_IDS else ""
        since_2024_button = f"""
          <button type="button" data-period="since24"{active} style="--time-fill:.76;--time-width:82px" title="Since January 2024">
            <span class="time-main"><span class="time-icon" aria-hidden="true"><svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 2"/></svg></span><span class="selector-label">'24</span></span>
            <span class="time-scale" aria-hidden="true"><span></span></span>
          </button>"""
    app_config = json.dumps(
        {
            "primaryPortfolioId": PRIMARY_PORTFOLIO_ID,
            "since2024PortfolioIds": sorted(SINCE_2024_PORTFOLIO_IDS),
            "appVersion": DISPLAY_VERSION,
            "defaultProxyMode": DEFAULT_PROXY_MODE,
            "hasMultiplePortfolios": len(profiles) > 1,
            "annualRiskFreeRate": SETTINGS.annual_risk_free_rate,
            "portfolioFeatures": {
                PRIMARY_PORTFOLIO_ID: [],
                **{
                    portfolio_id: sorted(profile.features)
                    for portfolio_id, profile in SETTINGS.portfolios.items()
                },
            },
        }
    ).replace("<", "\\u003c")
    return render_template(
        "index.html",
        portfolio_buttons="".join(buttons),
        show_rankings=len(profiles) > 1,
        since_2024_button=since_2024_button,
        app_version=DISPLAY_VERSION,
        app_config=app_config,
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8050)
