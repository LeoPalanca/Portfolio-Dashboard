"""Fineco XLSX statement inspection and current-account movement parsing."""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

FINECO_BANK_REQUIRED_HEADERS = {
    "data_operazione",
    "data_valuta",
    "entrate",
    "uscite",
    "descrizione",
    "descrizione_completa",
    "stato",
}


def normalize_fineco_header(value: Any) -> str:
    """Normalize Fineco headings while preserving their semantic separators."""

    text = str(value or "").strip().casefold()
    return re.sub(r"[^a-z0-9à-ÿ]+", "_", text).strip("_")


def has_fineco_bank_headers(values: tuple[Any, ...] | list[Any]) -> bool:
    headers = {normalize_fineco_header(value) for value in values if value is not None}
    return FINECO_BANK_REQUIRED_HEADERS.issubset(headers)


def fineco_statement_kind(path: Path) -> str:
    """Return ``securities`` or ``bank`` for a supported Fineco workbook."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if any(name.casefold() == "movimenti dossier titoli" for name in workbook.sheetnames):
            return "securities"
        for sheet in workbook.worksheets[:3]:
            for row in sheet.iter_rows(max_row=30, values_only=True):
                if has_fineco_bank_headers(row):
                    return "bank"
    finally:
        workbook.close()
    return ""


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    if not raw:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:\s+.*)?", raw):
        try:
            return date.fromisoformat(raw[:10])
        except ValueError:
            return None
    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", raw):
        day, month, year = (int(part) for part in raw.split("/"))
        try:
            return date(year, month, day)
        except ValueError:
            return None
    return None


def _decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal(0)
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = str(value).strip().replace("€", "").replace(" ", "")
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw)
    except InvalidOperation:
        return Decimal(0)


def _cell_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def read_fineco_bank_movements(path: Path) -> list[dict[str, Any]]:
    """Read booked Fineco current-account rows from their native XLSX export."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        target = None
        header_row = 0
        headers: list[str] = []
        for sheet in workbook.worksheets[:3]:
            for row_number, row in enumerate(sheet.iter_rows(max_row=30, values_only=True), start=1):
                if has_fineco_bank_headers(row):
                    target = sheet
                    header_row = row_number
                    headers = [normalize_fineco_header(value) for value in row]
                    break
            if target is not None:
                break
        if target is None:
            return []

        movements: list[dict[str, Any]] = []
        for row in target.iter_rows(min_row=header_row + 1, values_only=True):
            record = {
                header: row[index] if index < len(row) else None
                for index, header in enumerate(headers)
                if header
            }
            state = _cell_text(record.get("stato")).casefold()
            if state and state not in {"contabilizzato", "booked", "completed"}:
                continue
            occurred_on = _date_value(record.get("data_operazione")) or _date_value(record.get("data_valuta"))
            if occurred_on is None:
                continue
            incoming = abs(_decimal_value(record.get("entrate")))
            outgoing = abs(_decimal_value(record.get("uscite")))
            amount = incoming - outgoing
            if amount == 0:
                continue
            short_description = _cell_text(record.get("descrizione"))
            full_description = _cell_text(record.get("descrizione_completa")) or short_description
            movements.append(
                {
                    "date": occurred_on,
                    "value_date": _date_value(record.get("data_valuta")),
                    "amount": amount,
                    "description": full_description,
                    "source_category": short_description,
                    "state": state,
                }
            )
        return movements
    finally:
        workbook.close()
