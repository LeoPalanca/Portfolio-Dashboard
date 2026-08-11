"""Detect supported statement formats without relying on their original filenames."""

from __future__ import annotations

import csv
import re
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

ALLOWED_EXTENSIONS = {".csv", ".xls", ".xlsx", ".pdf"}
SOURCE_LABELS = {
    "trade_republic": "Trade Republic",
    "fineco": "Fineco",
    "interactive_brokers": "Interactive Brokers",
    "etoro": "eToro",
    "revolut": "Revolut",
    "intesa": "Intesa Sanpaolo",
    "bbva": "BBVA",
    "manual": "Manual trade spreadsheet",
}
SOURCE_EXTENSIONS = {
    "trade_republic": (".csv",),
    "fineco": (".xlsx",),
    "interactive_brokers": (".pdf",),
    "etoro": (".xlsx",),
    "revolut": (".csv",),
    "intesa": (".xlsx",),
    "bbva": (".xls",),
    "manual": (".csv",),
}


def source_format_label(source: str) -> str:
    return "/".join(extension.removeprefix(".").upper() for extension in SOURCE_EXTENSIONS[source])


def _normalized_headers(path: Path) -> set[str]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open(newline="", encoding=encoding) as handle:
                row = next(csv.reader(handle), [])
            return {str(value).strip().casefold() for value in row if str(value).strip()}
        except (OSError, UnicodeDecodeError):
            continue
    return set()


def detect_statement_source(path: Path, requested_source: str = "auto") -> str:
    """Return the configured source id or raise a user-facing validation error."""

    requested = requested_source.strip().casefold().replace("-", "_").replace(" ", "_")
    suffix = path.suffix.casefold()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("Supported statement formats are CSV, XLS, XLSX, and PDF")
    if requested and requested != "auto":
        if requested not in SOURCE_LABELS:
            raise ValueError(f"Unsupported source: {requested_source}")
        if suffix not in SOURCE_EXTENSIONS[requested]:
            required = source_format_label(requested)
            raise ValueError(f"{SOURCE_LABELS[requested]} imports require {required}; this file is {suffix[1:].upper()}")
        return requested

    if suffix == ".pdf":
        return "interactive_brokers"
    if suffix == ".xls":
        return "bbva"
    if suffix == ".csv":
        headers = _normalized_headers(path)
        if {"type", "category", "shares", "amount"}.issubset(headers):
            return "trade_republic"
        if {"tipo", "data di completamento", "descrizione", "importo", "state"}.issubset(headers):
            return "revolut"
        if len(headers) >= 17:
            return "manual"
        raise ValueError("The CSV headers do not match a supported Trade Republic, Revolut, or manual export")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("The workbook could not be opened") from exc
    try:
        sheets = {name.casefold() for name in workbook.sheetnames}
        if "movimenti dossier titoli" in sheets:
            return "fineco"
        if "attività account" in sheets and "dividendi" in sheets:
            return "etoro"
        if "lista operazione" in sheets:
            return "intesa"
        for sheet in workbook.worksheets[:3]:
            for row in sheet.iter_rows(max_row=30, values_only=True):
                values = {str(value).strip().casefold() for value in row if value is not None}
                if {"data", "operazione", "importo"}.issubset(values):
                    return "intesa"
    finally:
        workbook.close()
    raise ValueError("The workbook sheets do not match a supported Fineco, eToro, or Intesa export")


def import_destination(source_dir: Path, source: str, digest: str, original_name: str) -> Path:
    """Choose a stable raw archive path that also matches legacy discovery patterns."""

    suffix = Path(original_name).suffix.casefold()
    stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    token = re.sub(r"[^a-z0-9]+", "-", source.casefold()).strip("-")
    if source == "revolut":
        filename = f"account-statement-{stamp}-{digest[:10]}{suffix}"
        return source_dir / "cash_exports" / source / filename
    if source == "intesa":
        filename = f"account_operations-{stamp}-{digest[:10]}{suffix}"
        return source_dir / "cash_exports" / source / filename
    if source == "bbva":
        filename = f"bbva-{stamp}-{digest[:10]}{suffix}"
        return source_dir / "cash_exports" / source / filename
    if source == "manual":
        return source_dir / "Spreadsheet - Trades.csv"
    filename = f"{token}-{stamp}-{digest[:10]}{suffix}"
    return source_dir / "broker_exports" / source / filename
